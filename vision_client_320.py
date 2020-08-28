from tkinter import *
from tkinter import ttk, messagebox, filedialog, simpledialog, font, colorchooser
from datetime import datetime,timedelta
from tkcalendar import DateEntry
from threading import _start_new_thread, Thread, Event
from selenium import webdriver
from openpyxl.styles import Alignment, Border, Side, colors, fills, Font
from openpyxl import Workbook, cell
import urllib.request, urllib.error, json, os, io, shutil, time, threading, subprocess, tkinter.scrolledtext as tkst,webbrowser,re
from PIL import Image, ImageTk
from operator import or_
from functools import reduce
import vlc, platform, base64
#Plotting tools#
import matplotlib, matplotlib.pyplot as plt, numpy as np
plt.style.use('ggplot')
matplotlib.use("TkAgg") # for backend
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg,NavigationToolbar2Tk
from matplotlib.backends._backend_tk import ToolTip
from matplotlib.figure import Figure
from matplotlib.widgets import SubplotTool
import pandas as pd, pickle, socket

pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)

app_version = "VISION v3.2b Official - FOX+ Media Monitoring Client"
app_date = 20200122

class Settings: #CHECKED
    """
    Data storage class for user preferences.
    """
    def __init__(self):
        self.pref = set()
        self.default_color = "DeepSkyBlue4"
        self.font_color = "white"
        self.button_color = self.default_color
        self.analyze_mode = False

    def change_fg(self, new):
        self.font_color = new

    def change_bg(self, new):
        self.default_color = new

    def save(self):
        with open("settings.pkl", "wb") as f:
            pickle.dump(self, f)

class AutocompleteEntry(Entry): #CHECKED
    """
    Auto complete entry
    Used in Schedule Checking -> Search entry
    Trace self.var for changes and auto-fill the rest
    """

    def __init__(self, *args, **kwargs):
        Entry.__init__(self, *args, **kwargs)
        self.var = self["textvariable"]
        if self.var == '':
            self.var = self["textvariable"] = StringVar()

        self.var.trace('w', self.changed)
        self.bind("<Right>", self.selection)
        self.bind("<Up>", self.up)
        self.bind("<Down>", self.down)

        self.lb_up = False

    def changed(self, name, index, mode):
        try:
            if not vision.filter_var.get():
                if self.var.get() == '':
                    try:
                        self.lb.destroy()
                    except AttributeError:
                        pass
                    self.lb_up = False
                    Entry.config(self,background="white")
                else:
                    words = self.comparison()
                    if words:
                        Entry.config(self,background="white")
                        if not self.lb_up:
                            self.lb = Listbox(width=50)
                            self.lb.bind("<Double-Button-1>", self.selection)
                            self.lb.bind("<Right>", self.selection)
                            root.bind("<Button-1>",self.delete_widget)
                            self.lb.place(relx=0.309,rely=0.119)
                            self.lb_up = True
                        try:
                            self.lb.delete(0, END)
                        except TclError:
                            return
                        for w in words:
                            self.lb.insert(END, w)
                    else:
                        if self.lb_up:
                            self.lb.destroy()
                            self.lb_up = False
                            Entry.config(self,background="pink")
        except NameError:
            pass

    def selection(self, event):
        if self.lb_up:
            self.var.set(self.lb.get(ACTIVE))
            self.lb.destroy()
            self.lb_up = False
            self.icursor(END)
            vision.filter_data() #execute search button when selected

    def delete_widget(self,event):
        if event.widget != self.lb:
            self.lb.destroy()
            root.unbind("<Button-1>")

    def up(self, event):
        if self.lb_up:
            if self.lb.curselection() == ():
                index = '0'
            else:
                index = self.lb.curselection()[0]
            if index != '0':
                self.lb.selection_clear(first=index)
                index = str(int(index) - 1)
                self.lb.selection_set(first=index)
                self.lb.activate(index)

    def down(self, event):
        if self.lb_up:
            if self.lb.curselection() == ():
                index = '0'
            else:
                index = self.lb.curselection()[0]
            if index != END:
                self.lb.selection_clear(first=index)
                index = str(int(index) + 1)
                self.lb.selection_set(first=index)
                self.lb.activate(index)

    def comparison(self):
        """
        :return: <list> object containing strings of titles of selected channels
        """
        pattern = re.compile('.*' + self.var.get() + '.*',re.IGNORECASE)
        search_channel = reduce(or_, [series_name if vision.series_var.get() else set(),
                                      scc_name if vision.scc_var.get() else set(),
                                      factual_name if vision.factual_var.get() else set(),
                                      movies_name if vision.movies_var.get() else set(),
                                      scm_name if vision.scm_var.get() else set()],
                                      trailer_name if vision.trailer_var.get() else set())
        return [w for w in search_channel if re.match(pattern, w)]

class CustomDateEntry(DateEntry): #CHECKED
    """
    Customized DateEntry object
    Added TODAY button for quick access
    """
    def __init__(self,master=None, **kw):
        DateEntry.__init__(self,master, **kw)
        self.today_button = Button(self._calendar,text="Today",command=self.go_today,
                                   background='darkblue',fg="yellow",relief="flat")
        self.today_button.place(relx=0.495,rely=0)
        self.today_button.bind("<Enter>", self.enter)
        self.today_button.bind("<Leave>", self.leave)
        self.today_button.bind("<ButtonPress>", self.leave)

    def enter(self,e):
        self.today_button.config(relief="groove",bg=settings.default_color)

    def leave(self,e):
        self.today_button.config(relief="flat",bg="darkblue")

    def go_today(self):
        try:
            current = datetime.now().strftime("%m/%d/%Y").lstrip("0").replace(" 0", " ")
        except:
            current = datetime.now().strftime("%m/%d/%Y")
        self.set_date(current)
        self._top_cal.withdraw()

    def set_date(self, date):
        """
        Monkey patching only...
        :param date: string in "%m/%d/%Y" format
        :return: None
        """
        self._set_text(date)

class PlotTool(SubplotTool): #CHECKED
    """
    Just to change the title of configuration page
    """
    def __init__(self, targetfig, toolfig):
        SubplotTool.__init__(self, targetfig, toolfig)
        self.axleft.set_title('Adjust Plot Parameters')

class Navigator(NavigationToolbar2Tk): #CHECKED
    """
    Customized Navigator object
    - Removed mouse_move event.x and event.y display
    - Changed buttons layout
    """
    def mouse_move(self,event):
        pass

    def pan(self, *args):
        if self._active == 'PAN':
            self._active = None
            self.tool_buttons[3]["relief"] = "flat"
        else:
            self._active = 'PAN'
            self.tool_buttons[3]["relief"] = "sunken"
            self.tool_buttons[4]["relief"] = "flat"
        if self._idPress is not None:
            self._idPress = self.canvas.mpl_disconnect(self._idPress)
            self.mode = ''

        if self._idRelease is not None:
            self._idRelease = self.canvas.mpl_disconnect(self._idRelease)
            self.mode = ''

        if self._active:
            self._idPress = self.canvas.mpl_connect(
                'button_press_event', self.press_pan)
            self._idRelease = self.canvas.mpl_connect(
                'button_release_event', self.release_pan)
            self.mode = 'pan/zoom'
            self.canvas.widgetlock(self)
        else:
            self.canvas.widgetlock.release(self)

        for a in self.canvas.figure.get_axes():
            a.set_navigate_mode(self._active)

        self.set_message(self.mode)

    def zoom(self, *args):
        if self._active == 'ZOOM':
            self._active = None
            self.tool_buttons[4]["relief"] = "flat"
        else:
            self._active = 'ZOOM'
            self.tool_buttons[4]["relief"] = "sunken"
            self.tool_buttons[3]["relief"] = "flat"

        if self._idPress is not None:
            self._idPress = self.canvas.mpl_disconnect(self._idPress)
            self.mode = ''

        if self._idRelease is not None:
            self._idRelease = self.canvas.mpl_disconnect(self._idRelease)
            self.mode = ''

        if self._active:
            self._idPress = self.canvas.mpl_connect('button_press_event',
                                                    self.press_zoom)
            self._idRelease = self.canvas.mpl_connect('button_release_event',
                                                      self.release_zoom)
            self.mode = 'zoom rect'
            self.canvas.widgetlock(self)
        else:
            self.canvas.widgetlock.release(self)

        for a in self.canvas.figure.get_axes():
            a.set_navigate_mode(self._active)

        self.set_message(self.mode)

    def configure_subplots(self):
        if self.tool_buttons[5]["relief"] == "flat":
            self.tool_buttons[5].config(relief="sunken")
            toolfig = Figure(figsize=(4,2))
            self.window = Frame(root,highlightbackground="black", highlightcolor="black", highlightthickness=1)
            canvas = type(self.canvas)(toolfig, master=self.window)
            toolfig.subplots_adjust(top=0.85)
            canvas.tool = PlotTool(self.canvas.figure, toolfig)
            canvas.draw()
            canvas.get_tk_widget().pack(side=TOP, fill=BOTH, expand=1)
            self.window.place(relx=0,rely=0.122)
        else:
            self.tool_buttons[5].config(relief="flat")
            self.window.destroy()

    def _Button(self, text, file, command, extension='.gif'):
        im = PhotoImage(master=self, file=file)
        b = Button(master=self, text=text, padx=2,image=im, command=command,bg=vision.default_color,relief="flat")
        b._ntimage = im
        b.pack(side=LEFT)
        self.tool_buttons.append(b)
        return b

    def set_message(self, s):
        pass

    def _init_toolbar(self):
        self.tool_buttons = []
        self.toolbar_icons = ["icons/home.png",
                              "icons/backward.png",
                              "icons/forward.png",
                              None,
                              "icons/pan.png",
                              "icons/zoom.png",
                              "icons/config.png",
                              None,
                              "icons/save.png"]
        xmin, xmax = self.canvas.figure.bbox.intervalx
        height, width = 50, xmax-xmin
        Frame.__init__(self, master=self.window,
                          width=500, height=int(height))
        self.update()
        num = 0
        for text, tooltip_text, image_file, callback in self.toolitems:
            if text is None:
                self._Spacer()
            else:
                try:
                    button = self._Button(text=text, file=self.toolbar_icons[num],
                                          command=getattr(self, callback))
                    if tooltip_text is not None:
                        ToolTip.createToolTip(button, tooltip_text)
                except IndexError:
                    pass
            num+=1
        self.pack(side=BOTTOM, fill=X)

    def destroy(self, *args):
        Frame.destroy(self)

class PlotPage(Frame):
    """
    Statistic Page for Matplotlib plots display
    - Plot number of titles over selected period
    - Can choose from Bar/Pie chart
    - Can choose one genre per chart
    """
    def __init__(self, parent,**kwargs):
        Frame.__init__(self, parent, **kwargs)
        self.canvas = None
        self.ax = None
        self.f = None
        self.text_store = []
        self.chart_type = "bar"
        self.x_store, self.y_store = [], []
        self.bar_store = []
        self.bar_legends = {0: ("SEA", "b"),
                            1: ("TW", "g"),
                            2: ("PH", "r"),
                            3: ("HK", "yellow"),
                            4: ("SG", "orange")}
        self.default_color = settings.default_color
        self.font_color = settings.font_color
        top_frame = Frame(self, bg=self.default_color)
        self.plot_icon = PhotoImage(file="icons/plot.png")
        self.empty = PhotoImage(file="icons/empty.png")
        self.buttonPlot = Button(top_frame, text=" Start \n Plot ", compound="left", image=self.plot_icon,
                                 bg=self.default_color, fg=self.font_color, command=self.plot_graph, relief="flat")
        self.buttonPlot.grid(row=0, column=0, rowspan=2, sticky="ns", padx=5)
        buttonClear = Button(top_frame, text=" Clear \n Plot ", compound="left", image=self.empty,
                             bg=self.default_color, fg=self.font_color, command=self.clear_plot, relief="flat")
        buttonClear.grid(row=0, column=1, rowspan=2, sticky="ns")
        self._start = PhotoImage(file="icons/start_date.png").subsample(2, 2)
        self._end = PhotoImage(file="icons/to_date.png").subsample(2, 2)
        Label(top_frame, image=self._start, bg=self.default_color, fg=self.font_color).grid(row=0, column=2)
        Label(top_frame, image=self._end, bg=self.default_color, fg=self.font_color).grid(row=1, column=2)
        self.from_date = CustomDateEntry(top_frame, background='darkblue', justify="center",
                                         foreground='white', borderwidth=2, columnspan=2)
        self.from_date.grid(row=0, column=3, sticky="ns", pady=1, padx=3)
        self.from_date.set_date(datetime.now().replace(day=1).strftime("%m/%d/%Y"))
        ToolTip.createToolTip(self.from_date, "Choose start date")
        self.to_date = CustomDateEntry(top_frame, background='darkblue', justify="center",
                                       foreground='white', borderwidth=2, columnspan=2)
        self.to_date.grid(row=1, column=3, sticky="ns", pady=1, padx=3, ipady=3)
        self.to_date.set_date(self.last_day_of_month())
        ToolTip.createToolTip(self.to_date, "Choose end date")
        AANTAL = [(0, "Series"), (1, "Movies"), (2, "Factual"), (3, "SCM"), (4, "SCC"), ]
        self.v = IntVar()
        self.v.set(0)
        self.v.trace("w",self.trace_v)
        self.buttons = []
        for num, mode in AANTAL:
            but = Radiobutton(top_frame, bd=2, relief="groove", selectcolor="purple3", width=6,  # padx=5,#pady=2,
                              bg=self.default_color, fg=self.font_color,
                              text=mode, variable=self.v, value=num, indicatoron=0)
            but.grid(row=0, column=num + 7)
            self.buttons.append(but)
        self.all_vars = [IntVar() for _ in range(5)]
        for i in range(3):
            self.all_vars[i].set(1)
        self.all_feeds = ("SEA", "TW", "PH", "HK", "SG")
        self.feed_names = ["sea_start", "tw_start", "ph_start","hk_start","sg_start"]
        self.checks = []
        for num, feed in enumerate(self.all_feeds, 7):
            check = Checkbutton(top_frame, variable=self.all_vars[num - 7], anchor="e",
                                borderwidth=0, relief="solid",
                                text=feed, selectcolor="dark slate gray", bg=self.default_color,
                                foreground=self.font_color, justify="center", command="",
                                activebackground=self.default_color,
                                activeforeground=self.font_color)
            check.grid(row=1, column=num)
            self.checks.append(check)
        top_frame.pack(side=TOP, fill=X)
        self.chart_var = IntVar()
        self.chart_var.set(4)
        self.chart_setting = [(4, PhotoImage(file="icons/bar_chart.png"), "Bar"),
                              (5, PhotoImage(file="icons/pie_chart.png"), "Pie"),
                              (6, PhotoImage(file="icons/line_chart.png"), "Line")]
        for num, image, chart_type in self.chart_setting:
            but = Radiobutton(top_frame, bd=2, selectcolor="purple3", offrelief="flat",  # padx=5,#pady=2,
                              bg=self.default_color, fg=self.font_color, image=image,
                              variable=self.chart_var, value=num, indicatoron=0)
            but.grid(row=0, column=num, rowspan=2)
            ToolTip.createToolTip(but, f"Show in {chart_type} chart")
            if chart_type!="Line":
                but.config(command=self.release_all)
            else:
                but.config(command=self.lock_all)

        self.value = PhotoImage(file="icons/value.png")
        self.value_button = Button(top_frame, image=self.value, command=self.dynamic_show_text, relief="flat",
                                   bg=self.default_color)
        self.category = PhotoImage(file="icons/category.png")
        self.category_button = Button(top_frame, image=self.category, command=self.show_legend, relief="sunken",
                                      bg=self.default_color)
        ToolTip.createToolTip(self.value_button, "Show summary")
        ToolTip.createToolTip(self.category_button, "Show legend")
        self.tool_bar_frame = Frame(top_frame, bg=self.default_color)

    def lock_all(self):
        for i in (*self.checks, *self.buttons):
            i.config(state="disabled")

    def release_all(self):
        for i in (*self.checks, *self.buttons):
            i.config(state="normal")

    def last_day_of_month(self):
        """
        :return: a string representing last day of the current month
        """
        now = datetime.now()
        next_month = now.replace(day=28) + timedelta(days=4)  # this will never fail
        result = next_month - timedelta(days=next_month.day)
        return result.strftime("%m/%d/%Y")

    def trace_v(self,*args):
        """
        Trace method to automatically fit feeds to selected genre
        :param args: items in self.all_vars (a <list> object)
        """
        def set_var(*args):
            for num, state in enumerate(args):
                self.all_vars[num].set(state)
        if self.v.get() == 0:
            set_var(1,1,1,0,0)
        elif self.v.get() == 1:
            set_var(1,0,0,0,0)
        elif self.v.get() == 2:
            set_var(1,1,0,0,0)
        elif self.v.get() == 3:
            set_var(0,1,1,1,1)
        elif self.v.get() == 4:
            set_var(1,1,0,0,0)

    def get_x_date(self,input_list):
        """
        This method will fit the X-axis nicely for arbitrary amount of date selected.
        :param input_list: <list> object of all the date strings selected between self.from_date & self.to_date
        :return: <list> object with dates in interval. Non-showing dates will be replaced by "", i.e. [2019-06-01,"","","",2019-06-05...]
        """
        interval = self.tick_interval(input_list)
        result = [x if num==0 or not num%interval else "" for num,x in enumerate(input_list)]
        return result

    def plot_graph(self):
        """
        Look up self.chart_var and determine which chart type to plot.
        """
        if self.chart_var.get() == 4:
            self.df_plot_bar(vision.df)
        elif self.chart_var.get() == 5:
            self.df_plot_pie(vision.df)
        elif self.chart_var.get() == 6:
            self.df_plot_line(vision.df)

    def dynamic_show_text(self):
        """
        Method to enable dynamic display of chart values. Apply to line chart only.
        """
        if self.text_store:
            if self.value_button["relief"] == "flat":
                for t in self.text_store:
                    t.set_visible(True)
            else:
                for t in self.text_store:
                    t.set_visible(False)
            self.value_button["relief"] = "sunken" if self.value_button["relief"] == "flat" else "flat"
            self.canvas.draw()
        else:
            self.value_button["relief"] = "sunken" if self.value_button["relief"] == "flat" else "flat"

    def plot_3d(self): #disabled
        # FIXME : Broken after switching to Pandas (self.get_data)
        pass

    def show_legend(self):
        """
        Method to enable dynamic display of legend. Apply to all charts.
        """
        if self.category_button["relief"] == "sunken":
            self.legend.set_visible(False)
            self.category_button["relief"] = "flat"
        else:
            self.legend.set_visible(True)
            self.category_button["relief"] = "sunken"
        self.canvas.draw()

    def tick_interval(self,ax):
        """
        Simple method to determine what the ticker interval should be.
        :param ax: <list> object of date strings.
        :return: interval value for get_x_date method
        """
        if len(ax) <= 7: return 1
        elif len(ax) <= 28: return 7
        elif len(ax) <= 60: return 10
        elif len(ax) <= 100: return 15
        elif len(ax) <= 200: return 30
        else: return 60

    def df_plot_line(self, data):
        if self.canvas:
            self.clear_plot()
        genre = ["series", "movies", "factual", "scm", "scc"]
        data = data[(~data.duplicated("ID")) | (data['ID'].isnull())]
        start = pd.Timestamp(self.from_date.get_date())
        end = pd.Timestamp(self.to_date.get_date())
        # start = datetime.strptime(self.from_date.get(),'%m/%d/%Y')
        # end = datetime.strptime(self.to_date.get(),'%m/%d/%Y')
        to_plot = (pd.melt(data.loc[data.genre.isin(genre)],
                           id_vars=["title", "genre"],
                           value_vars=[b for a,b in zip(self.all_vars, self.feed_names) if a.get()],
                           var_name="Feed", value_name="Date"))
        to_plot = to_plot.loc[to_plot.Date.between(start, end)]
        total_amount = len(to_plot)
        ppl = to_plot.groupby("Date")["title"].count()

        if not self.canvas:
            self.f = Figure(figsize=(5, 4), dpi=100)
            self.ax = self.f.add_subplot(111)
            ppl.plot(ax=self.ax)
            self.ax.set_ylabel("No. of media published")
            self.ax.set_xlabel("Date")
            self.ax.set_title(f'Total {total_amount} titles on shelf between'
                              f' {self.from_date.get()} and {self.to_date.get()} for all genres', pad=15.0)

            highest = "Days with highest traffic\n\n" + "".join(
                [f"{day.strftime('%Y/%m/%d')} - {int(count):<3}\n"
                 for day, count in ppl.nlargest(5).iteritems()]) +f"\nAverage: {ppl.mean():.2f}/day"

            self.text_store.append(self.ax.text(0.8, 0.8, highest,
                                   horizontalalignment='left',
                                   verticalalignment='center',
                                   transform=self.ax.transAxes,
                                   bbox=dict(facecolor='white', edgecolor='black', boxstyle='round')))

            self.canvas = FigureCanvasTkAgg(self.f, self)
            self.canvas.get_tk_widget().pack(fill=BOTH, expand=True)
            self.toolbar = Navigator(self.canvas, self.tool_bar_frame)
            self.toolbar.config(background=self.default_color)
            self.tool_bar_frame.grid(row=0, column=12, rowspan=2)
            self.value_button.grid(row=0, column=13, rowspan=2)
            if self.value_button["relief"] == "flat":
                self.value_button.invoke()
            #self.category_button.grid(row=0, column=14, rowspan=2)
            self.toolbar.update()

    def df_plot_bar(self, data):
        if self.canvas:
            self.clear_plot()
        genre = ["series", "movies", "factual", "scm", "scc"]
        data = data[(~data.duplicated("ID")) | (data['ID'].isnull())]

        to_plot = (pd.melt(data.loc[data.genre.eq(genre[self.v.get()])],
                           id_vars=["title"],
                           value_vars=[b for a,b in zip(self.all_vars, self.feed_names) if a.get()],
                           var_name="Feed", value_name="Date"))

        start = pd.Timestamp(self.from_date.get_date())
        end = pd.Timestamp(self.to_date.get_date())
        # start = datetime.strptime(self.from_date.get(),'%m/%d/%Y')
        # end = datetime.strptime(self.to_date.get(),'%m/%d/%Y')
        to_plot = to_plot.loc[to_plot.Date.between(start, end)]

        total_amount = len(to_plot)

        ppl = (to_plot.groupby(["Date", "Feed"])["title"].count()
               .unstack("Feed")
               .reindex(pd.date_range(start, end)))

        if not self.canvas:
            self.f = Figure(figsize=(5, 4), dpi=100)
            self.ax = self.f.add_subplot(111)
            ppl.plot(kind='bar', stacked=True, ax=self.ax, rot=0)
            self.ax.set_xticklabels(self.get_x_date(ppl.index.strftime("%Y-%m-%d")))
            self.ax.set_ylabel("No. of media published")
            self.ax.set_xlabel("Date")
            self.ax.set_title(f'Total {total_amount} titles on shelf between'
                              f' {self.from_date.get()} and {self.to_date.get()}',pad=15.0)

            self.legend = self.ax.legend([f"{feed.replace('_start','').upper():>3} total: {int(count):<3}"
                                          for feed, count in ppl.sum().iteritems()])

            self.canvas = FigureCanvasTkAgg(self.f, self)
            self.canvas.get_tk_widget().pack(fill=BOTH, expand=True)
            self.toolbar = Navigator(self.canvas, self.tool_bar_frame)
            self.toolbar.config(background=self.default_color)
            self.tool_bar_frame.grid(row=0, column=12, rowspan=2)
            #self.value_button.grid(row=0, column=13, rowspan=2)
            self.category_button.grid(row=0, column=14, rowspan=2)
            self.toolbar.update()

    def df_plot_pie(self, data):
        if self.canvas:
            self.clear_plot()
        genre = ["series", "movies", "factual", "scm", "scc"]
        data = data[(~data.duplicated("ID")) | (data['ID'].isnull())]

        to_plot = (pd.melt(data.loc[data.genre.eq(genre[self.v.get()])],
                           id_vars=["title"],
                           value_vars=[b for a,b in zip(self.all_vars, self.feed_names) if a.get()],
                           var_name="Feed", value_name="Date"))
        start = pd.Timestamp(self.from_date.get_date())
        end = pd.Timestamp(self.to_date.get_date())
        # start = datetime.strptime(self.from_date.get(),'%m/%d/%Y')
        # end = datetime.strptime(self.to_date.get(),'%m/%d/%Y')
        to_plot = to_plot.loc[to_plot.Date.between(start, end)]

        total_amount = len(to_plot)

        if not self.canvas:
            self.f = Figure(figsize=(5, 3), dpi=100)
            self.ax = self.f.add_subplot(111)
            ppl = to_plot.groupby("Feed")["title"].count()
            ax = ppl.plot.pie(
                title='Total titles on shelf between 2020-01-01 and 2020-01-31',ax=self.ax,
                autopct='%1.1f%%', startangle=90, labels=None, explode=[0.01 for _ in ppl])
            ax.set_ylabel('')
            self.ax.set_title(f'Total {total_amount} titles on shelf between'
                              f' {self.from_date.get()} and {self.to_date.get()}',pad=15.0)

            self.legend = self.ax.legend([f"{feed.replace('_start','').upper():>3} total: {int(count):<3}"
                                          for feed, count in ppl.iteritems()])
            self.f.set_size_inches(2,2)
            self.canvas = FigureCanvasTkAgg(self.f, self)
            self.canvas.get_tk_widget().pack(fill=BOTH, expand=True)
            self.toolbar = Navigator(self.canvas, self.tool_bar_frame)
            self.toolbar.config(background=self.default_color)
            self.tool_bar_frame.grid(row=0, column=12, rowspan=2)
            #self.value_button.grid(row=0, column=13, rowspan=2)
            self.category_button.grid(row=0, column=14, rowspan=2)
            self.toolbar.update()

    def clear_plot(self):
        if self.canvas:
            self.ax = None
            self.f.clear()
            self.buttonPlot.config(state="normal", relief="flat")
            self.toolbar.destroy()
            self.canvas.get_tk_widget().destroy()
            self.canvas = None
            self.tool_bar_frame.grid_forget()
            self.value_button.grid_forget()
            self.category_button.grid_forget()
            self.text_store = []

def error_log(message):
    """
    Function to log all error raised. Write to error_log.txt on client app folder.
    :param message: error message generated.
    """
    current_time = datetime.now().strftime("%Y-%m-%d %H:%m")
    with open("error_log.txt","a",encoding="utf-8") as f:
        print (str(message))
        f.write("{} - {}\n".format(current_time,str(message)))

class HouseBoxTransform: #CHECKED
    """
    Bring up the transformation window for user to quickly translate between house/box number.
    """
    def __init__(self,master):
        self.query_window = Canvas(master)
        top = Frame(self.query_window,background="DarkOrchid4",highlightbackground="green", highlightcolor="green", highlightthickness=1)
        top.pack(fill=X)
        Label(top,text="House number to Box number translator",bg="DarkOrchid4",fg="white").pack(side=LEFT)
        frame = Frame(self.query_window, highlightbackground="green", highlightcolor="green",highlightthickness=1,bg=vision.default_bg)
        frame.pack(fill=BOTH, side=LEFT)
        input_label = Label(frame,text="House/box no.: ",width=12,bg=vision.default_bg,fg=vision.default_font)
        input_label.grid(row=1,column=0)
        self.input_entry = Entry(frame, width=35,borderwidth=2)
        self.input_entry.grid(row=1,column=1)
        self.input_entry.focus_set()
        self.input_entry.bind("<Button-3>",vision.rClicker)
        CreateToolTip(self.input_entry,"Press ENTER to translate")
        output_label = Label(frame, text="Output: ", width=12,bg=vision.default_bg,fg=vision.default_font)
        output_label.grid(row=2, column=0)
        self.output_entry = Entry(frame, width=35,borderwidth=2)
        self.output_entry.grid(row=2, column=1)
        self.output_entry.bind("<Button-3>", vision.rClicker)
        self.input_entry.bind("<Return>", self.search_df)#(lambda event: number_search()))
        Label(frame,text="Hint: You can use ',' to input multiple house/box numbers.",
                font = "Calibri 8", foreground = "blue",bg=vision.default_bg).grid(row=3,column=0,columnspan=2)

    def search_df(self, event=None):
        if not len(vision.df):
            messagebox.showinfo("Error", "Unable to load database.")
        else:
            entry_list = self.input_entry.get().upper()
            entry_list = entry_list.replace(" ","").split(",")
            self.output_entry.delete(0,END)
            hse, box = vision.df["House Number"].values,vision.df["Box Number"].values
            result = []
            for i in entry_list:
                if i in hse:
                    result.append(vision.df.loc[vision.df["House Number"].eq(i), "Box Number"].iat[0])
                elif i in box:
                    result.append(vision.df.loc[vision.df["Box Number"].eq(i), "House Number"].iat[0])
                else:
                    result.append("N/A")
            self.output_entry.insert(END,",".join(result))

    def show_box(self):
        self.query_window.place(relx=0.001, rely=0) #(relx=0.001, rely=0.038)

    def hide_box(self):
        self.query_window.destroy()

class AboutWin:
    """
    Information and version history for the App.
    """
    def __init__(self,master):
        self.master = master
        window = Toplevel(master,highlightbackground="grey",highlightthickness=2) #relief='solid',borderwidth=2
        window.grab_set()
        window.overrideredirect(True)
        window.geometry("217x200+{}+{}".format(master.winfo_x() + 520, master.winfo_y()+250)) #220/30
        window.iconbitmap("")
        window_label = Label(window, text=" Program by Henry Yik", font="Calibri 11 bold",
                             justify=LEFT)
        window_label.grid(row=0, column=1, columnspan=2,sticky="w")
        modify_label = Label(window, text=" henry.yik@disney.com", font=("Calibri", 8), justify=LEFT)
        modify_label.grid(row=1, column=1, columnspan=2,sticky="w")
        fox_logo = PhotoImage(file="icons/fox.png")
        fox_label = Label(window, image=fox_logo, compound="left", justify="left")
        fox_label.grid(row=0, column=0, rowspan=2)
        fox_label.img = fox_logo
        about_info = tkst.ScrolledText(window, wrap=WORD, width=32, height=6)
        about_info.config(font=("Calibri", 8))
        about_info.grid(row=3, column=0, columnspan=4)
        general = ["Kind:", "Version:", "Copyright:"]
        info = ["Application", app_version.split(" ")[1], "Copyright @2018-20, Henry Yik\nAll Rights Reserved"]
        for i, (text, detail) in enumerate(zip(general,info)):
            Label(window,text=text, font=("Calibri", 8)).grid(row=4 + i, column=0, sticky="ne")
            Label(window,text=detail, font=("Calibri", 8), justify=LEFT).grid(row=4 + i, column=1,
                                                                              sticky="ws", columnspan=4)
        about_info.insert(END,  "v3.2b official - 2020/01/22\n"
                                "- Applied fix 2 for date error\n\n"
                                "v3.2a official - 2020/01/22\n"
                                "- Applied fix for date error\n\n"
                                "v3.2 official - 2020/01/21\n"
                                "- Added color customization\n"
                                "- Removed 3D chart\n"
                                "- Added line chart\n\n"
                                "v3.1 beta 2 - 2020/01/20\n"
                                "- Removed inspector mode\n"
                                "- Added analyaze mode (default on)\n"
                                "- Added Hide Taiwan only titles\n\n"
                                "v3.1 beta 1 - 2020/01/16\n"
                                "- Fixed search by end date\n"
                                "- Added color palette\n\n"
                                "v3.07 build 0115a - 2020/01/15\n"
                                "- Added request for data update\n\n"
                                "v3.06 build 0113a - 2020/01/13\n"
                                "- Remove theme selection\n"
                                "- Relocated show thumbnail to menu\n\n"
                                "v3.05 build 0110a - 2020/01/10\n"
                                "- Added support for trailers\n\n"
                                "v3.04 build 0109a - 2020/01/09\n"
                                "- Fixed plotting page\n\n"
                                "v3.03 build 0108a - 2020/01/08\n"
                                "- Fixed house/box translator\n"
                                "- Fixed Query by ID\n\n"
                                "v3.02 build 0107a - 2020/01/07\n"
                                "- Changed preference selection from version to box no.\n\n"
                                "v3.01 build 0104a - 2020/01/04\n"
                                "- Database now created by loader instead of local\n"
                                "- Fixed exporting result to excel\n\n"
                                "v3.00 build 0103a - 2020/01/03\n"
                                "- Changed backend to Pandas dataframe\n"

                          )
        about_info.bind("<Key>", lambda e: vision.txtEvent(e))
        window.title("About Vision Client")
        window.resizable(False, False)
        window.transient(master)
        window.bind("<Button-1>",lambda e: window.destroy())

class DragManager:
    """
    class created to simulate drag/drug behavior for Toplevel with overrideredict set to True.
    """

    def __init__(self,window):
        self.window = window

    def add_dragable(self, widget):
        widget.bind("<ButtonPress-1>", self.on_start)
        widget.bind("<B1-Motion>", self.on_drag)
        widget.bind("<ButtonRelease-1>", self.on_drop)
        #widget.configure(cursor="arrow")

    def on_start(self, event):
        pass

    def on_drag(self, event):
        pass

    def on_drop(self, event):
        x,y = event.widget.winfo_pointerxy()
        target = event.widget.winfo_containing(x,y)
        height,rest = self.window.winfo_geometry().split("x")
        width,xloc,yloc = rest.split("+")
        self.window.geometry(f"{height}x{width}+{x}+{y}")
        try:
            target.configure(image=event.widget.cget("image"))
        except:
            pass

class MainGui:

    def __init__(self,master):
        self.master = master
        self.style = ttk.Style(self.master)
        self.default_color = settings.default_color #"DeepSkyBlue4"
        self.font_color = settings.font_color
        self.button_color = settings.default_color
        self.tree_bg = "white"
        self.tree_font = "black"
        self.admin_right = False
        self.last_focus = ""
        self.search_by_start = True
        self.current_iid = None
        self.shorten_title = {}
        self.wrap_length = 38
        self.show_image = IntVar(0)
        self.default_bg,self.default_font = "white","black"
        self.cut_icon = PhotoImage(file="icons\cut.png")
        self.copy_icon = PhotoImage(file="icons\copy.png")
        self.paste_icon = PhotoImage(file="icons\paste.png")
        self.inspect_icon = PhotoImage(file="icons\inspect.png")
        self.link_icon = PhotoImage(file="icons\link.png")
        self.loading_icon = PhotoImage(file="icons\loading.png")
        self.image_ref = [self.cut_icon, self.copy_icon, self.paste_icon, self.inspect_icon, self.link_icon,
                          self.loading_icon]
        self.menubar = Menu(master)
        self.nb = ttk.Notebook(master)
        self.create_menu(master)
        self.create_notebook(master)
        self.searching, self.show_pref = False, False
        self.tok,self.uid = "",""
        self._t = 14400
        self.miss_dict,self.filter_dict = {},{}
        self.series_order = ["SOU", "TXMAU", "TXMA"]
        self.loading = None
        self.create_top_frame(self.series_page)
        self.create_top_right()
        self.create_top_left()
        self.create_top_mid()
        self.create_bottom_mid()
        self.create_tree_area()
        self.create_meta_page()
        self.create_lowest()
        VideoSubCheck(self.videosubcheck_page)
        SearchExpiration(self.search_expire)
        self.start_thread(self.initial_start)
        self.restart_but.config(command=sys_check.ask_restart)

    def create_menu(self,master):
        self.submenu = Menu(self.menubar, tearoff=0)
        self.menubar.add_cascade(label="File", menu=self.submenu)
        config_menu = Menu(self.menubar,tearoff=0)
        #theme_menu = Menu(self.menubar,tearoff=0)
        config_menu.add_checkbutton(label="Show thumbnails", onvalue=1, offvalue=0, variable=self.show_image)
        config_menu.add_command(label="Customize colors", command=lambda: ColorChooser(root))
        config_menu.add_command(label="Reset to default", command=self.reset_default)
        # theme_menu.add_command(label="Clam",command=lambda: self.theme_change(0))
        # theme_menu.add_command(label="XP Native", command=lambda: self.theme_change(1))
        self.menubar.add_cascade(label="Configuration", menu=config_menu)
        self.menubar.add_cascade(label="About", menu=Menu, command=lambda: AboutWin(master))
        master.config(menu=self.menubar)

    def reset_default(self):
        settings.default_color = "DeepSkyBlue4"
        settings.font_color = "white"
        try:
            settings.save()
            self.display_message.config(text="Reverting color settings to default after restart.")
            self.display_icon.config(image=self.check_icon)
        except PermissionError:
            message = " Failed to save settings. Please close the file or enable permission."
            self.display_message.config(text=message)
            self.display_icon.config(image=self.error_icon)

    def theme_change(self,num):
        if num == 0:
            self.style.theme_use("clam")
            self.style.configure("Treeview", background="black",
                                 fieldbackground="black", foreground="white")
            self.tree_bg,self.default_bg = "black","black"
            self.tree_font,self.default_font = "white","white"
            self.decolourize()
        elif num == 1:
            self.style.theme_use("xpnative")
            self.style.configure("Treeview", background="white",
                            fieldbackground="white", foreground="black")
            self.tree_bg,self.default_bg = "white","white"
            self.tree_font,self.default_font = "black","black"
            self.decolourize()

    def create_notebook(self,master):
        self.series_page = Frame(self.nb,bg=self.default_color)
        self.metadata_page = Frame(self.nb)
        self.videosubcheck_page = Frame(self.nb)
        self.revise_sub = Frame(self.nb,bg=self.default_color)
        self.search_expire = Frame(self.nb, bg=self.default_color)
        self.plot_page = PlotPage(self.nb, bg="white")
        self.nb.add(self.series_page, text=" Schedule Checking ")
        self.nb.add(self.metadata_page, text=" Metadata Checking ")
        self.nb.add(self.videosubcheck_page, text=" Manifest Checking")
        self.nb.add(self.search_expire, text=" Expired title Checking")
        self.nb.add(self.revise_sub,text=" Akamai Subtitle Purge",state="disabled")
        self.nb.add(self.plot_page, text=" Statistics Page")
        self.nb.pack(expand=1, fill="both")

    def txtEvent(self,event):
        if (event.state==12 and event.keysym=='c'):
            return
        else:
            return "break"

    def rClicker(self,e):
        try:
            def rClick_Copy(e, apnd=0):
                e.widget.event_generate('<Control-c>')

            def rClick_Cut(e):
                e.widget.event_generate('<Control-x>')

            def rClick_Paste(e):
                e.widget.event_generate('<Control-v>')

            e.widget.focus()

            nclst=[
                   ('  Cut',self.cut_icon, lambda e=e: rClick_Cut(e)),
                   ('  Copy',self.copy_icon, lambda e=e: rClick_Copy(e)),
                   ('  Paste',self.paste_icon, lambda e=e: rClick_Paste(e))
                   ]

            rmenu = Menu(None, tearoff=0, takefocus=0)

            for (txt, img, cmd) in nclst:
                rmenu.add_command(label=txt, image=img, command=cmd, compound="left")

            rmenu.tk_popup(e.x_root+40, e.y_root+10,entry="0")

        except TclError as e:
            message = 'Function MainGui.rClicker: '+str(e)
            error_log(message)

        return "break"

    def start_thread(self,func):
        _start_new_thread(func,("Thread",1))

    def mpx_connect_console(self):
        try:
            with open("login_info.json","r") as f:
                login_info = json.load(f)
        except FileNotFoundError:
            login_info = {}
        if len(self.tree.get_children()) ==0 and self.searching is not True and self.nb.index(self.nb.select()) == 0:
            message = " There is nothing to query - Search something."
            self.search_prompt()
            self.display_message.config(text=message)
            self.display_icon.config(image=self.error_icon)
            return
        elif self.tok == "" and login_info.get("autologin"):
            self.uid = Encryption.decode("Infinitywar@2018", login_info.get("username"))
            settings.username = self.uid
            pw = Encryption.decode("Infinitywar@2018", login_info.get("password"))
            settings.password = pw
            self.mpx_button.config(state="disabled")
            self.disable_all()
            def mpx_connect(a,b):
                if self.tok == "":
                    auth_url = 'https://identity.auth.theplatform.com/idm/web/Authentication/signIn?schema=1.0&pretty=true&form=json&username='+self.uid+'&password='+pw
                    self.bottom_frame.config(cursor="watch")
                    try:
                        with urllib.request.urlopen(auth_url) as url:
                            take = json.loads(url.read().decode())
                            self.tok = take["signInResponse"]["token"]
                        if self.tree.get_children():
                            another_add = {str(self.tree.item(child)["values"][15]):{"lib":"","dp":""} for child in self.tree.get_children() if self.tree.item(child)["values"][15]}
                            ### LIBRARY GET ###
                            library_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,:imageApproved,defaultThumbnailUrl,approved&byCustomValue=%7BhouseID%7D%7B" + "%7C".join( #"pl3$imageApproved"
                                self.get_house_id()) + "%7D&token=" + self.tok + "&account=Fox%20Play%20Asia"
                            with urllib.request.urlopen(library_id_url) as url:
                                data = json.loads(url.read().decode())
                                for item in data["entries"]:
                                    if "pl1$houseID" in item:
                                        another_add[str(item["pl1$houseID"])]["lib"] = item["approved"]
                                mpx_image_dict = {str(item['pl1$houseID']): item.get("defaultThumbnailUrl","") if self.show_image.get() else item.get("pl2$imageApproved",item.get("pl3$imageApproved",item.get("pl1$imageApproved",""))) for item in data["entries"] if 'pl1$houseID' in item}
                            ### DP GET ###
                            house_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,:channel_id,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B"+"%7C".join(self.get_house_id())+"%7D&token="+self.tok+"&account=FoxPlus%20Asia%20DP"
                            with urllib.request.urlopen(house_id_url) as url:
                                data = json.loads(url.read().decode())
                                for item in data["entries"]:
                                    if "pl1$houseID" in item:
                                        if self.search_by_start:
                                            another_add[str(item["pl1$houseID"])]["dp"] = datetime.fromtimestamp(item['availableDate'] / 1000).strftime('%Y-%m-%d %H:%M') if item["availableDate"] !=0 else ""
                                        else:
                                            another_add[str(item["pl1$houseID"])]["dp"] = datetime.fromtimestamp(item['expirationDate'] / 1000).strftime('%Y-%m-%d %H:%M') if item["expirationDate"] != 0 else ""
                            for child in self.tree.get_children():
                                try:
                                    if str(self.tree.item(child)["values"][15]) in another_add:
                                        old_value = list(self.tree.item(child)["values"])
                                        old_value[18] = another_add[str(self.tree.item(child)["values"][15])]["lib"]
                                        old_value[19] = another_add[str(self.tree.item(child)["values"][15])]["dp"]
                                        self.tree.item(child,value=old_value)
                                except IndexError as e:
                                    message = "Function MainGUI.mpx_connect_console.mpx_connect: "+str(e)
                                    error_log(message)
                            def populate_thumbnail(a,b): #Thread thumbnail event
                                for child in self.tree.get_children():
                                    h_id = str(self.tree.item(child)["values"][15])
                                    if h_id in mpx_image_dict:
                                        if self.show_image.get():
                                            try:
                                                raw_data = urllib.request.urlopen(mpx_image_dict[h_id]).read()
                                                im = Image.open(io.BytesIO(raw_data)).resize((40, 22))
                                                image = ImageTk.PhotoImage(im)
                                                self.image_ref.append(image)
                                                self.tree.item(child, image=image, text="")
                                            except (urllib.error.HTTPError,urllib.error.URLError,ValueError) as e:
                                                message = "Function MainGUI.mpx_connect_console.mpx_connect.populate_thumbnail: "+str(e)
                                                error_log(message)
                                        else:
                                            if mpx_image_dict.get(h_id):
                                                self.tree.item(child, image="", text="True")
                                            else:
                                                self.tree.item(child, image="", text="False")
                            self.start_thread(populate_thumbnail)
                        self.red_message.config(image=self.green_icon, text="",justify=LEFT) #  Connected to MPX  width=160,
                        self.reset_idle_timer()
                        self.set_idle_timer()
                    except urllib.error.HTTPError:
                        self.red_message.config(image=self.green_icon, text="",justify=LEFT) #  Connected to MPX  width=160,
                        login._quit_login()
                        messagebox.showinfo("Error", "Too many media selected for query.\nPlease reduce the amount of media to search.")
                    except (KeyError,urllib.error.URLError) as e:
                        message = "Function MainGUI.mpx_connect_console.mpx_connect: "+str(e)
                        error_log(message)
                    self.miss_dict = {}
                    self.mpx_button.config(state="normal")
                    self.enable_all()
                    if any(s in self.uid.upper() for s in ("HENRY.YIK","KENNETH.CHAN","ASIC","YUEN.LAU")):
                        self.admin_right = True
                        self.nb.tab(4,state="normal")
                        self.sub_purge.entry.insert(0, self.sub_purge.message)
                    self.bottom_frame.config(cursor="arrow")
            self.start_thread(mpx_connect)
        elif self.tok == "":
            login = LoginScreen()
            login.username.bind("<Return>", (lambda event: self.start_thread(mpx_connect)))
            login.password.bind("<Return>", (lambda event: self.start_thread(mpx_connect)))
            login_dnd = DragManager(login.log_win)
            login_dnd.add_dragable(login.top_frame)
            login_dnd.add_dragable(login.bg_label)
            def mpx_connect(a,b):
                login.username.config(state="disabled")
                login.password.config(state="disabled")
                login.eye_label.config(state="disabled")
                login.log_but.config(relief="sunken",state="disabled")
                self.uid = login.username.get()
                pw = login.password.get()
                if self.tok == "":
                    self.bottom_frame.config(cursor="watch")
                    auth_url = 'https://identity.auth.theplatform.com/idm/web/Authentication/signIn?schema=1.0&pretty=true&form=json&username='+self.uid+'&password='+pw
                    try:
                        with urllib.request.urlopen(auth_url) as url:
                            take = json.loads(url.read().decode())
                            self.tok = take["signInResponse"]["token"]
                            if login.var.get():
                                with open("login_info.json","w") as e:
                                    login.login_info["autologin"] = 1
                                    login.login_info["username"] = Encryption.encode("Infinitywar@2018", login.username.get())
                                    login.login_info["password"] = Encryption.encode("Infinitywar@2018", login.password.get())
                                    json.dump(login.login_info,e)
                            else:
                                with open("login_info.json", "w") as e:
                                    login.login_info = {"autologin": 0, "username": "", "password": ""}
                                    json.dump(login.login_info, e)
                        if self.tree.get_children():
                            another_add = {str(self.tree.item(child)["values"][15]):{"lib":"","dp":""} for child in self.tree.get_children() if self.tree.item(child)["values"][15]}
                            ### LIBRARY GET ###
                            library_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,:imageApproved,defaultThumbnailUrl,approved&byCustomValue=%7BhouseID%7D%7B" + "%7C".join( #"pl3$imageApproved"
                                self.get_house_id()) + "%7D&token=" + self.tok + "&account=Fox%20Play%20Asia"
                            with urllib.request.urlopen(library_id_url) as url:
                                data = json.loads(url.read().decode())
                                for item in data["entries"]:
                                    if "pl1$houseID" in item:
                                        another_add[str(item["pl1$houseID"])]["lib"] = item["approved"]
                                mpx_image_dict = {str(item['pl1$houseID']): item.get("defaultThumbnailUrl","") if self.show_image.get() else item.get("pl2$imageApproved",item.get("pl3$imageApproved",item.get("pl1$imageApproved",""))) for item in data["entries"] if 'pl1$houseID' in item}
                            ### DP GET ###
                            house_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,:channel_id,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B"+"%7C".join(self.get_house_id())+"%7D&token="+self.tok+"&account=FoxPlus%20Asia%20DP"
                            with urllib.request.urlopen(house_id_url) as url:
                                data = json.loads(url.read().decode())
                                for item in data["entries"]:
                                    if "pl1$houseID" in item:
                                        if self.search_by_start:
                                            another_add[str(item["pl1$houseID"])]["dp"] = datetime.fromtimestamp(item['availableDate'] / 1000).strftime('%Y-%m-%d %H:%M') if item["availableDate"] !=0 else ""
                                        else:
                                            another_add[str(item["pl1$houseID"])]["dp"] = datetime.fromtimestamp(item['expirationDate'] / 1000).strftime('%Y-%m-%d %H:%M') if item["expirationDate"] != 0 else ""
                            for child in self.tree.get_children():
                                try:
                                    if str(self.tree.item(child)["values"][15]) in another_add:
                                        old_value = list(self.tree.item(child)["values"])
                                        old_value[18] = another_add[str(self.tree.item(child)["values"][15])]["lib"]
                                        old_value[19] = another_add[str(self.tree.item(child)["values"][15])]["dp"]
                                        self.tree.item(child,value=old_value)
                                except IndexError as e:
                                    message = "Function MainGUI.mpx_connect_console.mpx_connect: "+str(e)
                                    error_log(message)
                            def populate_thumbnail(a,b): #Thread thumbnail event
                                for child in self.tree.get_children():
                                    h_id = str(self.tree.item(child)["values"][15])
                                    if h_id in mpx_image_dict:
                                        if self.show_image.get():
                                            try:
                                                raw_data = urllib.request.urlopen(mpx_image_dict[h_id]).read()
                                                im = Image.open(io.BytesIO(raw_data)).resize((40, 22))
                                                image = ImageTk.PhotoImage(im)
                                                self.image_ref.append(image)
                                                self.tree.item(child, image=image, text="")
                                            except (urllib.error.HTTPError,urllib.error.URLError,ValueError) as e:
                                                message = "Function MainGUI.mpx_connect_console.mpx_connect.populate_thumbnail: "+str(e)
                                                error_log(message)
                                        else:
                                            if mpx_image_dict.get(h_id):
                                                self.tree.item(child, image="", text="True")
                                            else:
                                                self.tree.item(child, image="", text="False")
                            self.start_thread(populate_thumbnail)
                        self.red_message.config(image=self.green_icon, text="",justify=LEFT) #  Connected to MPX  width=160,
                        login._quit_login()
                        self.reset_idle_timer()
                        self.set_idle_timer()
                    except urllib.error.HTTPError:
                        self.red_message.config(image=self.green_icon, text="",justify=LEFT) #  Connected to MPX  width=160,
                        login._quit_login()
                        messagebox.showinfo("Error", "Too many media selected for query.\nPlease reduce the amount of media to search.")
                    except (KeyError,urllib.error.URLError) as e:
                        login.login_message.config(text="Invalid username or password.")
                        login.username.config(state="normal")
                        login.password.config(state="normal")
                        login.log_but.config(relief="groove", state="normal")
                        login.eye_label.config(state="normal")
                        message = "Function Maingui.mpx_connect_console.mpx_connect: "+str(e)
                        error_log(message)
                    self.miss_dict = {}
                    self.bottom_frame.config(cursor="arrow")
                    if any(s in self.uid.upper() for s in ("HENRY.YIK", "KENNETH.CHAN", "ASIC","YUEN.LAU")):
                        self.nb.tab(4, state="normal")
                        self.sub_purge.entry.insert(0, self.sub_purge.message)
            login.log_but.config(command=lambda: self.start_thread(mpx_connect))
        else:
            self.mpx_button.config(state="disabled")
            self.disable_all()
            self.bottom_frame.config(cursor="watch")
            def mpx_connect_with_token(a,b):
                try:
                    another_add = {str(self.tree.item(child)["values"][15]): {"lib": "", "dp": ""} for child in self.tree.get_children() if self.tree.item(child)["values"][15]}
                    library_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,:imageApproved,defaultThumbnailUrl,approved&byCustomValue=%7BhouseID%7D%7B" + "%7C".join(
                        self.get_house_id()) + "%7D&token=" + self.tok + "&account=Fox%20Play%20Asia"
                    with urllib.request.urlopen(library_id_url) as url:
                        data = json.loads(url.read().decode())
                        for item in data["entries"]:
                            if "pl1$houseID" in item:
                                another_add[str(item["pl1$houseID"])]["lib"] = item["approved"]
                        mpx_image_dict = {str(item['pl1$houseID']): item.get("defaultThumbnailUrl","") if self.show_image.get() else item.get("pl2$imageApproved",item.get("pl3$imageApproved",item.get("pl1$imageApproved",""))) for item in data["entries"] if 'pl1$houseID' in item}
                    house_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B" + "%7C".join(
                        self.get_house_id()) + "%7D&token=" + self.tok + "&account=FoxPlus%20Asia%20DP"
                    with urllib.request.urlopen(house_id_url) as url:
                        data = json.loads(url.read().decode())
                        for item in data["entries"]:
                            if "pl1$houseID" in item:
                                if self.search_by_start:
                                    another_add[str(item["pl1$houseID"])]["dp"] = datetime.fromtimestamp(
                                        item['availableDate'] / 1000).strftime('%Y-%m-%d %H:%M') if item["availableDate"] != 0 else ""
                                else:
                                    another_add[str(item["pl1$houseID"])]["dp"] = datetime.fromtimestamp(
                                        item['expirationDate'] / 1000).strftime('%Y-%m-%d %H:%M') if item["expirationDate"] != 0 else ""
                    for child in self.tree.get_children():
                        if str(self.tree.item(child)["values"][15]) in another_add:
                            old_value = list(self.tree.item(child)["values"])
                            old_value[18] = another_add[str(self.tree.item(child)["values"][15])]["lib"]
                            old_value[19] = another_add[str(self.tree.item(child)["values"][15])]["dp"]
                            self.tree.item(child, value=old_value)
                    def populate_thumbnail(a, b):
                        for child in self.tree.get_children():
                            h_id = str(self.tree.item(child)["values"][15])
                            if h_id in mpx_image_dict:
                                if self.show_image.get():
                                    try:
                                        raw_data = urllib.request.urlopen(
                                            mpx_image_dict[h_id]).read()
                                        im = Image.open(io.BytesIO(raw_data)).resize((40, 22))
                                        image = ImageTk.PhotoImage(im)
                                        self.image_ref.append(image)
                                        self.tree.item(child, image=image, text="")
                                    except (urllib.error.HTTPError,ValueError):
                                        pass
                                else:
                                    if mpx_image_dict.get(h_id):
                                        self.tree.item(child, image="", text="True")
                                    else:
                                        self.tree.item(child, image="", text="False")
                    self.start_thread(populate_thumbnail)
                    self.reset_idle_timer()
                    if any(s in self.uid.upper() for s in ("HENRY.YIK", "KENNETH.CHAN", "ASIC","YUEN.LAU")):
                        self.nb.tab(4, state="normal")
                        self.sub_purge.entry.insert(0, self.sub_purge.message)
                    self.bottom_frame.config(cursor="arrow")
                    self.mpx_button.config(state="normal")
                    self.enable_all()
                except:
                    self.tok = ""
                    self.nb.tab(4, state="disabled")
                    self.mpx_connect_console()
                self.miss_dict,self.filter_dict = {},{}
            self.start_thread(mpx_connect_with_token)

    def set_idle_timer(self):
        hours, remainder = divmod(self._t, 3600)
        mins, secs = divmod(remainder, 60)
        user_name = self.uid.split("@")[0].split("/")[-1].split(".")[0].capitalize() + " " + \
                    self.uid.split("@")[0].split("/")[-1].split(".")[-1].capitalize()
        if user_name:
            timeformat = "Welcome, {} - Your session will expire in {:02d}:{:02d}:{:02d}".format(user_name, hours, mins,
                                                                                                 secs)
        else:
            timeformat = 'Your session will expire in {:02d}:{:02d}:{:02d}'.format(hours, mins, secs)
        self.right_message.config(text=timeformat)
        self._t -= 1
        if self._t <= 0:
            self.right_message.config(text="Your session has ended. Please re-login to MPX.")
            self.red_message.config(image=self.red_icon, text="") #  Not connected to MPX
            self.tok = ""
            self.nb.tab(4, state="disabled")
            return
        root.after(1000, self.set_idle_timer)

    def reset_idle_timer(self):
        self._t = 14400

    def refresh_inventory(self,a,b):
        self.translator_but.config(state=DISABLED)
        self.refresh_but.config(state=DISABLED)
        self.onair_but.config(state=DISABLED)
        self.search_button.config(state=DISABLED)
        self.mpx_button.config(state=DISABLED)
        self.tree.unbind("<Button-3>")

        self.load_df()

        self.display_message.config(text=" Inventory refreshed.")
        self.display_icon.config(image=self.check_icon)
        self.refresh_but.config(state=NORMAL)
        self.search_button.config(state=NORMAL)
        self.onair_but.config(state=NORMAL)
        self.mpx_button.config(state=NORMAL)
        #self.menubar.entryconfig("House/Box Translator", state="normal")
        self.translator_but.config(state=NORMAL)
        self.tree.bind("<Button-3>", self.popup)

    def socket_messaging(self):
        if not self.uid:
            self.display_message.config(text=" Please login first for database requests.")
            self.display_icon.config(image=self.error_icon)
            return
        self.import_but.config(relief="sunken",state="disabled")
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
                s.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
                s.connect(("10.219.2.18", 443)) #10.219.2.18 server
                text = "VISIONLOADER"#f"User {self.uid} requested update to database"
                s.sendall(text.encode())
            self.display_message.config(text=" Successfully submitted request to update database.")
            self.display_icon.config(image=self.check_icon)
        except (OSError, Exception):
            self.display_message.config(text=" Failed to request database update. Please try again later.")
            self.display_icon.config(image=self.error_icon)

    def create_top_frame(self,master): #FIXME color changes
        self.top_frame = Frame(master,bg=self.default_color)
        self.top_frame.grid()
        top_first = Frame(self.top_frame,bg=self.default_color)
        top_first.grid(row=1, column=0, rowspan=2)
        self.tutorial_icon = PhotoImage(file="icons/download.png") #tutorial3
        self.import_but = Button(top_first, image=self.tutorial_icon, compound="left",background=self.button_color,foreground="white", anchor="w",relief="flat",cursor="hand2")
        self.import_but.grid(row=0, column=1, padx=5, pady=2, sticky=W)
        self.import_but.config(command=self.socket_messaging) #self.first_time_use
        CreateToolTip(self.import_but, "Request database update")
        self.refresh_icon = PhotoImage(file="icons/reload2.png") #icons/reload2.png
        self.refresh_but = Button(top_first, image=self.refresh_icon, compound="left",background=self.button_color,relief="flat",
                             command=lambda: self.start_thread(self.refresh_inventory), anchor="w",cursor="hand2")
        self.refresh_but.grid(row=0, column=2, padx=5, pady=2)
        CreateToolTip(self.refresh_but, "Reload data from Console")
        self.inventory_icon = PhotoImage(file="icons/color2.png") #"icons/color2.png"
        self.inventory_but = Button(top_first, image=self.inventory_icon, compound="left", background=self.button_color,foreground="white",
                               command=self.colourize, relief="flat", anchor="w",cursor="hand2")
        self.inventory_but.grid(row=0, column=3, padx=5, pady=2)
        CreateToolTip(self.inventory_but, "Series: Brown\nSCC: Green\nNGC: Yellow\n"
                                          "Movies: Blue\nSCM: Purple\nTrailer: Black")
        self.onair_icon = PhotoImage(file="icons/info2.png")
        self.onair_but = Button(top_first, image=self.onair_icon, compound="left",background=self.button_color,foreground="white", command=self.open_data,
                           relief="flat", anchor="w",cursor="hand2")
        self.onair_but.grid(row=0, column=4, padx=5, pady=2)
        CreateToolTip(self.onair_but, "Show imported data information")
        self.error_icon = PhotoImage(file="icons/error.png")
        self.error_but = Button(top_first, image=self.error_icon)
        img = [self.tutorial_icon,self.refresh_icon,self.inventory_icon,self.onair_icon,self.error_icon]
        self.restart_icon = PhotoImage(file="icons/restart2.png")
        self.restart_but = Button(top_first, image=self.restart_icon, relief="flat", anchor="w", background=self.button_color,cursor="hand2")
        self.restart_but.grid(row=0, column=0, padx=5, pady=2)
        self.image_ref.append(self.restart_icon)
        CreateToolTip(self.restart_but, "Restart client")
        ########################### Frame creation ###########################
        self.top_left = Frame(self.top_frame,bg=self.default_color)
        self.top_left.grid(row=1, column=1, rowspan=2,sticky="e")

        self.top_mid = Frame(self.top_frame,bg=self.default_color)
        self.top_mid.grid(row=1, column=2,sticky="w")

        self.bot_mid = Frame(self.top_frame,bg=self.default_color)
        self.bot_mid.grid(row=2, column=2,sticky="w")

        self.top_right = Frame(self.top_frame,bg=self.default_color)
        self.top_right.grid(row=1, column=3, rowspan=2,sticky="w")

        self.top_frame.grid_columnconfigure(0, weight=1, uniform="group1")
        self.top_frame.grid_columnconfigure(1, weight=1, uniform="group1")
        self.top_frame.grid_columnconfigure(2, weight=1, uniform="group1")
        self.top_frame.grid_columnconfigure(3, weight=1, uniform="group1")
        self.top_frame.grid_rowconfigure(0, weight=1)

        self.bottom_frame = Frame(self.series_page,bg=self.default_color)
        self.bottom_frame.grid()

    def change_colour(self,wid):
        self.series_page.config(bg="brown")
        self.top_left.config(bg="brown")
        a = wid.winfo_children()
        for widget in a:
            try:
                if not "entry" in str(widget):
                    widget.config(bg="brown")
            except (TclError,AttributeError):
                pass
            if "frame" in str(widget):
                self.change_colour(widget)

    def open_data(self):
        self.data = DataInfo(self.series_page)
        self.onair_but.config(relief="sunken",command=self.close_data)
        self.data.show_data()

    def close_data(self):
        self.onair_but.config(relief="flat",command=self.open_data)
        self.data.hide_data()

    def disable_all(self):
        self.refresh_but.config(state=DISABLED)
        self.filter_entry.config(state=DISABLED)
        self.filter_go.config(state=DISABLED)
        self.search_button.config(state=DISABLED)
        self.mpx_button.config(state=DISABLED)
        self.from_date.config(state=DISABLED)
        self.to_date.config(state=DISABLED)
        self.tree.unbind("<Double-1>")
        self.tree.unbind("<Button-3>")

    def enable_all(self):
        self.refresh_but.config(state=NORMAL)
        self.filter_entry.config(state=NORMAL)
        self.filter_go.config(state=NORMAL)
        self.search_button.config(state=NORMAL)
        self.mpx_button.config(state=NORMAL)
        self.from_date.config(state=NORMAL)
        self.to_date.config(state=NORMAL)
        self.tree.bind("<Double-1>", self.OnDoubleClick)
        self.tree.bind("<Button-3>", self.popup)

    def colourize(self):
        self.tree.tag_configure("series", background="saddle brown", foreground="white")
        self.tree.tag_configure("scc", background="chartreuse2", foreground="black")
        self.tree.tag_configure("factual", background="yellow2", foreground="black")
        self.tree.tag_configure("movies", background="deep sky blue", foreground="black")
        self.tree.tag_configure("scm", background="purple", foreground="white")
        self.tree.tag_configure("scm_series", background="purple", foreground="white")
        self.tree.tag_configure("trailer", background="black", foreground="white")
        self.inventory_but.config(command=self.decolourize,relief="sunken")

    def decolourize(self):
        self.tree.tag_configure("series", background=self.tree_bg, foreground=self.tree_font)
        self.tree.tag_configure("scc", background=self.tree_bg, foreground=self.tree_font)
        self.tree.tag_configure("factual", background=self.tree_bg, foreground=self.tree_font)
        self.tree.tag_configure("movies", background=self.tree_bg, foreground=self.tree_font)
        self.tree.tag_configure("scm", background=self.tree_bg, foreground=self.tree_font)
        self.tree.tag_configure("scm_series", background=self.tree_bg, foreground=self.tree_font)
        self.tree.tag_configure("trailer", background=self.tree_bg, foreground=self.tree_font)
        self.inventory_but.config(command=self.colourize,relief="flat")

    def create_top_right(self):
        self.green_icon = PhotoImage(file="icons/connected.png") #connected
        self.green_icon = self.green_icon.subsample(4)
        self.green_message = Label(self.top_right, image="", text="", relief=FLAT,
                                   compound="center", anchor="e",justify=RIGHT,bg=self.default_color)
        self.green_message.img = self.green_icon
        self.red_icon = PhotoImage(file="icons/disconnected.png") #disconnected
        self.red_icon = self.red_icon.subsample(4)
        self.red_message = Label(self.top_frame, image=self.red_icon,
                                 relief=FLAT, compound="left",justify=CENTER,
                                 bg=self.default_color,fg=self.font_color,cursor="question_arrow")
        self.red_message.img = self.red_icon
        self.red_message.place(relx=0.968,rely=0.15)
        CreateToolTip(self.red_message, "Connectivity to MPX")
        self.all_var = IntVar()
        self.all_var.set(1)
        self.all_var_check = Checkbutton(self.top_right, variable=self.all_var, borderwidth=2, text="Select None",
                                         relief="flat", foreground=self.font_color,
                                         command=lambda: self.check_uncheck_all(self.all_var.get()), anchor="w",
                                         justify=LEFT, width=9, selectcolor="dark slate gray",
                                         bg=self.default_color, activebackground=self.default_color,
                                         activeforeground=self.font_color)
        self.all_var_check.grid(row=0, column=0, padx=5,sticky="n")
        self.search_icon = PhotoImage(file="icons\search2.png")
        self.search_button = Button(self.top_right, image=self.search_icon, compound="left",text=" Search ",
                                    command=self.display_tree, relief="flat",cursor="hand2",
                                    background=self.default_color, foreground=self.font_color,width=64)
        self.search_button.grid(row=1, column=0, padx=1,sticky="w")
        CreateToolTip(self.search_button,"Find result in schedule")
        self.search_button.img = self.search_icon
        self.mpx_icon = PhotoImage(file="icons\mpx.png")
        self.mpx_icon = self.mpx_icon.subsample(8)
        self.mpx_button = Button(self.top_right, image=self.mpx_icon, compound="top",
                                 text="", command=self.mpx_connect_console,
                                 relief="flat", background=self.default_color,foreground=self.font_color,cursor="hand2")
        self.mpx_button.grid(row=0, column=1,sticky="w",rowspan=2)
        self.mpx_button.img = self.mpx_icon
        CreateToolTip(self.mpx_button, "Query MPX")
        self.console_icon = PhotoImage(file="icons/inspection2.png") #sherlock
        self.console_button = Button(self.top_right,image=self.console_icon,command=self.analyze_media, #start_hover_info
                                     relief="flat", bg=self.default_color,fg=self.font_color,cursor="hand2")
        self.console_button.grid(row=0,column=2,sticky="w",padx=5,rowspan=2)
        self.console_button.img = self.console_icon
        CreateToolTip(self.console_button, "Analyze mode")
        self.logout_icon = PhotoImage(file="icons/logout2.png") #"icons/logout2.png"
        self.clear_login = Button(self.top_right,relief="flat",image=self.logout_icon,
                                  command=self.clear_login,bg=self.default_color,cursor="hand2")
        self.clear_login.grid(row=0, column=3, sticky="w",rowspan=2)
        self.clear_login.img = self.logout_icon
        CreateToolTip(self.clear_login,"Logout from MPX")
        if settings.analyze_mode:
            self.console_button.invoke()

    def analyze_media(self):
        settings.analyze_mode = True
        try:
            settings.save()
        except PermissionError:
            pass
        self.console_button.config(relief="sunken", command=self.stop_analyze_media)

    def stop_analyze_media(self):
        settings.analyze_mode = False
        try:
            settings.save()
        except PermissionError:
            pass
        self.console_button.config(relief="flat", command=self.analyze_media)

    def start_hover_info(self):
        """
        Deprecated
        :return:
        """
        for item in self.tree.selection():
            self.tree.selection_remove(item)
        #self.tree.config(selectmode="browse")
        #self.bottom_frame.config(cursor="hand2")
        #self.tree.bind("<<TreeviewSelect>>", self.hover_canvas_new)
        self.current_iid = None
        self.tree.bind("<Motion>", self.hover_canvas)
        self.console_button.config(relief="sunken",command=self.stop_hover_info)

    def stop_hover_info(self):
        """
        Deprecated
        :return:
        """
        #self.tree.config(selectmode="extended")
        #self.bottom_frame.config(cursor="arrow")
        self.tree.unbind("<<TreeviewSelect>>")
        self.console_button.config(relief="flat",command=self.start_hover_info)
        try:
            self.hover_info.destroy()
        except (AttributeError,TclError):
            pass

    def tree_callback_new(self, event):
        def something():
            self._iid = "".join(self.tree.selection())
            if self._iid != self.last_focus:
                if self.last_focus:
                    try:
                        self.hover_info.destroy()
                    except TclError:
                        pass
                try:
                    self.hover_canvas(event)
                except IndexError:
                    pass
                self.last_focus = self._iid

        root.after(10, something)

    def hover_canvas(self, event):
        self.hover_info = Canvas(self.series_page)
        values = self.tree.item(self._iid)["values"]
        hover_frame = Frame(self.hover_info, bg="PaleGreen4", highlightbackground="green", highlightcolor="green",
                            highlightthickness=1)
        hover_frame.pack(fill=X)
        Label(hover_frame, text=values[0], bg="PaleGreen4", fg="white").pack(side=LEFT, padx=1)
        lower_frame = Frame(self.hover_info, bg=self.default_bg, highlightbackground="green", highlightcolor="green",
                            highlightthickness=1)
        lower_frame.pack(fill=BOTH)
        for num, text in enumerate(
                ("Season:", "SEA start:", "SEA end:", "TW start:", "TW end:", "PH start:", "PH end:", "House No:")):
            Label(lower_frame, text=text, bg=self.default_bg, fg=self.default_font, width=9, anchor='w').grid(row=num,
                                                                                                              column=0,
                                                                                                              sticky="w")
        for num, text in enumerate(
                ("Episode:", "HK start:", "HK end:", "SG start:", "SG end:", "Version:", "Box No:", "On-Air ID")):
            Label(lower_frame, text=text, bg=self.default_bg, fg=self.default_font, width=8, anchor='w').grid(row=num,
                                                                                                              column=2,
                                                                                                              sticky="w")
        Label(lower_frame, text=values[1] if values[1] else "N/A",
              bg=self.default_bg, fg=self.default_font, width=10, anchor='w').grid(row=0, column=1, sticky="w")
        Label(lower_frame, text=values[2] if values[2] else "N/A",
              bg=self.default_bg, fg=self.default_font, width=10, anchor='w').grid(row=0, column=3, sticky="w")
        for num, text in enumerate(["SEA start", "SEA end", "TW start", "TW end", "PH start", "PH end"]):
            Label(lower_frame, text=values[num + 3],
                  bg=self.default_bg, fg=self.default_font, width=10, anchor='w').grid(row=num + 1, column=1,
                                                                                       sticky="w")
        for num, text in enumerate(["HK start", "HK end", "SG start", "SG end", "Version", "Box No."]):
            Label(lower_frame, text=values[num + 9],
                  bg=self.default_bg, fg=self.default_font, width=12, anchor='w').grid(row=num + 1, column=3,
                                                                                       sticky="w")
        Label(lower_frame, text=values[15] if values[15] else "",
              bg=self.default_bg, fg=self.default_font, width=10, anchor='w').grid(row=7, column=1, sticky="w")
        Label(lower_frame, text=values[21] if values[21] else "",
              bg=self.default_bg, fg=self.default_font, width=10, anchor='w').grid(row=7, column=3, sticky="w")
        self.hover_info.place(relx=0.001, rely=0.605)  # relx=0.001,rely=0.619
        current_iid = "".join(self.tree.selection())

        def added_info(a, b):
            time.sleep(0.5)
            house_list_dp, house_list_sch = {}, {}
            try:
                if self._iid != current_iid:
                    return
            except RuntimeError:
                return

            def dict_compare(d1, d2):
                d1_keys = set(d1.keys())
                d2_keys = set(d2.keys())
                intersect_keys = d1_keys.intersection(d2_keys)
                added = d1_keys - d2_keys
                removed = d2_keys - d1_keys
                modified = {o: (d1[o], d2[o]) for o in intersect_keys if d1[o] != d2[o]}
                same = set(o for o in intersect_keys if d1[o] == d2[o])
                return added, removed, modified, same

            house_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:channel_id,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B" \
                           + str(
                self.tree.item(self._iid)["values"][15]) + "%7D&token=" + self.tok + "&account=FoxPlus%20Asia%20DP"
            try:
                with urllib.request.urlopen(house_id_url) as url:
                    data = json.loads(url.read().decode())
                    house_list_dp = {"".join(item["pl2$channel_id"]):
                                         {"start": datetime.fromtimestamp(item['availableDate'] // 1000).strftime(
                                             '%Y-%m-%d %H:%M') if item['availableDate'] != 0 else "",
                                          "end": datetime.fromtimestamp(item['expirationDate'] // 1000).strftime(
                                              '%Y-%m-%d %H:%M') if item['expirationDate'] != 0 else ""}
                                     for item in data["entries"]}
                    house_list_dp = {k: house_list_dp[k] for k in sorted(house_list_dp.keys())}
                    if house_list_dp:
                        Label(lower_frame, text="DP scheduling info", bg="RoyalBlue3", fg="white", anchor="w").grid(
                            row=0, column=4, columnspan=6, sticky="ew")  # "dim gray"
                        ttk.Separator(lower_frame, orient=VERTICAL).grid(row=1, column=4, rowspan=8, sticky='ns')
                        ttk.Separator(lower_frame, orient=VERTICAL).grid(row=1, column=6, rowspan=8, sticky='ns')
                        ttk.Separator(lower_frame, orient=VERTICAL).grid(row=1, column=8, rowspan=8, sticky='ns')
                        start = 2
                        Label(lower_frame, text="Channel ID", bg=self.default_bg, fg=self.default_font,
                              anchor="w").grid(row=1, column=5, sticky="w")
                        Label(lower_frame, text="Start date", bg=self.default_bg, fg=self.default_font,
                              anchor="w").grid(row=1, column=7, sticky="w")
                        Label(lower_frame, text="End date", bg=self.default_bg, fg=self.default_font, anchor="w").grid(
                            row=1, column=9, sticky="w")
                        for k, v in house_list_dp.items():
                            Label(lower_frame, text=k, bg=self.default_bg, fg=self.default_font, anchor="w").grid(
                                row=start, column=5, sticky="w")
                            Label(lower_frame, text=v["start"], bg=self.default_bg, fg=self.default_font,
                                  anchor="w").grid(row=start, column=7, sticky="w")
                            Label(lower_frame, text=v["end"], bg=self.default_bg, fg=self.default_font,
                                  anchor="w").grid(row=start, column=9, sticky="w")
                            start += 1
            except:
                pass
            time.sleep(0.5)
            try:
                if self._iid != current_iid:
                    return
            except RuntimeError:
                return
            house_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:channel_id,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B" \
                           + str(self.tree.item(self._iid)["values"][
                                     15]) + "%7D&token=" + self.tok + "&account=FoxPlus%20Asia%20HK"  # FoxPlus Asia HK
            tw_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:channel_id,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B" \
                     + str(self.tree.item(self._iid)["values"][
                               15]) + "%7D&token=" + self.tok + "&account=FoxPlus%20Asia%20TW"  # FoxPlus Asia TW
            try:
                with urllib.request.urlopen(tw_url) as tw:
                    tw_data = json.loads(tw.read().decode())
                    tw_house_list = {"".join(item["pl2$channel_id"]):
                                         {"start": datetime.fromtimestamp(item['availableDate'] // 1000).strftime(
                                             '%Y-%m-%d %H:%M') if item['availableDate'] != 0 else "",
                                          "end": datetime.fromtimestamp(item['expirationDate'] // 1000).strftime(
                                              '%Y-%m-%d %H:%M') if item['expirationDate'] != 0 else ""}
                                     for item in tw_data["entries"]}
                with urllib.request.urlopen(house_id_url) as url:
                    data = json.loads(url.read().decode())
                    house_list_sch = {"".join(item["pl2$channel_id"]):
                                          {"start": datetime.fromtimestamp(item['availableDate'] // 1000).strftime(
                                              '%Y-%m-%d %H:%M') if item['availableDate'] != 0 else "",
                                           "end": datetime.fromtimestamp(item['expirationDate'] // 1000).strftime(
                                               '%Y-%m-%d %H:%M') if item['expirationDate'] != 0 else ""}
                                      for item in data["entries"]}
                    house_list_sch.update(tw_house_list)
                    house_list_sch = {k: house_list_sch[k] for k in sorted(house_list_sch.keys())}
                    added, removed, modified, same = dict_compare(house_list_dp, house_list_sch)
                    if house_list_sch:
                        Label(lower_frame, text="TW/HK scheduling info", bg="dark orchid", fg="white", anchor="w").grid(
                            row=0, column=10, columnspan=6, sticky="ew")
                        ttk.Separator(lower_frame, orient=VERTICAL).grid(row=1, column=10, rowspan=8, sticky='ns')
                        ttk.Separator(lower_frame, orient=VERTICAL).grid(row=1, column=12, rowspan=8, sticky='ns')
                        ttk.Separator(lower_frame, orient=VERTICAL).grid(row=1, column=14, rowspan=8, sticky='ns')
                        start = 2
                        Label(lower_frame, text="Channel ID", bg=self.default_bg, fg=self.default_font,
                              anchor="w").grid(row=1, column=11, sticky="w")
                        Label(lower_frame, text="Start date", bg=self.default_bg, fg=self.default_font,
                              anchor="w").grid(row=1, column=13, sticky="w")
                        Label(lower_frame, text="End date", bg=self.default_bg, fg=self.default_font, anchor="w").grid(
                            row=1, column=15, sticky="w")
                        for k, v in house_list_sch.items():
                            Label(lower_frame, text=k, bg=self.default_bg, anchor="w", compound="left",
                                  fg=self.default_font if k not in modified and k not in removed else "red",
                                  image=self.error_icon if k in modified or k in removed else ""
                                  ).grid(row=start, column=11, sticky="w")
                            Label(lower_frame, text=v["start"], bg=self.default_bg,
                                  fg=self.default_font if k not in modified and k not in removed else "red",
                                  anchor="w").grid(row=start, column=13, sticky="w")
                            Label(lower_frame, text=v["end"], bg=self.default_bg,
                                  fg=self.default_font if k not in modified and k not in removed else "red",
                                  anchor="w").grid(row=start, column=15, sticky="w")
                            start += 1
            except:
                pass

        if self.tok:
            self.start_thread(added_info)

    def clear_login(self):
        self.tok = ""
        self._t = 1
        login_info = {"autologin": 0, "username": "", "password": ""}
        self.nb.tab(4, state="disabled")
        try:
            with open("login_info.json","w") as f:
                json.dump(login_info,f)
        except PermissionError:
            pass

    def create_top_left(self):
        self.translator = PhotoImage(file="icons/translator.png")
        self.translator_but = Button(self.top_frame,relief="flat",image=self.translator,
                                     command=self.open_translator,bg=self.default_color,cursor="hand2")
        self.translator_but.place(relx=0.25,rely=0.01)
        self.translator_but.img = self.translator
        CreateToolTip(self.translator_but,"House/Box Translator")
        self.trailer_var = IntVar(value=0)
        Checkbutton(self.top_left, variable=self.trailer_var, anchor="e", borderwidth=0, relief="solid",
                    text="Include trailers", bg=self.default_color, selectcolor="dark slate gray",
                    foreground=self.font_color,
                    activebackground=self.default_color, activeforeground=self.font_color
                    ).grid(row=0,column=3,columnspan=1,sticky="e",padx=2)
        self.filter_combo = ttk.Combobox(self.top_left, state="readonly", width=9)
        self.filter_combo["values"] = ("Search by", "Filter by")
        self.filter_var = IntVar()
        self.filter_var.set(0)
        radio_but_one = Radiobutton(self.top_left, bd=2,relief="groove",selectcolor="purple3",
                              bg=self.default_color,fg=self.font_color,cursor="hand2",
                              text="Search",padx=10, variable=self.filter_var,value=0,indicatoron=0)
        radio_but_two = Radiobutton(self.top_left, bd=2,relief="groove",selectcolor="purple3",
                              bg=self.default_color,fg=self.font_color,cursor="hand2",
                              text="Filter",padx=10, variable=self.filter_var,value=1,indicatoron=0)
        radio_but_one.grid(row=0, column=1, sticky="ew")
        radio_but_two.grid(row=0, column=2, sticky="ew")
        self.vcmd = (self.top_left.register(self._columns_searcher), '%P')
        self.filter_entry = AutocompleteEntry(self.top_left, width=38, validate="key",borderwidth=2,
                            fg = "grey")
        self.filter_entry.insert(0," Search or filter result (case insensitive)")
        self.filter_entry.bind("<Return>", (lambda event: self.filter_data()))
        self.filter_entry.bind('<Button-3>', self.rClicker, add='')
        self.filter_entry.bind('<FocusIn>', self.entry_click)
        self.filter_entry.bind('<FocusOut>',
                               lambda e: self.entry_focusout(e, " Search or filter result (case insensitive)"))
        self.filter_entry.grid(row=1, column=1, columnspan=3, pady=5,sticky="e")
        self.filter_icon = PhotoImage(file="icons\search3.png")
        self.filter_go = Button(self.top_left, text="GO", image=self.filter_icon,
                                command=self.filter_data, relief="flat",bg="white")
        self.filter_go.place(relx=0.91,rely=0.566)
        self.filter_go.img = self.filter_icon

    def entry_click(self,event):
        if event.widget["foreground"] == "grey":
            event.widget.delete(0, "end")
            event.widget.insert(0, "")
            event.widget.configure(fg="black", validatecommand=self.vcmd)

    def entry_focusout(self,event, msg):
        if not event.widget.get():
            event.widget.configure(fg="grey", validatecommand="")
            event.widget.insert(0, msg)

    def open_translator(self):
        self.translator_but.config(relief="sunken",command=self.close_translator)
        self.hbox = HouseBoxTransform(self.series_page)
        self.hbox.show_box()
        self.hbox.input_entry.focus_set()

    def close_translator(self):
        self.translator_but.config(relief="flat",command=self.open_translator)
        self.hbox.hide_box()

    def create_top_mid(self):
        self.series_var = IntVar()
        self.series_var.set(1)
        series_check = Checkbutton(self.top_mid, variable=self.series_var, anchor="e", borderwidth=0,
                                   relief="solid", text="Series",selectcolor="dark slate gray",
                                   bg=self.default_color, foreground=self.font_color,
                                   activebackground=self.default_color,activeforeground=self.font_color)
        series_check.grid(row=0, column=0, padx=1)
        self.movies_var = IntVar()
        self.movies_var.set(1)
        movies_check = Checkbutton(self.top_mid, variable=self.movies_var, borderwidth=0, relief="solid",
                                   text="Movies",selectcolor="dark slate gray",
                                   bg=self.default_color, foreground=self.font_color,
                                   activebackground=self.default_color,activeforeground=self.font_color)
        movies_check.grid(row=0, column=1, padx=1)
        self.factual_var = IntVar()
        self.factual_var.set(1)
        factual_check = Checkbutton(self.top_mid, variable=self.factual_var, borderwidth=0, relief="solid",
                                    text="Factual",selectcolor="dark slate gray",
                                    bg=self.default_color, foreground=self.font_color,
                                    activebackground=self.default_color,activeforeground=self.font_color)
        factual_check.grid(row=0, column=2, padx=1)
        self.scc_var = IntVar()
        self.scc_var.set(1)
        scc_check = Checkbutton(self.top_mid, variable=self.scc_var, borderwidth=0, relief="solid",
                                text="SCC",selectcolor="dark slate gray",
                                bg=self.default_color, foreground=self.font_color,
                                activebackground=self.default_color,activeforeground=self.font_color)
        scc_check.grid(row=0, column=3, padx=1)
        self.scm_var = IntVar()
        self.scm_var.set(1)
        scm_check = Checkbutton(self.top_mid, variable=self.scm_var, borderwidth=0, relief="solid",
                                text="SCM",selectcolor="dark slate gray",
                                bg=self.default_color, foreground=self.font_color,
                                activebackground=self.default_color,activeforeground=self.font_color)
        scm_check.grid(row=0, column=4, padx=1)

    def create_bottom_mid(self):
        Label(self.bot_mid,text="Start date:",fg=self.font_color,bg=self.default_color).grid(row=0,column=0,padx=2)
        self.from_date = CustomDateEntry(self.bot_mid, width=10, background='darkblue',
                              foreground='white', borderwidth=2, columnspan=2)
        self.from_date.grid(row=0, column=1, padx=2, pady=1)
        CreateToolTip(self.from_date, "Choose media START date")
        Label(self.bot_mid, text="End date:", fg=self.font_color, bg=self.default_color).grid(row=0, column=3, padx=2)
        self.to_date = CustomDateEntry(self.bot_mid, width=10, background='darkblue',
                            foreground='white', borderwidth=2, columnspan=2)
        self.to_date.grid(row=0, column=4, padx=2, pady=1)
        self.to_date.set_date((datetime.now()+timedelta(days=7)).strftime("%x"))
        CreateToolTip(self.to_date, "Choose media END date")

    def create_lowest(self):
        self.lowest_frame = Frame(self.series_page,background=self.default_color,
                                  highlightbackground="black", highlightcolor="black", highlightthickness=1)
        self.lowest_frame.grid(sticky=EW)
        self.wait_icon = PhotoImage(file="icons/wait.png")
        self.check_icon = PhotoImage(file="icons/check.png")
        self.result_icon = PhotoImage(file="icons/result.png")
        self.config_icon = PhotoImage(file="icons/config.png")
        self.display_icon = Label(self.lowest_frame, width=8,bg=self.default_color)
        self.display_icon.grid(row=0, column=0,padx=5)
        self.display_icon.img = self.wait_icon
        self.display_message = Label(self.lowest_frame, image="", text="   ", compound="left",
                                     anchor="w", relief=FLAT, justify=LEFT, width=115,
                                     background=self.default_color,foreground=self.font_color) #62
        self.display_message.grid(row=0, column=1, sticky=W)
        self.middle_message = Label(self.lowest_frame, image="", text=" ", width=50, borderwidth=0,
                                    relief=FLAT,bg=self.default_color,fg=self.font_color)
        self.middle_message.img = self.result_icon
        self.right_message = Label(self.lowest_frame, image="", text="Waiting for connection to MPX...",
                                   compound="right", anchor="e", relief=FLAT,
                                   background=self.default_color,foreground=self.font_color,
                                   justify=RIGHT, width=50)
        self.right_message.grid(row=0, column=3, sticky=E)
        self.right_message.img = self.check_icon
        CreateToolTip(self.right_message,"Time before your token reaches idle limit and expires."
                                         " Token is automatically refreshed when you perform a query.")
        self.lowest_frame_mpb = ttk.Progressbar(self.lowest_frame, orient="horizontal",
                                                length=200, mode="determinate")
        self.lowest_frame_mpb["maximum"] = 100
        self.lowest_frame_mpb["value"] = 100

    def check_uncheck_all(self,state):
        self.series_var.set(state)
        self.movies_var.set(state)
        self.factual_var.set(state)
        self.scc_var.set(state)
        self.scm_var.set(state)
        if state == 1:
            self.all_var_check.config(text="Select None")
        else:
            self.all_var_check.config(text="Select All")

    def display_tree(self):
        """
        Search dataframe by selected date range and genres. Display results on Treeview widget.
        Start date: Read from self.from_date <class DateEntry>
        End date:   Read from self.to_date <class DateEntry>
        Genre:      Read from combinations of tk.IntVar
        :return: None
        """
        all_var = [self.trailer_var, self.series_var,self.movies_var,self.factual_var,self.scc_var,self.scm_var]
        self.genre = ["trailer","series", "movies", "factual", "scc", "scm"]
        channels = [self.genre[num] for num, v in enumerate(all_var) if v.get()]
        if not channels:    # return early if no channels selected
            self.display_message.config(text="Please select one or more genre.")
            self.display_icon.config(image=self.error_icon)
            return
        self.a = TreeTitle(self.tree, "")
        self.shorten_title = {}
        self.disable_all()
        self.tree.delete(*self.tree.get_children())
        #FIXME
        start_date = pd.Timestamp(self.from_date.get_date())
        end_date = pd.Timestamp(self.to_date.get_date())
        # x_start_date = datetime.strptime(self.from_date.get(), "%m/%d/%Y")
        # x_end_date = datetime.strptime(self.to_date.get(), "%m/%d/%Y")
        self.tree.delete(*self.tree.get_children())
        cond1 = (self.df["genre"].str.contains("|".join(channels)))
        if self.search_by_start:
            cond2 = (self.df["sea_start"].between(start_date, end_date)) | \
                    (self.df["tw_start"].between(start_date, end_date)) | \
                    (self.df["ph_start"].between(start_date, end_date)) | \
                    (self.df["hk_start"].between(start_date, end_date)) | \
                    (self.df["sg_start"].between(start_date, end_date))
        else:
            cond2 = (self.df["sea_end"].between(start_date, end_date)) | \
                    (self.df["tw_end"].between(start_date, end_date)) | \
                    (self.df["ph_end"].between(start_date, end_date)) | \
                    (self.df["hk_end"].between(start_date, end_date)) | \
                    (self.df["sg_end"].between(start_date, end_date))

        dates = ['sea_start', 'sea_end', 'tw_start', 'tw_end', 'ph_start',
                 'ph_end', 'hk_start', 'hk_end', 'sg_start', 'sg_end']

        temp = self.df[cond1 & cond2].copy()
        temp = temp[(~temp.duplicated("ID")) | (temp['ID'].isnull())]
        temp[dates] = temp[dates].astype(str).replace({"NaT": ""})
        for row in temp.itertuples():
            if len(row[1]) >= self.wrap_length:# and row[22] in ("series","movies"):
                name = row[1][:self.wrap_length] + "..."
                name = name.title()
                self.shorten_title[name] = row[1].title()
            else:
                name = row[1].title()
            self.tree.insert('', 0, text="", values=[name, *list(row[2:17]), "", row[18],
                                                     "", "", row[17], row[21]], tags=(row[23],))
        message = " Found {} results.".format(len(self.tree.get_children()))
        self.display_message.config(text=message)
        self.display_icon.config(image=self.result_icon)
        if len(self.tree.get_children()) >= 1:
            self.filter_entry.config(state=NORMAL)
            self.filter_go.config(state=NORMAL)
        else:
            self.filter_entry.config(state=DISABLED)
            self.filter_go.config(state=DISABLED)
        self.enable_all()
        self.bottom_frame.config(cursor="arrow")
        self.search_button.config(text=" Search", command=self.display_tree, image=self.search_icon)

    def filter_data(self):
        """
        Method to search title by name. Perform search when self.filter_var == 0.
        :return: None
        """
        if self.filter_entry.get() != "" and not self.filter_var.get() and len(self.filter_entry.get()) < 2:
            self.display_icon.config(image=self.error_icon)
            self.display_message.config(text=" Too few keywords entered - try to search with 2 characters or more.")
        elif self.filter_entry.get() != "" and not self.filter_var.get():
            self.tree.delete(*self.tree.get_children())
            schedule_order = ["series" if self.series_var.get() else "",
                              "scc" if self.scc_var.get() else "",
                              "factual" if self.factual_var.get() else "",
                              "movies" if self.movies_var.get() else "",
                              "scm" if self.scm_var.get() else '',
                              "trailer" if self.trailer_var.get() else '']
            schedule_order = [i for i in schedule_order if i]
            temp = self.df.loc[(self.df["genre"].isin(schedule_order))&(self.df["title"].str.contains(self.filter_entry.get(),case=False))].copy()
            temp = temp.sort_values(["ID",'pref'], ascending=False)
            temp = temp[(~temp.duplicated("ID")) | (temp['ID'].isnull())]
            dates = ['sea_start', 'sea_end', 'tw_start', 'tw_end', 'ph_start',
                     'ph_end', 'hk_start', 'hk_end', 'sg_start', 'sg_end']
            temp[dates] = temp[dates].astype(str).replace({"NaT": ""})
            for row in temp.itertuples():
                if len(row[1]) >= self.wrap_length:  # and row[22] in ("series","movies"):
                    name = row[1][:self.wrap_length] + "..."
                    name = name.title()
                    self.shorten_title[name] = row[1].title()
                else:
                    name = row[1].title()
                self.tree.insert('', 0, text="", values=[name, *list(row[2:17]), "", row[18], "", "", row[17], row[21]],
                                 tags=(row[23],))
            self.display_message.config(text=" Searching by keyword: \"{}\" - {} matches found."
                                        .format(self.filter_entry.get(),len(self.tree.get_children())))
            self.display_icon.config(image=self.result_icon)
            self.enable_all()
            self.bottom_frame.config(cursor="arrow")
            self.miss_dict, self.filter_dict = {},{}

    def filter_focus(self,event):
        self.filter_entry.focus_set()

    def create_tree_area(self):
        self.tree = ttk.Treeview(self.bottom_frame, selectmode='extended', height=24)
        self.tree.pack(side='left')
        vsb = ttk.Scrollbar(self.bottom_frame, orient="vertical",
                            command=self.tree.yview)  ## CREATE SCROLLBAR WIDGET AND ATTACH TO treeview.yview
        vsb.pack(side='right', fill='y')
        self.tree.configure(yscrollcommand=vsb.set)  ## ATTACH treeview TO SCROLLBAR
        header = (
        "Series", "Season", "Episode", "SEA start", "SEA end", "TW start", "TW end", "PH start", "PH end", "HK start",
        "HK end", "SG start", "SG end", "version", "box_no", "house_no", "D&D time", "qc_date", "Lib approved",
        "DP schedule date", "Channel")
        width = (220, 48, 52, 80, 80, 80, 80, 80, 80, 80, 80, 80, 80, 60, 85, 65, 80, 80, 80, 120, 120, 120)
        self.tree["columns"] = header
        # tree['show'] = 'headings'

        self.tree.column("#0", width=70, anchor="w",stretch=False)
        self.tree.heading("#0", text="Image", anchor="w")
        for i in range(len(header)):
            self.tree.column(header[i], width=width[i], anchor="w")
            self.tree.heading(header[i], text=header[i], anchor='w')
            self.tree["displaycolumns"] = (
        "Series", "Season", "Episode", "SEA start", "TW start", "PH start", "HK start", "SG start", "version", "box_no",
        "house_no", "Lib approved", "DP schedule date")

        for col in self.tree["columns"]:
            self.tree.heading(col, text=col, command=lambda _col=col: self.treeview_sort_column(self.tree, _col, False))

        self.tree.bind("<Double-1>", self.OnDoubleClick)
        self.tree.bind("<Button-3>", self.popup)
        self.last_focus = None

    def get_house_id(self):
        self.house_id_list = []
        for child in self.tree.get_children():
            try:
                self.house_id_list.append(str(self.tree.item(child)["values"][15]))
            except IndexError as e:
                message = "Function MainGUI.get_house_id: "+str(e)
                error_log(message)
        return self.house_id_list

    def treeview_sort_column(self, tv, col, reverse):
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)

        # rearrange items in sorted positions
        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        # reverse sort next time
        tv.heading(col, command=lambda: \
            self.treeview_sort_column(tv, col, not reverse))

    def OnDoubleClick(self,event):
        if self.tok != "":
            self.iid = self.tree.identify('item', event.x, event.y)
            item_text = self.tree.item(self.iid, "values")
            try:
                _ = str(item_text[15])
            except (IndexError, NameError):
                return
            try:
                found = self.df.loc[self.df["ID"].eq(int(item_text[-1])),"House Number"].values.tolist()
                if found:
                    self.house_id_combo = [i for i in found if i]
                else:
                    self.house_id_combo = [str(item_text[15])]
            except IndexError:
                self.house_id_combo = []
                print ("Index error")
            except Exception as e:
                self.house_id_combo = []
                print("Function: OnDoubleClick, cannot get house_id, {}".format(item_text[16]))
            if self.house_id_combo:
                self.house_id_event(self.house_id_combo)
        else:
            self.mpx_connect_console()

    def house_id_event(self,house_id_combo):
        if self.tok == "":
            self.mpx_connect_console()
        elif house_id_combo:
            try:
                library_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,title,approved,:channel_id,:workcode,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B" + "%7C".join(
                    house_id_combo) + "%7D&token=" + self.tok + "&account=Fox%20Play%20Asia"
                with urllib.request.urlopen(library_url) as url:
                    data = json.loads(url.read().decode())
                    lib_list = [("Library", item["pl1$houseID"], item["title"], item["pl2$workcode"], "", "",
                                 item["pl2$channel_id"][0], item["approved"]) for item in data["entries"]]
                sch_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,title,approved,:channel_id,:workcode,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B" + "%7C".join(
                    house_id_combo) + "%7D&token=" + self.tok + "&account=FoxPlus%20Asia%20HK"
                with urllib.request.urlopen(sch_id_url) as url:
                    data = json.loads(url.read().decode())
                    sch_list = [("HK", item["pl1$houseID"], item["title"], item["pl2$workcode"],
                                 datetime.fromtimestamp(item['availableDate'] // 1000).strftime('%Y-%m-%d %H:%M') if
                                 item['availableDate'] != 0 else "",
                                 datetime.fromtimestamp(item['expirationDate'] // 1000).strftime('%Y-%m-%d %H:%M') if
                                 item['expirationDate'] != 0 else "",
                                 item["pl2$channel_id"][0], item["approved"]) for item in data["entries"]]
                tw_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,title,approved,:channel_id,:workcode,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B" + "%7C".join(
                    house_id_combo) + "%7D&token=" + self.tok + "&account=FoxPlus%20Asia%20TW"
                with urllib.request.urlopen(tw_id_url) as url:
                    data = json.loads(url.read().decode())
                    tw_list = [("TW", item["pl1$houseID"], item["title"], item["pl2$workcode"],
                                datetime.fromtimestamp(item['availableDate'] // 1000).strftime('%Y-%m-%d %H:%M') if
                                item['availableDate'] != 0 else "",
                                datetime.fromtimestamp(item['expirationDate'] // 1000).strftime('%Y-%m-%d %H:%M') if
                                item['expirationDate'] != 0 else "",
                                item["pl2$channel_id"][0], item["approved"]) for item in data["entries"]]
                house_id_url = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:houseID,title,approved,:channel_id,:workcode,availableDate,expirationDate&byCustomValue=%7BhouseID%7D%7B" + "%7C".join(
                    house_id_combo) + "%7D&token=" + self.tok + "&account=FoxPlus%20Asia%20DP"
                with urllib.request.urlopen(house_id_url) as url:
                    data = json.loads(url.read().decode())
                    house_list = [("DP", item["pl1$houseID"], item["title"], item["pl2$workcode"],
                                   datetime.fromtimestamp(item['availableDate'] // 1000).strftime('%Y-%m-%d %H:%M') if
                                   item['availableDate'] != 0 else "",
                                   datetime.fromtimestamp(item['expirationDate'] // 1000).strftime('%Y-%m-%d %H:%M') if
                                   item['expirationDate'] != 0 else "",
                                   item["pl2$channel_id"][0], item["approved"]) for item in data["entries"]]
                TreePop(lib_list, sch_list, tw_list, house_list, house_id_combo)
                self.reset_idle_timer()
            except KeyError:
                self.mpx_connect_console()
        else:
            print ("Nothing performed for house_id_event")

    _detached = set()

    def _columns_searcher(self, P):
        #              originally a set            returns a tuple
        tree_child = {text:num for num, text in enumerate(self.tree.get_children())}
        children = {**self.filter_dict,**tree_child}
        self.filter_dict = {}
        self._brut_searcher(children, P)
        return True

    def _brut_searcher(self, children, query):
        try:
            if self.filter_var.get():
                for item_id,num in children.items():
                    text = self.tree.item(item_id)['values'][0]  # already contains the string-concatenation (over columns) of the row's values
                    if query.upper() in text.upper():
                        self.tree.reattach(item_id, '', self.filter_dict.get(item_id,num))
                    else:
                        self.filter_dict[item_id] = num
                        self.tree.detach(item_id)
                if not query:
                    self.display_message.config(text=f" Showing {len(self.tree.get_children())} results.")
                else:
                    self.display_message.config(
                        text=f" Filtering by keyword: \"{query}\" - {len(self.tree.get_children())} matches found.")
        except:
            pass

    def load_df(self): #TODO : Add scm_series
        try:
            shutil.copy(r"S:\OZ Migration\T&O\Overseer\database.pkl", "database.pkl")
            shutil.copy(r"S:\OZ Migration\T&O\Overseer\trailer.pkl", "trailer.pkl")
        except (FileNotFoundError, PermissionError):
            pass
        df = pd.read_pickle("database.pkl")
        trailer = pd.read_pickle("trailer.pkl")
        df = pd.concat([df, trailer],ignore_index=True,sort=False)
        df["title"] = df["title"].astype(str)
        global series_name,scm_name,movies_name,factual_name,scc_name, trailer_name
        series_name = set(df.loc[df["genre"].eq("series"),"title"].unique())
        movies_name = set(df.loc[df["genre"].eq("movies"),"title"].unique())
        factual_name = set(df.loc[df["genre"].eq("factual"),"title"].unique())
        scc_name = set(df.loc[df["genre"].eq("scc"),"title"].unique())
        scm_name = set(df.loc[df["genre"].eq("scm"),"title"].unique())
        trailer_name = set(df.loc[df["genre"].eq("trailer"),"title"].unique())
        #scm_name.update(set(df.loc[df["genre"].eq("movies"),"title"].unique()))

        df["pref"] = np.where(df["Box Number"].isin(settings.pref),10,df["pref"])
        df = df.sort_values(["genre", 'pref'], ascending=False)
        self.df = df

    def initial_start(self,*args):
        self.load_df()
        root.deiconify()
        sys_check.check_data_update_new()
        self.display_message.config(text=" Inventory loading completed.")
        self.display_icon.config(image=self.check_icon)
        self.sub_purge = PurgeSubtitle(self.revise_sub)
        self.sub_purge.entry.bind("<Button-3>", vision.rClicker)

    def search_prompt(self):
        self.search_button.flash()

    def popup(self,event):
        def configure_columns():
            column_filter = Frame(root, highlightbackground="black",highlightcolor="black",
                                  highlightthickness=1)
            full_header = ["SEA start", "SEA end", "TW start", "TW end", "PH start", "PH end", "HK start", "HK end",
                           "SG start", "SG end", "D&D time"]
            top_frame = Frame(column_filter, bg="midnight blue", highlightbackground="green", highlightcolor="green",
                              highlightthickness=1)
            top_frame.grid(row=0,column=0,columnspan=6,sticky="ew")
            Label(top_frame, text="Customize columns", bg="midnight blue", fg="white").pack(side=LEFT)
            for num, var in enumerate(full_var):
                e = Checkbutton(column_filter, text=full_header[num], variable=var, onvalue=1, offvalue=0)
                e.grid(row=(num//5)+1, column=num % 5, sticky="w")
                if full_header[num] in self.tree["displaycolumns"]:
                    var.set(1)

            def save_exit():
                result = sum(item.get() for item in full_var)
                if result == 5:
                    display_column = [full_header[i] for i in range(len(full_var)) if full_var[i].get() == 1]
                    self.tree["displaycolumns"] = ["Series", "Season", "Episode"] + display_column + ["version", "box_no",
                                                                                                 "house_no",
                                                                                                 "Lib approved",
                                                                                                 "DP schedule date"]
                    self.tree.column("#0", width=70, anchor="w")
                else:
                    column_label.config(text="You must select precisely 5 items to display.")

            save_button = Button(column_filter, text="SAVE", relief="groove",
                                 command=save_exit, bg=self.default_color, fg=self.font_color)
            quit_button = Button(column_filter, text="EXIT", relief="groove",
                                 command=column_filter.destroy, bg=self.default_color, fg=self.font_color)
            column_label = Label(column_filter, text="Select 5 items to display", font="Arial 9", fg="blue")
            save_button.grid(row=3, column=3, sticky="ew")
            quit_button.grid(row=3, column=4, sticky="ew")
            column_label.grid(row=3, column=0, columnspan=3,stick="ew")
            column_filter.place(relx=0.001,rely=0.18)

        def right_click_search_by_houseid(event): #FIXME
            user_input = StringDialog.ask_string("Query MPX by House/Box no.", "Input number below (1 only)")

            def id_check(string):
                global searching
                if string.isdigit and len(string) == 8:
                    if 10000000 < int(string) < 11000000:
                        searching = True
                        self.house_id_event([string])
                    else:
                        messagebox.showinfo("Error", "Invalid number entered.")
                elif string in self.df["Box Number"].values:
                    searching = True
                    #.df.loc[vision.df["Box Number"].eq(box_no),"sec_hdd"].iloc[0]
                    self.house_id_event([self.df.loc[self.df["Box Number"].eq(string),"House Number"].iloc[0]])
                else:
                    messagebox.showinfo("Error", "Invalid number entered.")

            if user_input:
                id_check(user_input)

        def right_click_copy_selected(event,id):
            try:
                root.clipboard_clear()
                copy_all = [str(self.tree.item(child)["values"][id]) for child in self.tree.selection()]
                root.clipboard_append(",".join(copy_all))
            except:
                pass

        def right_click_copy_all(event, id):
            try:
                root.clipboard_clear()
                copy_all = [str(self.tree.item(child)["values"][id]) for child in self.tree.get_children()]
                root.clipboard_append(",".join(copy_all))
            except:
                pass

        def right_click_show_missing(event):
            try:
                for num, item_id in enumerate(self.tree.get_children()):
                    text = self.tree.item(item_id)['values'][19]  # already contains the string-concatenation (over columns) of the row's values
                    if text:
                        self.miss_dict[item_id] = num
                        self.tree.detach(item_id)
                message = " Showing {} missing media.".format(len(self.tree.get_children()))
                self.display_message.config(text=message)
            except:
                pass

        def right_click_hide_tw():
            try:
                for num, item_id in enumerate(self.tree.get_children()):
                    text = self.tree.item(item_id)['values']  # already contains the string-concatenation (over columns) of the row's values
                    if text[5] and not any(s for s in (text[3],*text[7:12:2])):
                        self.miss_dict[item_id] = num
                        self.tree.detach(item_id)
                message = " Showing {} missing media.".format(len(self.tree.get_children()))
                self.display_message.config(text=message)
            except:
                pass

        def right_click_hide_tw_upload():
            exclude = ("Lady Commander","Super Funny Show","Witty Star","Travel Cheering Squad")
            try:
                for num, item_id in enumerate(self.tree.get_children()):
                    text = self.tree.item(item_id)['values']
                    if text[5] and any(s in text[0] for s in exclude):
                        self.miss_dict[item_id] = num
                        self.tree.detach(item_id)
                message = " Showing {} missing media.".format(len(self.tree.get_children()))
                self.display_message.config(text=message)
            except:
                pass

        def right_click_show_series(event, string):
            try:
                for num, item_id in enumerate(self.tree.get_children()):
                    if string not in self.tree.item(item_id)["tags"]:
                        self.miss_dict[item_id] = num
                        self.tree.detach(item_id)
            except:
                pass
            message = " Showing {} media from {}.".format(len(self.tree.get_children()), string.upper())
            self.display_message.config(text=message)

        def right_click_show_all(event):
            try:
                for item_id, i_r in self.miss_dict.items():
                    self.tree.reattach(item_id, '', i_r)
            except:
                pass
            message = " Showing all {} results.".format(len(self.tree.get_children()))
            self.display_message.config(text=message)
            self.miss_dict = {}

        def right_click_query(event):
            self.mpx_connect_console()

        def right_click_eject():
            if self.tree.selection():
                for item in self.tree.selection():
                    self.tree.detach(item)
            message = " Showing {} results.".format(len(self.tree.get_children()))
            self.display_message.config(text=message)

        def right_click_clear_all(event):
            self.tree.delete(*self.tree.get_children())
            self.missing_dict,self.filter_dict = {},{}
            self.display_message.config(text=" All results cleared.")

        def right_click_export(*args):
            f = filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=(("Excel file", "*.xlsx"), ("All Files", "*.*")))
            if f:
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Media report generated"
                    header = ["Series title", "Season", "Episode", "SEA start", "SEA end", "TW start", "TW end",
                              "PH start", "PH end", "HK start", "HK end", "SG start", "SG end", "Version", "Box ID",
                              "House ID", "Library approved", "DP schedule date", "Channel", "Image approved", "HDD"]
                    if not self.search_by_start:
                        header[17] = "DP end date"
                    ws.append(header)
                    ws.auto_filter.ref = 'A:U'
                    grey = colors.Color(rgb='444444')
                    my_fill = fills.PatternFill(patternType='solid', fgColor=grey)
                    thin_border = Border(left=Side(style='thin'),
                                         right=Side(style='thin'),
                                         top=Side(style='thin'),
                                         bottom=Side(style='thin'))
                    cols = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'H1', 'I1', 'J1', 'K1', 'L1', 'M1', 'N1', 'O1', 'P1', 'Q1', 'R1', 'S1', 'T1', 'U1']
                    for cell in cols:
                        ws[cell].fill = my_fill
                        ws[cell].font = Font(color='FFFFFF')
                        ws[cell].border = thin_border

                    for child in self.tree.get_children():
                        try:
                            box_no = self.tree.item(child)["values"][14]
                            hdd_no = vision.df.loc[vision.df["Box Number"].eq(box_no),"sec_hdd"].iloc[0]# maui_dict.get(self.tree.item(child)["values"][14])
                            img_app = self.tree.item(child)["text"]
                            result = self.tree.item(child)["values"]
                            del result[16:18]
                            del result[-1]
                            result.append(img_app if img_app else "")
                            result.append(hdd_no if hdd_no else "")
                            full_title = self.shorten_title.get(result[0])
                            if full_title:
                                result[0] = full_title
                            ws.append(result)
                        except:
                            continue

                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter  # Get the column name
                        for cell in col:
                            try:  # Necessary to avoid error on empty cells
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 1) * 1.05
                        ws.column_dimensions[column].width = adjusted_width
                    wb.save(filename=f)
                    filename = f.split("/")[-1]
                    message = " Exported as {}. (Click to view)".format(filename)
                    self.display_message.config(text=message)
                    self.display_message.bind("<Button-1>",lambda e: export_open(e,f))
                    self.display_icon.config(image=self.check_icon)
                except PermissionError:
                    message = " Failed to export. Please close the file or enable permission."
                    self.display_message.config(text=message)
                    self.display_icon.config(image=self.error_icon)
                except (IndexError,Exception) as e:
                    message = "Function MainGUI.popup.right_click_export: An unknown error occurred."
                    print (e)
                    error_log(message)

        def export_open(event,file_path):
            subprocess.Popen(file_path,shell=True)
            self.display_message.config(text="Launched exported file by excel.")
            self.display_message.unbind("<Button-1>")

        def right_click_set_pref(event):
            set_pref.clear_canvas()
            try:
                item = self.tree.identify('item', event.x, event.y)
                values = self.tree.item(item, "values")
                set_pref.draw_canvas(values,item,event.x,event.y)
            except:
                pass

        def right_click_sort_uploaded():
            try:
                for num,item_id in enumerate(self.tree.get_children()):
                    text = self.tree.item(item_id)['values'][18]  # already contains the string-concatenation (over columns) of the row's values
                    if text:
                        self.miss_dict[item_id] = num
                        self.tree.detach(item_id)
                message = " Showing {} media requiring upload.".format(len(self.tree.get_children()))
                self.display_message.config(text=message)
            except:
                pass

        def search_by_end():
            if self.search_by_start:
                self.search_by_start = False
                self.tree.heading("DP schedule date", text="DP end date", anchor='w')
                message = " Switched search mode to END date."
                self.display_message.config(text=message)
                self.display_icon.config(image=self.check_icon)
                self.tree["displaycolumns"] = (
                    "Series", "Season", "Episode", "SEA end", "TW end", "PH end", "HK end", "SG end",
                    "version", "box_no",
                    "house_no", "Lib approved", "DP schedule date")
            else:
                self.search_by_start = True
                self.tree.heading("DP schedule date", text="DP schedule date", anchor='w')
                message = " Switched search mode to START date."
                self.tree["displaycolumns"] = (
                    "Series", "Season", "Episode", "SEA start", "TW start", "PH start", "HK start", "SG start",
                    "version", "box_no",
                    "house_no", "Lib approved", "DP schedule date")
                self.display_message.config(text=message)
                self.display_icon.config(image=self.check_icon)

        def sort_season(s):
            try:
                for num,item_id in enumerate(self.tree.get_children()):
                    season = self.tree.item(item_id)['values'][1]  # already contains the string-concatenation (over columns) of the row's values
                    if str(season) != s:
                        self.miss_dict[item_id] = num
                        self.tree.detach(item_id)
                message = " Showing {} missing media.".format(len(self.tree.get_children()))
                self.display_message.config(text=message)
            except:
                pass

        top_commands = [
            ("Set version preference", lambda e=event: right_click_set_pref(event)),
            ('Query MPX', lambda e=event: right_click_query(event)),
            ('Query by number', lambda e=event: right_click_search_by_houseid(event)),
            ("Search by end date" if self.search_by_start else "Search by start date", search_by_end),
        ]
        house_copy_commands = [
            ('Copy HouseID', lambda e=event: right_click_copy_selected(event, 15)),
            ('Copy all HouseID', lambda e=event: right_click_copy_all(event, 15)),
        ]
        box_copy_commands = [
            ('Copy BoxID', lambda e=event: right_click_copy_selected(event, 14)),
            ('Copy all BoxID', lambda e=event: right_click_copy_all(event, 14)),
        ]
        show_commands = [
            ('Show missing only', lambda e=event: right_click_show_missing(event)),
            #('Show flagged only', lambda e=event: right_click_show_series(event, "flagged")),
            ('Show not uploaded',right_click_sort_uploaded),
            ('Hide Taiwan only', right_click_hide_tw),
            ("Hide Taiwan upload",right_click_hide_tw_upload),
            ('Show ALL', lambda e=event: right_click_show_all(event)),
            ("Clear ALL", lambda e=event: right_click_clear_all(event)),
        ]
        sort_commands = [
            ('Series', lambda e=event: right_click_show_series(event, "series")),
            ('SCC', lambda e=event: right_click_show_series(event, "scc")),
            ('Factual', lambda e=event: right_click_show_series(event, "factual")),
            ('Movies', lambda e=event: right_click_show_series(event, "movies")),
            ('SCM', lambda e=event: right_click_show_series(event, "scm")),
        ]
        additional_commands = [
            ('Customize columns', configure_columns),
            ('Delete selected item(s)', right_click_eject),
            ('Export result', lambda e=event: right_click_export(event)),
        ]
        command_list = [top_commands, house_copy_commands, box_copy_commands, show_commands]
        rmenu = Menu(None, tearoff=0, takefocus=0)
        sort_menu = Menu(None, tearoff=0, takefocus=0)
        season_menu = Menu(None, tearoff=0, takefocus=0)
        all_season = set()
        all_season.add("")
        for num, item_id in enumerate(self.tree.get_children()):
            season = self.tree.item(item_id)['values'][1]
            all_season.add(str(season))
        for k in command_list:
            for (txt, cmd) in k:
                rmenu.add_command(label=txt, command=cmd, compound="left")
            rmenu.add_separator()
        rmenu.add_cascade(label="Sort by", menu=sort_menu)
        sort_menu.add_cascade(label="Season",menu=season_menu)
        all_season = sorted(list(all_season))
        for sea_no in all_season:
            if sea_no:
                season_menu.add_command(label=sea_no, command=lambda e=sea_no:sort_season(e))
            else:
                season_menu.add_command(label="None", state="disabled")
        for (txt, cmd) in sort_commands:
            sort_menu.add_command(label=txt, command=cmd, compound="left")
        rmenu.add_separator()
        for (txt, cmd) in additional_commands:
            rmenu.add_command(label=txt, command=cmd, compound="left")
        rmenu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")

    def first_time_use(self):
        tutorial = Toplevel(root)
        tutorial.config(bd=0,bg="black")
        tutorial.geometry("1236x670+{}+{}".format(root.winfo_x() + 1, root.winfo_y() - 2))
        tutorial.overrideredirect(True)
        tutorial.grab_set()
        page_no = IntVar()
        page_no.set(1)
        tutorial_pic = PhotoImage(file="tk/images/tutor_page{}.png".format(page_no.get()))

        def next_page():
            page_no.set(page_no.get() + 1)
            tutorial_pic = PhotoImage(file="tk/images/tutor_page{}.png".format(page_no.get()))
            l.img = tutorial_pic
            l.config(image=tutorial_pic)
            if page_no.get() == 3:
                tutorial.geometry("1220x723+{}+{}".format(root.winfo_x() + 10, root.winfo_y() + 20))
                l.config(command=tutorial.destroy)

        l = Button(tutorial, image=tutorial_pic, relief=SUNKEN, borderwidth=0,
                   command=next_page)
        l.img = tutorial_pic
        l.pack()

    def create_meta_page(self):
        self.meta_top_frame = Frame(self.metadata_page)
        self.meta_top_frame.pack()
        self.meta_bottom_frame = Frame(self.metadata_page)
        self.meta_bottom_frame.pack()
        self.meta_header = ("No.", "Title", "Series title", "Season", "Ep", "On-Air ID", "Channel", "Eng desc",
                       "Zh-hk Title", "Zh-tw Title", "Zh-Title", "Zh-hk Desc", "Zh-tw Desc", "Zh-Desc",
                       "Genre", "Year", "Cast count", "Cast")
        self.meta_width = (40, 190, 180, 45, 45, 70, 60, 70, 70, 70, 60, 70, 70, 60, 55, 45, 45, 100)

        self.sub_header = ("No.", "Title", "Series title", "Season", "Ep", "House ID", "Channel", "English sub",
                      "Chinese HK sub", "Chinese TW sub", "Simp. Chinese Sub", "Count", "", "",
                      "", "")
        self.sub_width = (40, 190, 180, 45, 45, 70, 60, 130, 130, 130, 130, 50, 0, 0, 0, 0)
        self.meta_tree = ttk.Treeview(self.meta_bottom_frame, selectmode='browse', height=26)
        meta_vsb = ttk.Scrollbar(self.meta_bottom_frame, orient="vertical",
                                 command=self.meta_tree.yview)  ## CREATE SCROLLBAR WIDGET AND ATTACH TO treeview.yview

        self.meta_tree["columns"] = self.meta_header
        self.meta_tree["displaycolumns"] = (
        "No.", "Title", "Series title", "Season", "Ep", "On-Air ID", "Channel", "Eng desc",
        "Zh-hk Title", "Zh-tw Title", "Zh-Title", "Zh-hk Desc", "Zh-tw Desc", "Zh-Desc",
        "Genre", "Year")
        self.meta_tree['show'] = 'headings'

        for i in range(len(self.meta_header)):
            self.meta_tree.column(self.meta_header[i], width=self.meta_width[i], anchor="w", stretch=YES)
            self.meta_tree.heading(self.meta_header[i], text=self.meta_header[i], anchor='w')

        meta_image = PhotoImage(file="icons/meta.png")
        self.meta_button = Button(self.meta_top_frame, text=" Start metadata check ", image=meta_image,
                                  compound="left",bg=self.default_color,fg=self.font_color,
                                  font=("Arial", 14), command=lambda: self.start_thread(self.get_metadata))
        self.meta_button.grid(row=0, column=0, padx=1)
        self.meta_button.img = meta_image
        meta_label = Label(self.meta_top_frame, text="?", font=("Arial", 10), width=2, anchor="w", underline=0)
        meta_label.grid(row=0, column=1)
        CreateToolTip(meta_label,"This will check the most recent 10000 media on PA feed. Media will all metadata"
                                 " fulfilled with be skipped. Anything else will be shown below. You can right click"
                                 " and export the result for later processing.")
        chi_sub_image = PhotoImage(file="icons/subtitle.png")
        self.chi_sub_button = Button(self.meta_top_frame, text=" Start subtitle check ", image=chi_sub_image,
                                     compound="left",bg=self.default_color,fg=self.font_color,
                                     font=("Arial", 14), command=lambda: self.start_thread(self.chi_sub_check))
        self.chi_sub_button.grid(row=0, column=2, padx=1)
        chi_sub_label = Label(self.meta_top_frame, text="?", font=("Arial", 10), width=2, anchor="w", underline=0)
        chi_sub_label.grid(row=0, column=3)
        CreateToolTip(chi_sub_label,"This will check the most recent 1000 media on DP. "
                                    "Sports and test media will be skipped. "
                                    "Any media with 2 or less subtitles will be shown below. "
                                    "You can double click on each media to view detailed info.")
        self.chi_sub_button.img = chi_sub_image
        self.meta_info = tkst.ScrolledText(self.meta_top_frame, wrap=WORD, width=100, height=3)
        self.meta_info.config(font=("Arial", 8))
        self.meta_info.grid(row=0, column=4, padx=20)
        self.meta_info.insert(END, "Welcome to Data Services console. "
                                   "Click a button on left to start corresponding checking.\n"
                                   "You can continue to perform other tasks while waiting for result.")
        self.meta_info.bind("<Key>", lambda e: vision.txtEvent(e))
        self.meta_info.yview(END)
        self.meta_tree.tag_configure("flagged", background="firebrick1", foreground="white")
        self.meta_tree.pack(side='left')
        self.meta_tree.bind("<Button-3>", self.meta_popup)
        meta_vsb.pack(side='right', fill='y')

    def get_metadata(self,a, b):
        self.meta_tree.delete(*self.meta_tree.get_children())
        meta_width = (40, 190, 180, 45, 45, 70, 60, 70, 70, 70, 60, 70, 70, 60, 55, 45, 45,
                      100)  # (40, 170, 120, 45, 45, 70, 60, 70, 70, 70, 60, 70, 70, 60, 55, 45)
        self.meta_tree["displaycolumns"] = ()
        self.meta_tree["columns"] = self.meta_header
        self.meta_tree["displaycolumns"] = (
        "No.", "Title", "Series title", "Season", "Ep", "On-Air ID", "Channel", "Eng desc",
        "Zh-hk Title", "Zh-tw Title", "Zh-Title", "Zh-hk Desc", "Zh-tw Desc", "Zh-Desc",
        "Genre", "Year")
        for i in range(len(self.meta_header)):
            self.meta_tree.column(self.meta_header[i], width=self.meta_width[i], anchor="w", stretch=YES)
            self.meta_tree.heading(self.meta_header[i], text=self.meta_header[i], anchor='w')
        self.meta_button.config(state=DISABLED)
        self.chi_sub_button.config(state=DISABLED)
        self.meta_info.insert(END, "\nCalculating total number of items...")
        self.meta_info.yview(END)
        error_count = 0
        total_count = 0
        metadata_url = "https://feed.entertainment.tv.theplatform.com/f/NIi3EC/foxplus_pa?byRequiredAvailability=media&byMediaAvailabilityState=available&pretty=true&fields=title&byMediaAvailabilityTags=fam|ffm|ffm_catchup|fmp_fam_catchup|fmp|fox|foxcrime|fx|ngc|ngp|ngw|scci|scc_ph|scc_sg|scm|scm_hk|scm_sg|scm_ph|starworld&form=cjson&sort=added|desc&range=1-8000"
        with urllib.request.urlopen(metadata_url) as url:
            data = json.loads(url.read().decode())
            try:
                total_entry = len(data["entries"])
            except KeyError:
                messagebox.showerror("Error", "Fail to retrieve data from PA feed.")
                self.meta_info.insert(END, "\nOperation aborted. Please try again. ")
                self.meta_button.config(state=NORMAL)
                self.chi_sub_button.config(state=NORMAL)
                return
        self.meta_info.insert(END, "\nTotal number of items found: {}".format(total_entry))
        self.meta_info.insert(END, "\nFetching data from PA feed (This will take a while)... ")
        self.meta_info.yview(END)
        metadata_url = "https://feed.entertainment.tv.theplatform.com/f/NIi3EC/foxplus_pa?byRequiredAvailability=media&byMediaAvailabilityState=available&pretty=true&byMediaAvailabilityTags=fam|ffm|ffm_catchup|fmp_fam_catchup|fmp|fox|foxcrime|fx|ngc|ngp|ngw|scci|scc_ph|scc_sg|scm|scm_hk|scm_sg|scm_ph|starworld&form=cjson&sort=added|desc&range=1-{}".format(
            total_entry)  #
        with urllib.request.urlopen(metadata_url) as url:
            try:
                data = json.loads(url.read().decode())
            except:
                messagebox.showerror("Error", "Fail to retrieve data from PA feed.")
                self.meta_info.insert(END, "\nOperation aborted. Please try again. ")
                self.meta_button.config(state=NORMAL)
                self.chi_sub_button.config(state=NORMAL)
                return
        for item in data["entries"]:
            guid = item["guid"]
            title = item["title"]
            e_des = item["description"]
            c_title = item["secondaryTitleLocalized"]
            c_title2 = item["sortTitleLocalized"]
            c_des = item["descriptionLocalized"]
            c_des2 = item["longDescriptionLocalized"]
            tags = item["tags"]
            eng_title_check_index = 0
            eng_des_check_index = 0
            chinese_title_check_index = 0
            chinese_des_check_index = 0

            if item["credits"]:
                cast_count = len(item["credits"])
                cast = ["{}: {}".format(type["creditType"].capitalize(), type["personName"]) for type in
                        item["credits"]]
            else:
                cast_count = 0
                cast = ""

            if "pl1$series-Title" in item:
                series_title = item["pl1$series-Title"]
            else:
                series_title = ""

            if "tvSeasonEpisodeNumber" in item:
                ep_no = item["tvSeasonEpisodeNumber"] if item["tvSeasonEpisodeNumber"] is not None else ""
            else:
                ep_no = ""

            if "tvSeasonNumber" in item:
                season_no = item["tvSeasonNumber"] if item["tvSeasonNumber"] is not None else ""
            else:
                season_no = ""

            if title == "":
                pass
            else:
                eng_title_check_index = eng_title_check_index + 1

            if e_des == "":
                e_des = "Missing"
                pass
            else:
                e_des = "OK"
                eng_des_check_index = eng_des_check_index + 1

            # Begin Chinese title check###
            if "zh-hant-hk" in c_title:
                zh_hant_hk_title = "OK"
                chinese_title_check_index = chinese_title_check_index + 1
            elif "zh-hant-hk" in c_title2:
                zh_hant_hk_title = "OK"
                chinese_title_check_index = chinese_title_check_index + 1
            else:
                zh_hant_hk_title = "Missing"

            if "zh-hant" in c_title:
                zh_hant_title = "OK"
                chinese_title_check_index = chinese_title_check_index + 1
            elif "zh-hant" in c_title2:
                zh_hant_title = "OK"
                chinese_title_check_index = chinese_title_check_index + 1
            else:
                zh_hant_title = "Missing"

            if "zh-hans" in c_title:
                zh_hans_title = "OK"
                chinese_title_check_index = chinese_title_check_index + 1
            elif "zh-hans" in c_title2:
                zh_hans_title = "OK"
                chinese_title_check_index = chinese_title_check_index + 1
            else:
                zh_hans_title = "Missing"

            ###Begin Chinese description check###
            if "zh-hant-hk" in c_des:
                zh_hant_hk_des = "OK"
                chinese_des_check_index = chinese_des_check_index + 1
            elif "zh-hant-hk" in c_des2:
                zh_hant_hk_des = "OK"
                chinese_des_check_index = chinese_des_check_index + 1
            else:
                zh_hant_hk_des = "Missing"

            if "zh-hant" in c_des:
                zh_hant_des = "OK"
                chinese_des_check_index = chinese_des_check_index + 1
            elif "zh-hant" in c_des2:
                zh_hant_des = "OK"
                chinese_des_check_index = chinese_des_check_index + 1
            else:
                zh_hant_des = "Missing"

            if "zh-hans" in c_des:
                zh_hans_des = "OK"
                chinese_des_check_index = chinese_des_check_index + 1
            elif "zh-hans" in c_des2:
                zh_hans_des = "OK"
                chinese_des_check_index = chinese_des_check_index + 1
            else:
                zh_hans_des = "Missing"

            if "pl1$foxplay-Genre" in item:
                genre = item["pl1$foxplay-Genre"]
            else:
                genre = ""

            if "year" in item:
                production_year = item["year"] if item["year"] is not None else ""
            else:
                production_year = ""

            total_count = total_count + 1
            channel = list(set(i["scheme"] for i in tags))
            try:
                channel = channel[0]
            except:
                pass
            # channel = str(tags).strip("[]").strip("{}").replace("'scheme': ", "").replace("'title':", "")
            # channel = channel.split(",")[0]
            if chinese_title_check_index + chinese_des_check_index + eng_des_check_index + eng_title_check_index == 8:
                if cast_count == 0 and "ng" in channel:  # pass
                    pass  # delete below
                elif cast_count == 0:
                    result = [total_count, str(title).replace(",", ""), series_title, season_no, ep_no, guid[5:],
                              channel, e_des, zh_hant_hk_title, zh_hant_title, zh_hans_title,
                              # channel.replace("'", "")
                              zh_hant_hk_des, zh_hant_des, zh_hans_des, genre, production_year, ", ".join(cast)]
                    error_count = error_count + 1
                    self.meta_tree.insert('', END, text="", value=result)
            else:
                result = [total_count, str(title).replace(",", ""), series_title, season_no, ep_no, guid[5:],
                          channel, e_des, zh_hant_hk_title, zh_hant_title, zh_hans_title,  # channel.replace("'", "")
                          zh_hant_hk_des, zh_hant_des, zh_hans_des, genre, production_year, ", ".join(cast)]
                error_count = error_count + 1
                self.meta_tree.insert('', END, text="", value=result)
        self.meta_info.insert(END, "\nData retrieve success - {} results found.".format(len(self.meta_tree.get_children())))
        self.meta_info.yview(END)
        pop = PopUpMessage(1,"Task completed","Metadata checking has completed.","(CLICK TO VIEW)")
        pop.can.bind("<Button-1>",pop.jump_tab)
        self.meta_button.config(state=NORMAL)
        self.chi_sub_button.config(state=NORMAL)
        # meta_tree.configure(yscrollcommand=meta_vsb.set)
        self.meta_tree.bind("<Double-1>", self.meta_double_click)

    def meta_popup(self,event):
        item = self.meta_tree.identify('item', event.x, event.y)
        item_text = self.meta_tree.item(item, "values")

        def metadata_export():
            f = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=(("Excel file", "*.xlsx"), ("All Files", "*.*")))
            if f and self.meta_tree["columns"] == self.meta_header:
                wb = Workbook()
                ws = wb.active
                ws.title = "Metadata from PA feed"
                ws_headers1 = (
                "No.", "Title", "Series title", "Season", "Episode", "On-Air ID", "Channel", "English description",
                "Chinese title", "", "",
                "Chinese description", "", "",
                "Genre", "Year", "Cast")
                ws_headers2 = ("", "", "", "", "", "", "", "",
                               "香港繁體", "台灣繁體", "簡體",
                               "香港繁體", "台灣繁體", "簡體",
                               "", "", "")

                ws.append(ws_headers1)
                ws.append(ws_headers2)
                ws.merge_cells("A1:A2")
                ws['A1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("B1:B2")
                ws['B1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("C1:C2")
                ws['C1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("D1:D2")
                ws['D1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("E1:E2")
                ws['E1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("F1:F2")
                ws['F1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("G1:G2")
                ws['G1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("H1:H2")
                ws['H1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("I1:K1")
                ws['I1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("L1:N1")
                ws['L1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("O1:O2")
                ws['O1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("P1:P2")
                ws['P1'].alignment = Alignment(vertical="center", horizontal="center")
                ws.merge_cells("Q1:Q2")

                def set_border(worksheet, cell_range):
                    rows = list(worksheet.iter_rows(cell_range))
                    side = Side(border_style='thin', color="FF000000")
                    rows = list(
                        rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
                    max_y = len(rows) - 1  # index of the last row
                    for pos_y, cells in enumerate(rows):
                        max_x = len(cells) - 1  # index of the last cell
                        for pos_x, cell in enumerate(cells):
                            border = Border(
                                left=cell.border.left,
                                right=cell.border.right,
                                top=cell.border.top,
                                bottom=cell.border.bottom
                            )
                            if pos_x == 0:
                                border.left = side
                            if pos_x == max_x:
                                border.right = side
                            if pos_y == 0:
                                border.top = side
                            if pos_y == max_y:
                                border.bottom = side

                            # set new border only if it's one of the edge cells
                            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                                cell.border = border

                set_border(ws, "A1:Q2")

                for child in self.meta_tree.get_children():
                    ws.append(self.meta_tree.item(child)["values"])

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # Get the column name
                    for cell in col:
                        try:  # Necessary to avoid error on empty cells
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 1) * 1.05
                    ws.column_dimensions[column].width = adjusted_width

                ws.column_dimensions["A"].width = 5
                ws.column_dimensions["B"].width = 40
                ws.column_dimensions["C"].width = 25
                ws.column_dimensions["F"].width = 14
                ws.column_dimensions["I"].width = 10
                ws.column_dimensions["J"].width = 10
                ws.column_dimensions["K"].width = 10
                ws.column_dimensions["L"].width = 10
                ws.column_dimensions["M"].width = 10
                ws.column_dimensions["N"].width = 10
                ws.column_dimensions["O"].width = 29
                ws.column_dimensions["P"].width = 6
                ws.column_dimensions["Q"].width = 100
                wb.save(filename=f)
            elif f and self.meta_tree["columns"] == self.sub_header:
                wb = Workbook()
                ws = wb.active
                ws.title = "Subtitle availability DP"
                ws_headers = ("No.", "Title", "Series title", "Season", "Ep", "House ID", "Channel", "English sub",
                              "Chinese HK sub", "Chinese TW sub", "Simp. Chinese Sub", "Count")
                ws.append(ws_headers)
                for child in self.meta_tree.get_children():
                    ws.append(self.meta_tree.item(child)["values"])
                wb.save(filename=f)

        def RightClick_meta_house_id():
            root.clipboard_clear()
            root.clipboard_append(item_text[5])

        if len(self.meta_tree.get_children()) != 0:
            rmenu = Menu(None, tearoff=0, takefocus=0)
            if self.meta_tree["columns"] == self.sub_header:
                rmenu.add_command(label="Copy House ID", command=RightClick_meta_house_id, compound="left")
            rmenu.add_command(label="Export result", command=metadata_export, compound="left")
            rmenu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")

    def meta_double_click(self, event):
        global iid
        iid = self.meta_tree.identify('item', event.x, event.y)
        item_text = self.meta_tree.item(iid, "values")
        if self.meta_tree["columns"] == self.sub_header:
            try:
                self.house_id_event([item_text[5]])
            except:
                pass

    def chi_sub_check(self,a, b):
        if self.tok == "":
            self.searching = True
            self.mpx_connect_console()
            while True:
                time.sleep(0.2)
                if self.tok != "":
                    self.chi_sub_check("", "")
                    break
            return
        self.meta_tree.delete(*self.meta_tree.get_children())
        self.meta_tree["displaycolumns"] = ()
        self.meta_tree["columns"] = self.sub_header
        self.meta_tree["displaycolumns"] = (
        "No.", "Title", "Series title", "Season", "Ep", "House ID", "Channel", "English sub",
        "Chinese HK sub", "Chinese TW sub", "Simp. Chinese Sub", "Count")
        for i in range(len(self.sub_header)):
            self.meta_tree.column(self.sub_header[i], width=self.sub_width[i], anchor="w", stretch=YES)
            self.meta_tree.heading(self.sub_header[i], text=self.sub_header[i], anchor='w')
        self.meta_button.config(state=DISABLED)
        self.chi_sub_button.config(state=DISABLED)
        self.meta_info.insert(END, "\nFetching recent 1000 media on DP...")
        self.meta_info.yview(END)
        self.searching = True
        media_item = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:season,:episodeNumber,:houseID,title,:channel_id,:series-Title,content&range=1-{}&sort=added|desc&token={}&account=FoxPlus%20Asia%20DP".format(
            1000, self.tok)
        total_count, sports_count = 0, 0
        with urllib.request.urlopen(media_item) as url:
            data = json.loads(url.read().decode())
        try:
            for item in data['entries']:
                if "pl3$channel_id" in item:
                    channel_id = ",".join(item["pl3$channel_id"])
                elif 'pl2$channel_id' in item:
                    channel_id = ",".join(item["pl2$channel_id"])
                elif 'pl1$channel_id' in item:
                    channel_id = ",".join(item["pl1$channel_id"])
                else:
                    channel_id = ""
                title = item["title"]
                if "pl3$series-Title" in item:
                    series_title = item["pl3$series-Title"]
                elif "pl2$series-Title" in item:
                    series_title = item["pl2$series-Title"]
                elif "pl1$series-Title" in item:
                    series_title = item["pl1$series-Title"]
                else:
                    series_title = ""
                if "pl1$houseID" in item:
                    house_id = item["pl1$houseID"]
                else:
                    house_id = ""
                if "pl1$episodeNumber" in item:
                    ep = item["pl1$episodeNumber"]
                else:
                    ep = ""
                if "pl1$season" in item:
                    season = item["pl1$season"]
                else:
                    season = ""
                if "foxsports" in channel_id or "test" in channel_id:
                    sports_count += 1
                    pass
                else:
                    dfxp = ["", "", "", ""]
                    for media_item in item["content"]:
                        if media_item["url"].upper().endswith("EN.DFXP"):
                            dfxp[0] = media_item["url"].split("/")[-1]
                        elif media_item["url"].endswith("91.DFXP"):
                            dfxp[3] = media_item["url"].split("/")[-1]
                        elif media_item["url"].upper().endswith("CT.DFXP"):
                            dfxp[1] = media_item["url"].split("/")[-1]
                        elif media_item["url"].upper().endswith("ZH.DFXP") or media_item["url"].endswith("TW.DFXP"):
                            dfxp[2] = media_item["url"].split("/")[-1]
                        else:
                            pass
                    total_sub = [x for x in dfxp if x]
                    if len(total_sub) >= 3 or dfxp[1] != "" or dfxp[2] != "":
                        pass
                    else:
                        total_count += 1
                        self.meta_tree.insert("", END, text="", value=(
                        total_count, title, series_title, season, ep, house_id, channel_id, dfxp[0], dfxp[1], dfxp[2],
                        dfxp[3], len(total_sub)))
        except KeyError:
            self.mpx_connect_console()
            self.meta_button.config(state=NORMAL)
            self.chi_sub_button.config(state=NORMAL)
            self.searching = False
        self.meta_info.insert(END, "\n{} sports or test item skipped.".format(sports_count))
        self.meta_info.insert(END, "\nData retrieve success - {} results found without Chinese subtitle.".format(
            len(self.meta_tree.get_children())))
        self.meta_info.yview(END)
        self.meta_button.config(state=NORMAL)
        self.chi_sub_button.config(state=NORMAL)
        pop = PopUpMessage(1,"Task completed","Subtitle checking has completed.","(CLICK TO VIEW)") #"Task completed" #message #"(CLICK TO VIEW)"
        pop.can.bind("<Button-1>", pop.jump_tab)
        self.reset_idle_timer()
        self.meta_tree.bind("<Double-1>", self.meta_double_click)

class ColorChooser(Frame):
    def __init__(self,master=None, **kwargs):
        super().__init__(master,**kwargs)
        self.check_icon = PhotoImage(file=r"icons\check.png")
        self.delete_icon = PhotoImage(file=r"icons\delete.png")
        self.revert_icon = PhotoImage(file=r"icons\revert.png")
        top_frame = Frame(self, bg="midnight blue", highlightbackground="green",
                          highlightcolor="green", highlightthickness=1, width=200)
        top_frame.grid(row=0, column=0, columnspan=6, sticky="ew")

        Label(top_frame, text="Customize colors", bg="midnight blue", fg="white", anchor="w"
                 ).grid(row=0,column=0,sticky="ew", columnspan=3)
        Label(top_frame, text="Background color: ", anchor="e", bg="white").grid(row=1,column=0,sticky="nesw")
        Label(top_frame, text="Font color: ", anchor="e", bg="white").grid(row=2,column=0,sticky="nesw")

        self.buttons = [Button(top_frame, image=i, command=j,
                               bg="midnight blue",relief="flat",anchor="e")
                        for i, j in zip((self.check_icon, self.delete_icon, self.revert_icon),
                                        (self.save_settings, self.destroy, self.revert_settings))]

        for num, i in enumerate(self.buttons,2):
            i.grid(row=3, column=num, sticky="ew")

        self.test_message = Label(top_frame, bg=settings.default_color,
                                  fg=settings.font_color, text="**This is sample message**")
        self.test_message.grid(row=3,column=0, sticky="ew")
        self.bg = Button(top_frame, bg=settings.default_color, relief="groove", width=10)
        self.bg.config(command=lambda: self.choose_color(self.bg))
        self.bg.grid(row=1,column=1, sticky="ew", columnspan=4)
        self.fg = Button(top_frame, bg=settings.font_color, relief="groove", width=10)
        self.fg.config(command=lambda: self.choose_color(self.fg))
        self.fg.grid(row=2,column=1, sticky="ew", columnspan=4)
        self.place(relx=0.001,rely=0.001)

    def choose_color(self, widget):
        result = colorchooser.askcolor()
        if result:
            widget.config(bg=result[-1])
            self.test_message.config(bg=self.bg["background"],fg=self.fg["background"])

    def save_settings(self):
        settings.change_bg(self.bg["background"])
        settings.change_fg(self.fg["background"])
        try:
            with open("settings.pkl", "wb") as f:
                pickle.dump(settings, f)
            vision.display_message.config(text=f" Color changes will apply after restart.")
            vision.display_icon.config(image=vision.check_icon)
        except PermissionError:
            message = " Failed to save settings. Please close the file or enable permission."
            vision.display_message.config(text=message)
            vision.display_icon.config(image=vision.error_icon)
        self.destroy()

    def revert_settings(self):
        self.bg.config(bg="DeepSkyBlue4")
        self.fg.config(bg="white")
        self.test_message.config(fg="white",bg="DeepSkyBlue4")

class CreateToolTip:
    def __init__(self, widget, text='widget info'):
        self.waittime = 100 #500     #miliseconds
        self.wraplength = 180   #pixels
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.enter)
        self.widget.bind("<Leave>", self.leave)
        self.widget.bind("<ButtonPress>", self.leave)
        self.color = settings.default_color
        self.id = None
        self.tw = None

    def enter(self, event=None):
        try:
            if self.widget["text"] == "?":
                pass
            elif not "entry" in str(self.widget):
                if not self.widget["relief"] == "sunken":
                    self.widget.config(bg=self.color, relief="groove")
        except AttributeError:
            pass
        self.schedule()

    def leave(self, event=None):
        try:
            if self.widget["text"] == "?":
                pass
            elif not "entry" in str(self.widget):
                self.widget.config(bg=self.color)
                if not self.widget["relief"] == "sunken":
                    self.widget["relief"] = "flat"
        except AttributeError:
            pass
        self.unschedule()
        self.hidetip()

    def schedule(self):
        self.unschedule()
        self.id = self.widget.after(self.waittime, self.showtip)

    def unschedule(self):
        id = self.id
        self.id = None
        if id:
            self.widget.after_cancel(id)

    def showtip(self, event=None):
        x = y = 0
        x, y, cx, cy = self.widget.bbox("insert")
        if "entry" in str(self.widget):
            x += self.widget.winfo_rootx() - 20
            y += self.widget.winfo_rooty() + 20
        else:
            x += self.widget.winfo_rootx() + 25
            y += self.widget.winfo_rooty() + 40
        # creates a toplevel window
        self.tw = Toplevel(self.widget)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = Label(self.tw, text=self.text, justify='left',
                       background="#ffffff", relief='solid', borderwidth=1,
                       wraplength = self.wraplength)
        label.pack(ipadx=1)

    def hidetip(self):
        tw = self.tw
        self.tw= None
        if tw:
            tw.destroy()

class TreeTitle(CreateToolTip):
    def __init__(self, widget, text='widget info'):
        super().__init__(widget, text=text)
        self.widget.bind("<Motion>", self.hover)
        self.current_iid = None
        self.widget.unbind("<Enter>")
        self.widget.bind("<Leave>", self.leave)
        self.widget.bind("<ButtonPress>", self.leave)

    def hover(self,event): #ver 13, box 14, house 15, onair 21
        try:
            iid = self.widget.identify_row(event.y)
            if self.current_iid == iid:
                return
            else:
                self.leave(event=None)
            name = self.widget.item(iid)["values"][0]
            self.text = vision.shorten_title.get(name,"")
            if self.text and not settings.analyze_mode:
                self.schedule()
                self.current_iid = iid
            elif settings.analyze_mode:
                if self.text:
                    self.text= f"Title: {self.text}\n"
                result = self.widget.item(iid)["values"]
                ver, box, onair = result[13], result[14], result[21]
                if result[18]:
                    if result[18]=="True" and result[19]:
                        self.text+="Media has been uploaded and scheduled."
                    elif result[18]=="True":
                        self.text+="Media has been uploaded and approved. Please schedule accordingly."
                    else:
                        self.text+="Media has been uploaded and pending approval.\n" \
                                   "The files are either being published or subtitles are missing.\n" \
                                   "If pending, confirm with MSCO team on approval status."
                elif not onair or onair=="nan":
                    self.text+="On-air ID is missing for this media.\n" \
                              "Please input in FOX+ schedule and request data update."
                elif not ver:
                    self.text+="On-air version is missing. Please wait for media creation.\n" \
                              "If pending, confirm the correct OTT version has been selected in On-Air."
                elif not box:
                    self.text+="Media is not yet logged by MSCO team.\n" \
                              "This is common for D&D titles - please inform MSCO otherwise."
                else:
                    hdd = result[17]
                    if hdd:
                        self.text+="Media is OK and ready for upload.\n" \
                                  "If pending, check whether other versions are available.\n" \
                                  "(Right click and select another box number)"
                    else:
                        self.text+="Mezzanine file has not been or is still being produced.\n" \
                                  "This is common for titles require dubbing or TW uploaded titles.\n" \
                                  "Please check wih MSCO on status on mezzanine production."
                self.schedule()
                self.current_iid = iid
            else:
                self.current_iid = None
        except (TclError,IndexError):
            pass

    def showtip(self, event=None):
        x = y = 0
        try:
            x, y, cx, cy = self.widget.bbox(self.current_iid,column="Series" if not settings.analyze_mode else "box_no")
        except ValueError:
            return
        x += self.widget.winfo_rootx() + (25 if not settings.analyze_mode else 0)
        y += self.widget.winfo_rooty() + (30 if not settings.analyze_mode else 0)
        # creates a toplevel window
        self.tw = Toplevel(self.widget)
        # Leaves only the label and removes the app window
        self.tw.wm_overrideredirect(True)
        self.tw.wm_geometry("+%d+%d" % (x, y))
        label = Label(self.tw, text=self.text, justify='left',
                       background="light yellow", relief='solid', borderwidth=1,
                       wraplength = len(self.text)*8)
        label.pack(ipadx=1)

    def leave(self, event=None):
        try:
            self.unschedule()
            self.hidetip()
        except AttributeError:
            pass

class DataInfo:
    def __init__(self,master):
        self.top = Canvas(master)
        try:
            shutil.copy(r"S:\OZ Migration\T&O\Overseer\data_info.pkl", "data_info.pkl")
        except (FileNotFoundError, PermissionError):
            pass
        data_info = pd.read_pickle("data_info.pkl")
        top_frame = Frame(self.top,background="midnight blue",highlightbackground="green", highlightcolor="green", highlightthickness=1)
        top_frame.pack(fill=X)
        Label(top_frame,text="Data information",bg="midnight blue",fg="white").pack(side=LEFT)
        low_frame = Frame(self.top,highlightbackground="green", highlightcolor="green", highlightthickness=1)
        low_frame.pack(fill=BOTH)
        self.temp_tree = ttk.Treeview(low_frame, selectmode='browse', height=11)
        self.temp_tree.pack(side='left')
        header = ("Data type","Version/date","Link")
        width = (120,460,200)
        self.temp_tree["columns"] = header
        self.temp_tree['show'] = 'headings'
        self.temp_tree["displaycolumns"] = ("Data type","Version/date")
        for i in range(len(header)):
            self.temp_tree.column(header[i], width=width[i], anchor="w")
            self.temp_tree.heading(header[i], text=header[i], anchor='w')
        for _, i in data_info.iterrows():
            self.temp_tree.insert("", END, text="", values=(i[0], i[1].split("\\")[-1],i[1]))

        self.temp_tree.bind("<Double-1>", self.go_to_excel)

    def go_to_excel(self, event=None):
        iid = self.temp_tree.identify_row(event.y)
        file = self.temp_tree.item(iid)["values"][2]
        if not "Last" in file:
            try:
                os.rename(file,file)
                subprocess.Popen(file,shell=True) #explorer /select,
            except PermissionError:
                subprocess.Popen(f'explorer /select,"{file}"')  # explorer /select
                return
            except Exception as e:
                pass

    def show_data(self):
        self.top.place(relx=0.001, rely=0.1455) #0.178

    def hide_data(self):
        self.top.destroy()

class SubCheck: #sub_check
    def __init__(self,link):
        headstop = "begin="
        chi_check = re.compile(
            "[^\u0000-\u0007\u0009-\u05C0\u200B-\u2010\u2012-\u202E\u2030-\u205E\u2100-\u214F\u2200-\u22FF\u2500\u25CB\u2E80-\u2FD5\u3000-\u303D\u30FB\u3100-\u312F\u3400-\u4DBF\u4E00-\u9FFF\uF900-\uFAD0\uFE30-\uFE6B\uFEFF\uFF01-\uFFEF\uD844\uDCC1\u210C1\u25683\u282E2\u29D5A\u3040-\u309F]")
        eng_check = re.compile(
            "[^\u0000-\u0007\u0009-\u05C0\u200B-\u2010\u2012-\u202E\u2030-\u205E\u20A0-\u20BF\u2100-\u214F\uFEFF]")
        bur_check = re.compile(
            "[^\u0000-\u0007\u0009-\u05C0\u1000-\u109F\u200B-\u2010\u2012-\u202E\u2030-\u205E\u2200-\u22FF\u2600-\u26FF\u3000\uFEFF]")
        msid_check = re.compile("[^\u0000-\u0007\u0009-\u05C0\u200B-\u2010\u2012-\u202E\u2030-\u205E\u2100-\u214F\u3000\uFEFF]")
        thai_check = re.compile("[^\u0000-\u0007\u0009-\u05C0\u200B-\u2010\u2012-\u202E\u2030-\u205E\u20AC\uFEFF\u3000ก-๛]")
        ara_check = re.compile(
            "[^\u0000-\u0007\u0009-\u05C0\u200B-\u2010\u2012-\u202E\u2030-\u205E\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF\u3000\uFB50-\uFDFF\uFEFF\uFE70-\uFEFF\u10E60-\u10E7F\u1EE00-\u1EEFF]")
        kor_check = re.compile(
            "[^\u0000-\u0007\u0009-\u05C0\u1100-\u11FF\u200B-\u2010\u2012-\u202E\u2030-\u205E\u2100-\u214F\u3000\u3131-\u318E\u3200-\u35FF\uA960-\uA97C\uAC00-\uBFFF\uC058-\uD79D\uFFE1\uFEFF]")
        viet_check = re.compile("[^\u0000-\u0007\u0009-\u05C0\u200B-\u2010\u2012-\u202E\u2030-\u205E\u1EA0-\u1EF9\u3000\uFEFF]")

        sub_dict = {'91': chi_check,
                    'TW': chi_check,
                    'ZH': chi_check,
                    'CT': chi_check,
                    "96": chi_check,
                    "EN": eng_check,
                    "09": eng_check,
                    "61": msid_check,
                    "MS": msid_check,
                    "ID": msid_check,
                    "TH": thai_check,
                    "76": bur_check,
                    "AR": ara_check,
                    "KO": kor_check,
                    "VI": viet_check,
                    "46": viet_check,
                    }
        win = Toplevel()
        win.iconbitmap('icons/subtitle.ico')
        win.geometry("620x295+{}+{}".format(root.winfo_x() + 300, root.winfo_y() + 200))
        win.resizable(width=False, height=False)
        win.grab_set()
        frame1 = Frame(
            master=win,
            bg='#808000'
        )
        frame1.pack()
        editArea = tkst.ScrolledText(frame1, width=100, height=15)
        editArea.config(font=("Arial", 9), wrap="word")
        frame2 = Frame(win)
        frame2.pack()
        temp_dfxp_storage = []
        def get_dfxp_details(a,b):
            try:
                with urllib.request.urlopen(link) as url:
                    if link.split(".")[-2][-2:].upper() in sub_dict and not "__" in link.split("/")[-1]:
                        file = url.readlines()
                        excpt = []
                        start = True
                        for num, line in enumerate(file, 0):
                            if headstop in line.decode("utf8") and start is True:
                                tc_start = ((line.decode("utf8").split(headstop, 1)[1]).split(" ")[0]).replace("\"", "")
                                start_num = num
                                start = False
                            result = re.findall(sub_dict[link.split(".")[-2][-2:]], line.decode("utf8"))
                            if result:
                                try:
                                    for i in result:
                                        excpt.append((i, num + 1 - start_num))
                                except:
                                    print(num)
                            tc_in = line.decode("utf8").split(r'begin="')[-1][:12]
                            tc_out = line.decode("utf8").split(r'end="')[-1][:12]
                            content = re.findall(r'<.*?>', line.decode("utf8"))
                            e = line.decode("utf8")
                            for i in content:
                                e = e.replace(i, "") if i != "<br/>" else e.replace(i, " ")
                            if start is False and "</p>" in line.decode("utf8"):
                                if not result:
                                    temp_dfxp_storage.append(
                                        "Line {:04d} -  {}   {}   {}".format((num + 1 - start_num), tc_in, tc_out, e))
                                    editArea.insert(INSERT,
                                                    "Line {:04d} -  {}   {}   {}".format((num + 1 - start_num), tc_in, tc_out, e))
                                else:
                                    editArea.insert(INSERT,
                                                    "Line {:04d} -  {}   {}   {}".format((num + 1 - start_num), tc_in, tc_out, e),
                                                    "error")
                    else:
                        file = url.readlines()
                        start = True
                        for num, line in enumerate(file, 0):
                            if headstop in line.decode("utf8") and start is True:
                                tc_start = ((line.decode("utf8").split(headstop, 1)[1]).split(" ")[0]).replace("\"", "")
                                start_num = num
                                start = False
                            tc_in = line.decode("utf8").split(r'begin="')[-1][:12]
                            tc_out = line.decode("utf8").split(r'end="')[-1][:12]
                            content = re.findall(r'<.*?>', line.decode("utf8"))
                            e = line.decode("utf8")
                            for i in content:
                                e = e.replace(i, "")
                            if start is False and "</p>" in line.decode("utf8"):
                                editArea.insert(INSERT,
                                                "Line {:04d} -  {}   {}   {}".format((num + 1 - start_num), tc_in, tc_out, e))

                    def tc_check():
                        if int(str(tc_start).replace(":", "").replace(".", "")) > 3000000:
                            tc_result = "TC in error"
                        elif int(str(tc_start).replace(":", "").replace(".", "")) == 0:
                            tc_result = "SUCCESS"
                        else:
                            tc_result = "SUCCESS"
                        tc_entry.insert(END, " " + tc_start)
                        tc_result_entry.insert(END, " " + tc_result)
                        if excpt:
                            illegal_entry.insert(END, " YES")
                            illegal_count_entry.insert(END, " {}".format(len(excpt)))
                            msg = ["{} at line {}".format(str(details[0]), details[1]) for details in excpt]
                            illegal_result_entry.insert(END, ",".join(msg))
                        else:
                            illegal_entry.insert(END, " N/A")
                            illegal_count_entry.insert(END, " N/A")
                            illegal_result_entry.insert(END, " N/A")
                    tc_check()
            except:
                editArea.insert(INSERT,"Cannot read subtitle file.")
        win.title("Inspecting {} - Powered by Subcheck 7.0".format(link.split("/")[-1]))
        tc_label = Label(frame2, text=" Starting timecode : ", width=17, anchor="e")
        tc_label.grid(row=0, column=0)
        tc_entry = Entry(frame2, width=15)
        tc_entry.grid(row=0, column=1)
        tc_result_label = Label(frame2, text=" Timecode validation : ", width=17, anchor="e")
        tc_result_label.grid(row=1, column=0)
        tc_result_entry = Entry(frame2, width=15)
        tc_result_entry.grid(row=1, column=1)
        illegal_label = Label(frame2, text=" Illegal char spotted: ", width=17, anchor="e")
        illegal_label.grid(row=0, column=2)
        illegal_entry = Entry(frame2, width=15)
        illegal_entry.grid(row=0, column=3)
        illegal_count_label = Label(frame2, text=" Count: ", width=8, anchor="e")
        illegal_count_label.grid(row=0, column=4)
        illegal_count_entry = Entry(frame2, width=16)
        illegal_count_entry.grid(row=0, column=5, padx=5)
        illegal_result_label = Label(frame2, text=" Detailed breakdown: ", width=17, anchor="e")
        illegal_result_label.grid(row=1, column=2)
        illegal_result_entry = Entry(frame2, width=46)
        illegal_result_entry.grid(row=1, column=3, columnspan=3, padx=5)
        editArea.pack(padx=10, pady=10, fill=BOTH, expand=True)
        editArea.tag_config("error", foreground="red")
        editArea.bind("<Key>", lambda e: vision.txtEvent(e))
        _start_new_thread(get_dfxp_details,("Thread",1))

class TreePop:
    def __init__(self,input_list1,input_list2,input_list_tw,input_list3,house_id_combination):
        top = Toplevel()
        top.geometry("1210x412+{}+{}".format(root.winfo_x()+5, root.winfo_y()+50))
        try:
            top.title("Details for house ID: "+str(input_list1[0][1]))
        except:
            top.title("Details")
        top.iconbitmap('icons/schedule.ico')
        top.resizable(width=False, height=False)
        top.transient(root)
        bottom_left_frame = Frame(top)
        bottom_left_frame.grid(row=0, column=0, padx=10)
        image = ImageTk.PhotoImage(Image.open("icons/dummy.jpg"))
        media_image = Label(bottom_left_frame, image=image,borderwidth=2,relief="solid")
        media_image.grid(row=1,column=0,columnspan=2)
        media_image_status = Label(bottom_left_frame, text="Loading library info...",font=("Helvetica", 12))
        media_image_status.grid(row=0,column=0,columnspan=2)
        image_detail,stream_link = {},{}
        ass_link = []
        def get_image_link(a,b):
            try:
                link = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:imageApproved,:marketRating,adminTags,content,thumbnails,added,:foxplusRating,:audioCode,:channel_id&byCustomValue=%7BhouseID%7D%7B"+str(input_list1[0][1])+"%7D&token="+vision.tok+"&account=Fox%20Play%20Asia"
                with urllib.request.urlopen(link) as url:
                    data = json.loads(url.read().decode())
                    for i in data["entries"][0]["content"]:
                        stream_link[i.get("title")] = i.get("streamingUrl")
                        if i.get("title").endswith(".ass") and not i.get("title").endswith("76.ass"):
                            ass_link.append(i.get("streamingUrl"))
                    for item in data["entries"][0]["thumbnails"]:
                        if item['assetTypes'] == ['L_XS_xhdpi']:
                            image_detail["link"] = item["streamingUrl"]
                    added_info = ["pl2$imageApproved","pl2$channel_id","pl2$foxplusRating","pl2$audioCode","added","adminTags","pl2$marketRating"]
                    added_dic = ["approve","channel","foxplusrating","audioCode","added","workflowlabel","marketRating"]
                    for i in range(len(added_info)):
                        try:
                            image_detail[added_dic[i]] = data["entries"][0][added_info[i]]
                        except KeyError:
                            pass
                    dfxp = [media_item["assetTypes"] for media_item in data["entries"][0]["content"] if media_item["url"].endswith(".dfxp")]
                    global sub_info,video_info,jpeg_info
                    sub_info = {"".join(media_item["assetTypes"]).replace("Sub ",""):(media_item["streamingUrl"],media_item["added"],media_item["fileSize"]) for media_item in data["entries"][0]["content"] if media_item["url"].endswith(".dfxp")}
                    video_info = {media_item["title"]: (",".join(media_item["assetTypes"]),media_item["releases"][0]["url"] if media_item["releases"] else "",media_item["added"],media_item["fileSize"]) for media_item in data["entries"][0]["content"] if media_item["title"].endswith((".mp4",".m3u8",".mpd",".ism"))}
                    jpeg_info = {media_item["title"]: (",".join(media_item["assetTypes"]),media_item["streamingUrl"] if media_item["streamingUrl"] else "",media_item["added"],media_item["fileSize"]) for media_item in data["entries"][0]["thumbnails"] if media_item["format"] == "JPEG"}
                    flat_dfxp = [item for sublist in dfxp for item in sublist]
                if image_detail:
                    try:
                        raw_data = urllib.request.urlopen(image_detail.get("link","")).read()
                        im = Image.open(io.BytesIO(raw_data))
                        image = ImageTk.PhotoImage(im)
                        media_image.config(image=image)
                        media_image.img = image
                    except:
                        pass
                else:
                    media_image_status.config(text="Cannot find image.")
                    image_approve_status.delete(0, END)
                    image_approve_status.insert(END, "False")
                media_image_status.config(text="Library info")
                image_approve_status.delete(0, END)
                image_approve_status.insert(END, str(image_detail.get("approve", "False")))
                foxplus_rating.delete(0,END)
                foxplus_rating.insert(END,image_detail.get("foxplusrating",""))
                market_rating.delete(0,END)
                if image_detail.get("marketRating",""):
                    market_rating_list = [k + " : " + v for k, v in image_detail["marketRating"].items()]
                else:
                    market_rating_list = []
                market_rating.insert(END,", ".join(market_rating_list))
                added_date.delete(0,END)
                added_date.insert(END,datetime.fromtimestamp(image_detail.get("added",1000) / 1000).strftime('%Y-%m-%d %H:%M'))
                channel.delete(0,END)
                channel.insert(END,",".join(image_detail.get("channel","")))
                audio.delete(0,END)
                audio.insert(END,image_detail.get("audioCode",""))
                subtitle.delete(0,END)
                subtitle.insert(END,",".join(flat_dfxp).replace("Sub ",""))
                workflowlabel.delete(0,END)
                workflowlabel.insert(END,",".join(image_detail.get("workflowlabel","")))
                click_more.grid(row=0, column=0, columnspan=2)
            except IndexError:
                media_image_status.config(text="Library info not found.")
                delete_list = [image_approve_status,foxplus_rating,market_rating,added_date,channel,audio,subtitle,workflowlabel]
                for i in delete_list:
                    i.delete(0,END)
                    i.insert(END, "Not found")
            except TclError as e:
                message = "Function TreePop.get_image_link"+str(e)
                error_log(message)
        image_approve_label = Label(bottom_left_frame,text="Image Approved:  ",width=15, anchor="e")
        image_approve_label.grid(row=2,column=0)
        image_approve_status = Entry(bottom_left_frame, width=30)
        image_approve_status.grid(row=2,column=1)
        added_date_label = Label(bottom_left_frame,text="Date Added:  ", width=15, anchor="e")
        added_date_label.grid(row=3,column=0)
        added_date = Entry(bottom_left_frame,width=30)
        added_date.grid(row=3,column=1)
        foxplus_rating_label = Label(bottom_left_frame,text="Fox Plus Rating:  ", width=15, anchor="e")
        foxplus_rating_label.grid(row=4,column=0)
        foxplus_rating = Entry(bottom_left_frame,width=30)
        foxplus_rating.grid(row=4,column=1)
        market_rating_label = Label(bottom_left_frame,text="Market Rating:  ", width=15, anchor="e")
        market_rating_label.grid(row=5,column=0)
        market_rating = Entry(bottom_left_frame,width=30)
        market_rating.grid(row=5,column=1)
        channel_label = Label(bottom_left_frame,text="Channel:  ",width=15, anchor="e")
        channel_label.grid(row=6,column=0)
        channel = Entry(bottom_left_frame,width=30)
        channel.grid(row=6,column=1)
        audio_label = Label(bottom_left_frame,text="Audio Code:  ",width=15, anchor="e")
        audio_label.grid(row=7,column=0)
        audio = Entry(bottom_left_frame,width=30)
        audio.grid(row=7,column=1)
        workflowlabel_label = Label(bottom_left_frame,text="Workflow Label:  ",width=15, anchor="e")
        workflowlabel_label.grid(row=8,column=0)
        workflowlabel = Entry(bottom_left_frame,width=30)
        workflowlabel.grid(row=8,column=1)
        subtitle_label = Label(bottom_left_frame,text="Subtitle:  ",width=15, anchor="e")
        subtitle_label.grid(row=9,column=0)
        subtitle = Entry(bottom_left_frame,width=30)
        subtitle.grid(row=9,column=1)
        sep = ttk.Separator(bottom_left_frame, orient=HORIZONTAL)
        sep.grid(row=10, column=0, columnspan=2, sticky='ew',pady=5)
        bottom_frame = Frame(top)
        bottom_frame.grid(row=1,column=0,columnspan=2)
        more_icon = PhotoImage(file="icons/down.png")

        def RightClickCopyHouseID(event):
            item = temp_tree.identify('item', event.x, event.y)
            item_text = temp_tree.item(item, "values")
            if item_text:
                def temp_copy_house_id():
                    root.clipboard_clear()  # clear clipboard contents
                    root.clipboard_append(item_text[1])
                rmenu = Menu(None, tearoff=0, takefocus=0)
                rmenu.add_command(label="Copy House ID", command=temp_copy_house_id, compound="left")
                rmenu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")

        def _more():
            click_more.grid_forget()
            bottom_frame.grid_forget()
            top.geometry("1210x687")
            media_label = Label(top_frame, text="Media items (Library vs DP)",font=("Helvetica", 12))
            media_label.grid(row=4, column=0)

            def RightClickSubtitle(event):
                item = media_tree.identify('item', event.x, event.y)
                item_text = media_tree.item(item, "values")
                if item_text:
                    tag = "".join(media_tree.item(item,"tags"))
                    if "sub" in tag:
                        def open_web_browser(event):
                            if tag == "lib_sub":
                                for k, v in sub_info.items():
                                    try:
                                        if item_text[0] in v[0]:
                                            webbrowser.open(v[0])
                                    except:
                                        pass
                            elif tag == "dp_sub":
                                for k, v in dp_sub_info.items():
                                    try:
                                        if item_text[0] in v[0]:
                                            webbrowser.open(v[0])
                                    except:
                                        pass
                        additional_commands = [
                            (' Open subtitle link', vision.link_icon, lambda e=event: open_web_browser(event)),
                            (' Inspect subtitle', vision.inspect_icon, lambda e=event: DoubleClick(event)),
                        ]
                        rmenu = Menu(None, tearoff=0, takefocus=0)
                        for (txt, img,cmd) in additional_commands:
                            rmenu.add_command(label=txt,image=img, command=cmd, compound="left")
                        rmenu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")
                    elif "lib_video" in tag:
                        if item_text[0] in video_info:
                            if "Accelerate" in video_info[item_text[0]][0]:
                                def play_video():
                                    if stream_link.get(item_text[0]):
                                        try:
                                            vlc_player = Player("")
                                            player_dnd = DragManager(vlc_player.parent)
                                            player_dnd.add_dragable(vlc_player.top_frame)
                                            vlc_player.load_url(stream_link.get(item_text[0]),ass_link)
                                        except RuntimeError:
                                            pass
                                def web_open_video():
                                    driver = webdriver.Firefox()
                                    driver.set_window_size(720, 540)
                                    driver.get("http://52.74.14.209/samples/players/qc.html")
                                    link = driver.find_element_by_id("url")
                                    link.clear()
                                    link.send_keys(video_info[item_text[0]][1])
                                    driver.find_element_by_xpath("//button[@onclick=\"setReleaseCall();\"]").click()
                                rmenu = Menu(None, tearoff=0, takefocus=0)
                                rmenu.add_command(label=" Play video", command=play_video, compound="left")
                                rmenu.add_command(label=" View video in QC player", command=web_open_video, compound="left")
                                rmenu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")
                            else:
                                rmenu = Menu(None, tearoff=0, takefocus=0)
                                rmenu.add_command(label=" View video in QC player", command="", compound="left",state=DISABLED)
                                rmenu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")

            def DoubleClick(event):
                item = media_tree.identify('item', event.x, event.y)
                item_text = media_tree.item(item, "values")
                if item_text:
                    tag = "".join(media_tree.item(item, "tags"))
                    if tag == "lib_sub":
                        for k,v in sub_info.items():
                            try:
                                if item_text[0] in v[0]:
                                    SubCheck(v[0])
                            except:
                                pass
                    elif tag == "dp_sub":
                        for k,v in dp_sub_info.items():
                            try:
                                if item_text[0] in v[0]:
                                    SubCheck(v[0])
                            except:
                                pass
                    elif "jpeg" in tag:
                        if item_text[0] in jpeg_info:
                            try:
                                webbrowser.open(jpeg_info[item_text[0]][1])
                            except:
                                pass
                    elif tag == "lib_video":
                        if item_text[0] in video_info:
                            if "Accelerate" in video_info[item_text[0]][0]:
                                try:
                                    if stream_link.get(item_text[0]):
                                        vlc_player = Player("")
                                        player_dnd = DragManager(vlc_player.parent)
                                        player_dnd.add_dragable(vlc_player.top_frame)
                                        vlc_player.load_url(stream_link.get(item_text[0]),ass_link)
                                except:
                                    pass
            media_tree = ttk.Treeview(top_frame,selectmode="browse",height=12)
            media_tree.grid(row=5, column=0)
            media_tree.bind("<Double-1>", DoubleClick)
            media_tree.bind("<Button-3>",RightClickSubtitle)
            media_header = ("Name","Asset Type","Added Date","File Size")
            media_tree["columns"] = media_header
            media_width = (405, 170, 110,70)
            media_tree.column("#0",width=120)
            for i in range(len(media_header)):
                media_tree.column(media_header[i], width=media_width[i], anchor="w")
                media_tree.heading(media_header[i], text=media_header[i], anchor='w')
            library_media = media_tree.insert("",1,"",text="Library",values="",open=True)
            sub_folder = media_tree.insert(library_media,1,"",text="Subtitles ({})".format(len(sub_info)),values="",open=True)
            for k,v in sub_info.items():
                media_tree.insert(sub_folder,END,text="",values=(v[0].split("/")[-1],k,datetime.fromtimestamp(v[1] / 1000).strftime('%Y-%m-%d %H:%M'),str(round(v[2]/1024))+" KB"),tags="lib_sub")
            media_folder = media_tree.insert(library_media,2,"",text="Video files ({})".format(len(video_info)),values="",open=False)
            for k,v in video_info.items():
                media_tree.insert(media_folder,END,text="",values=(k,v[0],datetime.fromtimestamp(v[2] / 1000).strftime('%Y-%m-%d %H:%M'),str(round(v[3]/1048576))+" MB"),tags="lib_video")
            jpeg_folder = media_tree.insert(library_media,3,"",text="Image ({})".format(len(jpeg_info)),values="",open=False)
            for k,v in jpeg_info.items():
                media_tree.insert(jpeg_folder, END, text="", values=(k, v[0], datetime.fromtimestamp(v[2] / 1000).strftime('%Y-%m-%d %H:%M'),str(round(v[3] / 1024)) + " KB"), tags="lib_jpeg")
            DP_label = Label(bottom_left_frame, text="Loading DP info...", width=15,font=("Helvetica", 12))
            DP_label.grid(row=11, column=0, columnspan=2,pady=3)
            dp_foxplus_rating_label = Label(bottom_left_frame, text="Fox Plus Rating:  ", width=15, anchor="e")
            dp_foxplus_rating = Entry(bottom_left_frame, width=30)
            dp_market_rating_label = Label(bottom_left_frame,text="Market Rating:  ", width=15, anchor="e")
            dp_market_rating = Entry(bottom_left_frame, width=30)
            dp_workflowlabel_label = Label(bottom_left_frame, text="Workflow Label:  ", width=15, anchor="e")
            dp_workflowlabel = Entry(bottom_left_frame,width=30)
            dp_subtitle_label = Label(bottom_left_frame, text="Subtitle:  ", width=15, anchor="e")
            dp_subtitle = Entry(bottom_left_frame, width=30)
            dp_label = [dp_foxplus_rating_label,dp_market_rating_label,dp_workflowlabel_label,dp_subtitle_label]
            dp_entry = [dp_foxplus_rating,dp_market_rating,dp_workflowlabel,dp_subtitle]
            start_row = 13
            start_index = 0
            dp_media_image = Label(bottom_left_frame, image=image, borderwidth=2, relief="solid")
            dp_media_image.grid(row=12, column=0, columnspan=2)
            for i in dp_label:
                i.grid(row=start_row,column=0)
                dp_entry[start_index].grid(row=start_row,column=1)
                dp_entry[start_index].insert(END,"Loading...")
                start_row+=1
                start_index +=1
            dp_detail = {}
            def get_dp_info(a,b):
                try:
                    link = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=:marketRating,adminTags,content,thumbnails,:foxplusRating&byCustomValue=%7BhouseID%7D%7B" + str(
                        input_list1[0][1]) + "%7D&token=" + vision.tok + "&account=FoxPlus%20Asia%20DP"
                    with urllib.request.urlopen(link) as url:
                        data = json.loads(url.read().decode())
                        for item in data["entries"][0]["thumbnails"]:
                            if item['assetTypes'] == ['L_XS_xhdpi']:
                                image_detail["link"] = item["streamingUrl"]
                        added_info = ["pl2$foxplusRating","adminTags", "pl2$marketRating"]
                        added_dic = ["foxplusrating","workflowlabel", "marketRating"]
                        for i in range(len(added_info)):
                            try:
                                dp_detail[added_dic[i]] = data["entries"][0][added_info[i]]
                            except KeyError:
                                pass
                        dfxp = [media_item["assetTypes"] for media_item in data["entries"][0]["content"] if media_item["url"].endswith(".dfxp")]
                        global dp_sub_info,dp_video_info
                        dp_sub_info = {"".join(media_item["assetTypes"]).replace("Sub ", ""): (
                        media_item["streamingUrl"], media_item["added"], media_item["fileSize"]) for media_item in
                                    data["entries"][0]["content"] if media_item["url"].endswith(".dfxp")}
                        dp_video_info = {media_item["title"]: (
                        ",".join(media_item["assetTypes"]), media_item["streamingUrl"], media_item["added"],
                        media_item["fileSize"]) for media_item in data["entries"][0]["content"] if
                                      media_item["title"].endswith((".mp4",".m3u8",".mpd",".ism"))}
                        dp_jpeg_info = {media_item["title"]: (",".join(media_item["assetTypes"]),media_item["streamingUrl"] if media_item["streamingUrl"] else "",media_item["added"],media_item["fileSize"]) for media_item in data["entries"][0]["thumbnails"] if media_item["format"] == "JPEG"}
                        flat_dfxp = [item for sublist in dfxp for item in sublist]
                    if image_detail:
                        try:
                            raw_data = urllib.request.urlopen(image_detail.get("link","")).read()
                            im = Image.open(io.BytesIO(raw_data))
                            dp_image = ImageTk.PhotoImage(im)
                            dp_media_image.config(image=dp_image)
                            dp_media_image.img = dp_image
                        except:
                            pass
                    for i in dp_entry: #dp_foxplus_rating,dp_market_rating,dp_workflowlabel,dp_subtitle
                        i.delete(0,END)
                    dp_foxplus_rating.insert(0,dp_detail.get("foxplusrating",""))
                    if dp_detail.get("marketRating", ""):
                        dp_detail_list = [k + " : " + v for k, v in dp_detail["marketRating"].items()]
                    else:
                        dp_detail_list = []
                    dp_market_rating.insert(0,", ".join(dp_detail_list))
                    dp_subtitle.insert(END, ",".join(flat_dfxp).replace("Sub ", ""))
                    dp_workflowlabel.insert(END,",".join(dp_detail.get("workflowlabel","")))
                    DP_label.config(text="DP info")
                    dp_media = media_tree.insert("", 1, "", text="DP", values="", open=True)
                    dp_sub_folder = media_tree.insert(dp_media, 1, "", text="Subtitles ({})".format(len(dp_sub_info)),
                                                   values="", open=True)
                    for k, v in dp_sub_info.items():
                        media_tree.insert(dp_sub_folder, END, text="", values=(
                        v[0].split("/")[-1], k, datetime.fromtimestamp(v[1] / 1000).strftime('%Y-%m-%d %H:%M'),
                        str(round(v[2] / 1024)) + " KB"),tags="dp_sub")
                    dp_media_folder = media_tree.insert(dp_media, 2, "", text="Video files ({})".format(len(dp_video_info)),
                                                     values="", open=False)
                    for k, v in dp_video_info.items():
                        media_tree.insert(dp_media_folder, END, text="", values=(
                        k, v[0], datetime.fromtimestamp(v[2] / 1000).strftime('%Y-%m-%d %H:%M'),
                        str(round(v[3] / 1048576)) + " MB"),tags="dp_video")
                    dp_jpeg_folder = media_tree.insert(dp_media, 3, "", text="Images ({})".format(len(dp_jpeg_info)),
                                                        values="", open=False)
                    for k, v in dp_jpeg_info.items():
                        media_tree.insert(dp_jpeg_folder, END, text="", values=(
                        k, v[0], datetime.fromtimestamp(v[2] / 1000).strftime('%Y-%m-%d %H:%M'),
                        str(round(v[3] / 1024)) + " KB"),tags="dp_jpeg")
                except IndexError:
                    DP_label.config(text="DP info not found.")
                    for i in dp_entry:
                        i.delete(0,END)
                        i.insert(END,"Not found")

            vision.start_thread(get_dp_info)

        click_more = Button(bottom_frame,text=" MORE ",relief=FLAT, command=_more,image=more_icon,compound="left",font=("Helvetica", 11))
        click_more.img = more_icon
        all_field = [image_approve_status,added_date,foxplus_rating,market_rating,channel,audio,subtitle,workflowlabel]
        for i in all_field:
            i.insert(END,"Loading...")

        top_frame = Frame(top)
        top_frame.grid(row=0,column=1)
        def flag_media():
            if vision.nb.index("current") == 0:
                vision.tree.item(iid,tags="flagged")
                flag_this.config(text="Media flagged", state=DISABLED)
                vision.temp_store = [[vision.tree.item(child)["values"], vision.tree.item(child)["tags"][0],vision.tree.item(child)["image"]] for child in vision.tree.get_children()]
                vision.filtered_data = [[vision.tree.item(child)["values"], vision.tree.item(child)["tags"][0],vision.tree.item(child)["image"]] for child in vision.tree.get_children()]
            else:
                vision.meta_tree.item(iid,tags="flagged")
                flag_this.config(text="Media flagged", state=DISABLED)
        flag_image = PhotoImage(file="icons/flag.png")
        flag_this = Button(top_frame, text="Flag this media",image=flag_image,compound="left",relief="ridge",command=flag_media,justify=LEFT,anchor="w")
        #flag_this.place(x=758, y=0, width=120)
        flag_this.img = flag_image
        top_label = Label(top_frame,text="MPX availability",font=("Helvetica", 12))
        top_label.grid(row=0,column=0)
        temp_tree = ttk.Treeview(top_frame, selectmode='browse', height=10)
        temp_tree.grid(row=1,column=0)
        header = ("Account","House_no","Title","Version","Available date","Expiration date", "Channel","Approved")
        width = (80,80,260,60,110,110,110,65)
        temp_tree["columns"] = header
        temp_tree['show'] = 'headings'
        for i in range(len(header)):
            temp_tree.column(header[i], width=width[i], anchor="w")
            temp_tree.heading(header[i], text=header[i], anchor='w')
        for item in input_list1:
            temp_tree.insert('', 0, text="",values=item,tags="library")
        for item in input_list_tw:
            temp_tree.insert('', 0, text="",values=item,tags="schedule")
        for item in input_list2:
            temp_tree.insert('', 0, text="",values=item,tags="schedule")
        for item in input_list3:
            temp_tree.insert('', 0, text="", values=item, tags="DP")
        temp_tree.tag_configure("library", background="light blue")
        temp_tree.tag_configure("schedule", background="gold")
        temp_tree.tag_configure("DP", background="DarkOliveGreen1")
        temp_tree.bind("<Button-3>",RightClickCopyHouseID)
        bottom_label = Label(top_frame,text="MAUI record",font=("Helvetica", 12))
        bottom_label.grid(row=2,column=0)
        maui_tree = ttk.Treeview(top_frame, selectmode='browse', height=4)
        maui_tree.grid(row=3,column=0)
        maui_header = ("Box No","Filename","Pri HDD","Sec HDD","Create Date","Remarks")
        maui_width = (80, 340, 70, 70, 130, 185)
        maui_tree["columns"] = maui_header
        maui_tree['show'] = 'headings'
        for i in range(len(maui_header)):
            maui_tree.column(maui_header[i], width=maui_width[i], anchor="w")
            maui_tree.heading(maui_header[i], text=maui_header[i], anchor='w')
        result = vision.df.loc[vision.df["House Number"].isin(house_id_combination),["Box Number","file","pri_hdd","sec_hdd","archive_date","qc_by"]]
        result = result.loc[result.notnull().all(1)]
        if len(result):
            for i in result.values.tolist():
                maui_tree.insert("",0,text="",values=i,tags="maui")
        maui_tree.tag_configure("maui", background="DarkSeaGreen1")
        vision.start_thread(get_image_link)
        top.grab_set()

class StringDialog(simpledialog._QueryString):
    def body(self, master):
        super().body(master)
        self.iconbitmap('icons/vision3.ico')

    def ask_string(title, prompt, **kargs):
        d = StringDialog(title, prompt, **kargs)
        return d.result

class SystemCheck:
    """
    A class to perform routine check for updated data.
    Requires connection to folder in S:. Access has to be granted for non-BTO users.
    """
    def __init__(self):
        self.json_time = {}
        self.update_interval = 5
        self.path = "S:/OZ Migration/T&O/Overseer/"

    def check_data_update_new(self):
        current = os.path.getmtime("database.pkl")
        target = os.path.getmtime(r"S:\OZ Migration\T&O\Overseer\database.pkl")
        if current < target and PopUpMessage.count==0:
            pop = PopUpMessage(0, "Update detected",
                               "A database update is available.",
                               "(CLICK TO REFRESH DATA)")
            pop.can.bind("<Button-1>", pop.refresh_data)
        root.after(self.update_interval * 60 * 1000, self.check_data_update_new)

    def check_vision_update_available(self):
        path = "S:/OZ Migration/T&O/Overseer/" #os.getcwd()+r"\tk"
        try:
            for file in os.listdir(path):
                if "update" in file and int(file.split("_")[1]) > app_date:
                    answer = messagebox.askokcancel("Update available", "An update is available. Restart your client now?")
                    if answer:
                        try:
                            self.t.start()
                            root.destroy()
                        except:
                            root.destroy()
        except PermissionError:
            print ("Cannot access folder.")
        except:
            pass
        root.after(600000,self.check_vision_update_available)

    def restart_self(self):
        """
        Method to relaunch APP after close down.
        """
        time.sleep(2)
        subprocess.call(os.getcwd() + r"\vision_client.exe", shell=False)

    def ask_restart(self):
        """
        Method to thread the restart operation.
        During trial and error I found that a threaded process will not terminate even after calling gui.destroy().
        Therefore threading a subprocess.call can be utilized in mimic auto-restart behaviour.
        :return:
        """
        self.t = threading.Thread(target=self.restart_self)
        answer = messagebox.askokcancel("Restart client", "Do you want to restart and refresh data?")
        if answer:
            try:
                self.t.start()
                root.destroy()
            except:
                root.destroy()

class VideoSubCheck:

    def __init__(self,master):
        self.sports_count = 0
        self.default_color = settings.default_color
        self.font_color = settings.font_color
        self._detached = set()
        self.master = master
        self.create_control_frame(master)
        self.top_frame = Frame(master)
        self.top_frame.grid(row=1,column=0)
        self.create_tree()
        self.bot_frame = Frame(master)
        self.bot_frame.grid(row=2,column=0)

    def create_control_frame(self,master):
        self.control_frame = Frame(master)
        self.control_frame.grid(row=0, column=0, sticky="w")
        AANTAL = [(0, "House ID"), (1, "Series/Name"),(2, "Channel"), (3, "Additional Information"),(4,"Added date"),]
        self.v = IntVar()
        self.v.set(0)
        for text, mode in AANTAL:
            but = Radiobutton(self.control_frame, padx=20, pady=2,bd=2,relief="groove",selectcolor="purple3",
                              bg=self.default_color,fg=self.font_color,
                              text=mode, variable=self.v, value=text, indicatoron=0)
            if text == 3:
                but.grid(row=0, column=text, pady=2,columnspan=2)
            elif text == 4:
                but.grid(row=0, column=text+1, pady=2)
            else:
                but.grid(row=0, column=text, pady=2)
        vcmd = (self.control_frame.register(self._columns_searcher), '%P')
        self.fil_label = Label(self.control_frame, text="Choose filter: ")
        self.fil_label.grid(row=1, column=0, sticky="e", padx=5)
        self.hse_entry = Entry(self.control_frame,width=33, validate="key", validatecommand=vcmd)
        self.hse_entry.grid(row=1,column=1,columnspan=2,sticky="w")
        self.hse_entry.config(state="disabled")
        chi_sub_image = PhotoImage(file="icons/subtitle.png")
        self.chi_sub_button = Button(self.control_frame, text=" Start manifest check ",
                                     image=chi_sub_image, compound="left", font=("Arial", 14),
                                     command=lambda: _start_new_thread(self.click_me, ("Thread", 1)),
                                     height=44, width=250, bd=2, relief="ridge",
                                     bg=self.default_color,fg=self.font_color,cursor="hand2")
        self.chi_sub_button.grid(row=0, column=8, rowspan=2,padx=3)
        self.hse_entry.focus_set()
        self.chi_sub_button.img = chi_sub_image
        CreateToolTip(self.chi_sub_button, "This will check the most recent media on DP. "
                                           "Sports and test media will be skipped. "
                                           "Any media created before subtitles creation will be highlighted. "
                                           "You can double click on each media to view detailed info.")
        self.error_only = IntVar()
        self.error_only.set(1)
        Checkbutton(self.control_frame, variable=self.error_only, borderwidth=2,#relief="groove",
                                         text="Show error only",command="", anchor="w",
                                         justify=LEFT).grid(row=1,column=5)
        Label(self.control_frame,text="No. of items: ").grid(row=1,column=3,sticky="e",padx=5)
        self.num_of_item = ttk.Combobox(self.control_frame, state="readonly", width=10)
        self.num_of_item.grid(row=1,column=4)
        self.num_of_item["values"] = [i for i in range(100,1000,100)]+[i for i in range(1000,10001,500)]
        self.num_of_item.set(1000)
        self.msg_box = tkst.ScrolledText(self.control_frame, wrap=WORD, width=58, height=3)
        self.msg_box.grid(row=0,column=9,rowspan=2,padx=3)
        self.msg_box.config(font=("Arial", 8))
        self.msg_box.insert(END,"Manifest console. Click button on left to start checking.\n"
                                "You can continue to perform other tasks while waiting for result.\n")
        self.msg_box.bind("<Key>", lambda e: vision.txtEvent(e))

    def _columns_searcher(self, P):
        #              originally a set            returns a tuple
        children = list(self._detached) + list(self.tree.get_children())
        self._detached = set()
        self._brut_searcher(children, P)
        return True

    def _brut_searcher(self, children, query):
        i_r = -1
        try:
            for item_id in children:
                if self.v.get() == 0:
                    text = self.tree.item(item_id)['text']  # already contains the string-concatenation (over columns) of the row's values
                elif self.v.get() == 1:
                    text = self.tree.item(item_id)["values"][1]
                elif self.v.get() == 2:
                    text = self.tree.item(item_id)["values"][4]
                elif self.v.get() == 4:
                    text = self.tree.item(item_id)["values"][6]
                else:
                    text = self.tree.item(item_id)["values"][7]
                if query.upper() in str(text).upper():
                    i_r += 1
                    self.tree.reattach(item_id, '', i_r)
                else:
                    self._detached.add(item_id)
                    self.tree.detach(item_id)
        except:
            pass

    def create_tree(self):
        self.tree = ttk.Treeview(self.top_frame,height=25,selectmode='browse')
        self.tree.pack(side='left')
        vsb = ttk.Scrollbar(self.top_frame, orient="vertical",command=self.tree.yview)
        vsb.pack(side='right', fill='y')
        self.tree.configure(yscrollcommand=vsb.set)
        header = ("Media ID", "Series", "Season", "Ep", "Channel", "URL", "Added Date", "Additional information")
        width = (170, 300, 50, 50, 150, 300, 100, 230)
        self.tree["columns"] = header
        self.tree.column("#0", width=150, anchor="w")
        self.tree.heading("#0", text="House ID", anchor="w")
        for i in range(len(header)):
            self.tree.column(header[i], width=width[i], anchor="w")
            self.tree.heading(header[i], text=header[i], anchor='w')
            self.tree["displaycolumns"] = ("Media ID", "Series", "Season", "Ep", "Channel", "Added Date", "Additional information")
        #self.tree.tag_configure("problem", background="red",foreground="white")
        self.tree.bind("<Double-1>",self.double_click)
        self.tree.bind("<Button-3>", self.popup_menu)

    def click_me(self,a,b):
        if vision.tok == "":
            vision.mpx_connect_console()
            while True:
                time.sleep(0.2)
                if vision.tok != "":
                    self.click_me("","")
                    break
            return
        else:
            self.chi_sub_button.config(relief="sunken",command="",state="disabled", text=" Fetching result... ")
            self.get_content()

    def get_content(self):
        self.tree.delete(*self.tree.get_children())
        self.top_frame.config(cursor="wait")
        tok = vision.tok
        start_count = 1
        self.hse_entry.config(state="normal")
        self.hse_entry.focus_set()
        self.num_of_item.config(state="disabled")
        while start_count < int(self.num_of_item.get()):
            media_item = f"http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=id,added,categories,:season,:episodeNumber,:houseID,title,:channel_id,:series-Title,content&range={start_count}-{start_count+100}&token={tok}&account=FoxPlus%20Asia%20DP" #&sort=added|desc
            try:
                with urllib.request.urlopen(media_item) as url:
                    data = json.loads(url.read().decode())
                    for item in data['entries']:
                        channel_id = ", ".join(item.get("pl3$channel_id",item.get("pl2$channel_id",item.get("pl1$channel_id",""))))
                        media_id = str(item.get("id","").split("/")[-1])
                        title = item.get("title","")
                        series_title = item.get("pl3$series-Title",item.get("pl2$series-Title",item.get("pl1$series-Title","")))
                        house_id = item.get("pl1$houseID","")
                        ep = str(item.get("pl1$episodeNumber",""))
                        season = str(item.get("pl1$season",""))
                        add_date = datetime.fromtimestamp(item.get("added", 1000) / 1000).strftime('%Y-%m-%d %H:%M')
                        category = ", ".join(i.get("name","") for i in item.get("categories",""))
                        video, subtitle = {}, {}
                        sub_date,vid_date = [],[]
                        if "foxsports" in channel_id or "test" in channel_id:
                            self.sports_count += 1
                            pass
                        else:
                            for media_item in item["content"]:
                                if media_item["url"].upper().endswith(".DFXP"):
                                    date = datetime.fromtimestamp(media_item.get("added", 1000) / 1000).strftime('%Y-%m-%d %H:%M')
                                    subtitle[media_item.get("title")] = {"streamingUrl":media_item.get("streamingUrl",""),"added":date,"asset":', '.join(media_item.get("assetTypes",""))}
                                    sub_date.append(date)
                                elif any(media_item["url"].upper().endswith(s) for s in (".M3U8",".ISM",".MPD")):
                                    date = datetime.fromtimestamp(media_item.get("added",1000) / 1000).strftime('%Y-%m-%d %H:%M')
                                    video[media_item.get("title")] = {"streamingUrl":media_item.get("streamingUrl",""),"added":date,"asset":', '.join(media_item.get("assetTypes",""))}
                                    vid_date.append(date)
                            def populate_tree(tags):
                                master = self.tree.insert("", END, "", text=house_id, values=(media_id, series_title if series_title else title, season, ep, channel_id, "", add_date,
                                                            category), open=False)
                                sub_folder = self.tree.insert(master, 1, "", text="Subtitle ({})".format(len(subtitle)),values="", open=True)
                                vid_folder = self.tree.insert(master, 2, "", text="Video ({})".format(len(video)),values="", open=True)
                                for k, v in subtitle.items():
                                    self.tree.insert(sub_folder, END, text="", values=(k, v.get("asset"), "", "", "", v.get("streamingUrl"), v.get("added"), ""))
                                for k, v in video.items():
                                    self.tree.insert(vid_folder, END, text="", values=(k.split("-")[-1], v.get("asset"), "", "", "", v.get("streamingUrl"), v.get("added"),""))
                            try:
                                if any(s > max(sub_date) for s in vid_date) or len(video) < 3:
                                    if not self.error_only.get():
                                        master = self.tree.insert("", END, "", text=house_id, values=(media_id,series_title if series_title else title,season,ep,channel_id,"",add_date,category), open=False)
                                        sub_folder = self.tree.insert(master, 1, "",text="Subtitle ({})".format(len(subtitle)),values="", open=True)
                                        vid_folder = self.tree.insert(master, 2, "",text="Video ({})".format(len(video)), values="",open=True)
                                        for k, v in subtitle.items():
                                            self.tree.insert(sub_folder, END, text="", values=(k, v.get("asset"), "", "", "", v.get("streamingUrl"), v.get("added"), ""))
                                        for k, v in video.items():
                                            self.tree.insert(vid_folder, END, text="", values=(k.split("-")[-1], v.get("asset"), "", "", "", v.get("streamingUrl"),v.get("added"), ""))
                                else:
                                    master = self.tree.insert("", END, "", text=house_id,values=(media_id,series_title if series_title else title, season, ep, channel_id,"",add_date,category),open=False, tag=("problem",))
                                    sub_folder = self.tree.insert(master, 1, "",text="Subtitle ({})".format(len(subtitle)), values="",open=True)
                                    vid_folder = self.tree.insert(master, 2, "", text="Video ({})".format(len(video)),values="", open=True)
                                    for k, v in subtitle.items():
                                        self.tree.insert(sub_folder, END, text="", values=(k, v.get("asset"), "", "", "", v.get("streamingUrl"), v.get("added"), ""))
                                    for k, v in video.items():
                                        self.tree.insert(vid_folder, END, text="", values=(k.split("-")[-1], v.get("asset"), "", "", "", v.get("streamingUrl"),v.get("added"), ""))
                            except ValueError:
                                if not self.error_only.get():
                                    master = self.tree.insert("", END, "", text=house_id, values=(str(media_id),series_title if series_title else title, season, ep, channel_id,"", add_date,category), open=False)
                                    sub_folder = self.tree.insert(master, 1, "", text="Subtitle ({})".format(len(subtitle)), values="",open=True)
                                    vid_folder = self.tree.insert(master, 2, "", text="Video ({})".format(len(video)), values="",open=True)
                                    for k, v in subtitle.items():
                                        self.tree.insert(sub_folder, END, text="", values=(k,v.get("asset"),"","","",v.get("streamingUrl"),v.get("added"),""))
                                    for k,v in video.items():
                                        self.tree.insert(vid_folder, END, text="", values=(k.split("-")[-1],v.get("asset"),"","","",v.get("streamingUrl"),v.get("added"),""))
            except:
                pass
            start_count += 100
        self.chi_sub_button.config(relief="ridge",command=lambda: _start_new_thread(self.click_me, ("Thread", 1)),state="normal",text=" Start video/subtitle check ")
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col, command=lambda _col=col: self.treeview_sort_column(self.tree, _col, False))
        self.msg_box.insert(END, f"Manifest checking has completed. Showing {len(self.tree.get_children())} results.\n")
        self.num_of_item.config(state="normal")
        pop = PopUpMessage(2,"Task completed","Manifest checking has completed.","(CLICK TO VIEW)")
        pop.can.bind("<Button-1>", pop.jump_tab)
        self.top_frame.config(cursor="arrow")

    def double_click(self,event):
        item = self.tree.identify('item', event.x, event.y)
        try:
            if self.tree.item(item)["values"] and self.tree.item(item)["values"][0].endswith("dfxp"):
                try:
                    SubCheck(self.tree.item(item)["values"][5])
                except:
                    pass
        except AttributeError:
            pass

    def popup_menu(self,event):
        item = self.tree.identify('item', event.x, event.y)
        current = self.tree.item(item)["text"]
        rmenu = Menu(None, tearoff=0, takefocus=0)
        all_items = ",".join([self.tree.item(child)["text"] for child in self.tree.get_children()])
        commands = {"Copy HouseID": lambda: root.clipboard_append(current if current else ""),
                    "Copy all HouseID": lambda: root.clipboard_append(all_items),
                    "Query MPX": lambda: vision.house_id_event([current,])
                    if current else ""}

        for txt, command in commands.items():
            rmenu.add_command(label=txt, compound="left", command=command)

        rmenu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")
        # item = self.tree.identify('item', event.x, event.y)
        # if self.tree.item(item)["text"]:
        #     vision.house_id_event([self.tree.item(item)["text"],])

    def treeview_sort_column(self,tv, col, reverse):
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)

        # rearrange items in sorted positions
        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        # reverse sort next time
        tv.heading(col, command=lambda: \
            self.treeview_sort_column(tv, col, not reverse))

class SearchExpiration:
    """

    """
    def __init__(self,*args):
        self.master = args[0]
        self.tok = ""
        self.create_control_frame(self.master)
        self.top_frame = Frame(self.master)
        self.top_frame.grid(row=1,column=0)
        self.create_tree()
        self.bot_frame = Frame(self.master)
        self.bot_frame.grid(row=2,column=0)
        self.filter_dict = {}

    def create_control_frame(self,master):
        self.control_frame = Frame(master)
        self.control_frame.grid(row=0, column=0, sticky="w")
        AANTAL = [(0, "House ID"), (1, "Media ID"), (2, "Series/Name"), (3, "Channel"),
                  (4, "Available date"), ]
        self.v = IntVar()
        self.v.set(0)
        for text, mode in AANTAL:
            if text ==4:
                but = Radiobutton(self.control_frame, padx=8, pady=2, bd=2, relief="groove", selectcolor="purple3",
                                  bg=settings.default_color, fg=settings.font_color,
                                  text=mode, variable=self.v, value=text, indicatoron=0)
            else:
                but = Radiobutton(self.control_frame, padx=16, pady=2, bd=2, relief="groove", selectcolor="purple3",
                                  bg=settings.default_color, fg=settings.font_color,
                                  text=mode, variable=self.v, value=text, indicatoron=0)
            but.grid(row=0, column=text, pady=2)
        vcmd = (self.control_frame.register(self._columns_searcher), '%P')
        self.fil_label = Label(self.control_frame, text="Choose filter: ")
        self.fil_label.grid(row=1, column=0, sticky="e", padx=5)
        self.hse_entry = Entry(self.control_frame, width=33, validate="key", validatecommand=vcmd)
        self.hse_entry.grid(row=1, column=1, columnspan=2, sticky="w")
        self.hse_entry.config(state="disabled")
        self.hse_entry.focus_set()
        Label(self.control_frame, text="No. of items: ").grid(row=0, column=5, sticky="e", padx=2)
        self.num_of_item = ttk.Combobox(self.control_frame, state="readonly", width=10)
        self.num_of_item.grid(row=0, column=6,sticky="w")
        self.num_of_item["values"] = [i for i in range(100, 1000, 100)] + [i for i in range(1000, 10001, 500)]
        self.num_of_item.set(1000)
        self.msg_box = tkst.ScrolledText(self.control_frame, wrap=WORD, width=58, height=3)
        self.msg_box.grid(row=0, column=10, rowspan=2, padx=3)
        self.msg_box.config(font=("Arial", 8))
        self.msg_box.insert(END,
                            "Expired title console. Click button on left to start checking.\nYou can continue to perform other tasks while waiting for result.\n")
        self.msg_box.bind("<Key>", lambda e: vision.txtEvent(e))

        Label(self.control_frame,text="Expire date: ").grid(row=1,column=3,sticky="e",padx=2)
        self.expire_date = DateEntry(self.control_frame,width=13,justify="center")
        self.expire_date.grid(row=1,column=4)
        Label(self.control_frame, text="Expire time: ").grid(row=1, column=5, sticky="e", padx=2)
        self.expire_time = ttk.Combobox(self.control_frame, state="readonly", width=10)
        self.expire_time.grid(row=1,column=6)
        self.expire_time["values"] = [f'{h}:{m} {ap}' for ap in ('AM', 'PM') for h in ([12] + list(range(1,12))) for m in ('00', '30')]
        self.expire_time.set("12:00 AM")
        self.expire_icon = PhotoImage(file="icons/expire.png")
        self.get_button = Button(self.control_frame, text=" Generate report ",command=lambda: _start_new_thread(self.click_me,("Thread",3)),
                                 compound="left",font=("Arial", 14),image=self.expire_icon,height=44, width=195,
                                 bd=2, relief="ridge", bg=settings.default_color, fg=settings.font_color,
                                 cursor="hand2")
        self.get_button.img = self.expire_icon
        self.get_button.grid(row=0,column=7,rowspan=2,columnspan=2,sticky="news",padx=2,pady=3)

    def get_unix_timestamp(self):
        date = self.expire_date.get() +" "+ self.expire_time.get()
        unix_time = int(time.mktime(datetime.strptime(date, "%m/%d/%Y %I:%M %p").timetuple())*1000)
        return unix_time

    def _columns_searcher(self, P):
        tree_child = {text:num for num, text in enumerate(self.tree.get_children())}
        children = {**self.filter_dict,**tree_child}
        self.filter_dict = {}
        self._brut_searcher(children, P)
        return True

    def _brut_searcher(self, children, query):
        i_r = -1
        try:
            for item_id,num in children.items():
                if self.v.get() == 0:
                    text = self.tree.item(item_id)['text']
                elif self.v.get() == 1:
                    text = self.tree.item(item_id)["values"][0]
                elif self.v.get() == 2:
                    text = self.tree.item(item_id)["values"][1]
                elif self.v.get() == 3:
                    text = self.tree.item(item_id)["values"][4]
                else:
                    text = self.tree.item(item_id)["values"][5]
                if query.upper() in str(text).upper():
                    i_r += 1
                    self.tree.reattach(item_id, '', self.filter_dict.get(item_id, num))
                    self.tree.reattach(item_id, '', i_r)
                else:
                    self.filter_dict[item_id] = num
                    self.tree.detach(item_id)
        except:
            pass

    def click_me(self,*args):
        self.get_button.config(relief="sunken", state="disabled")
        if vision.tok == "":
            vision.mpx_connect_console()
            while True:
                time.sleep(0.2)
                if vision.tok != "":
                    self.click_me("","")
                    break
            return
        else:
            self.get_link()

    def del_selected(self,event=None,mode="select"):
        if mode == "select":
            for child in self.tree.selection():
                self.tree.detach(child)
        elif mode == "all":
            self.tree.delete(*self.tree.get_children())

    def get_link(self):
        if vision.tok:
            tok = vision.tok
            #self.tree.delete(*self.tree.get_children())
            self.top_frame.config(cursor="wait")
            self.hse_entry.config(state="normal")
            self.hse_entry.focus_set()
            self.num_of_item.config(state="disabled")
            self.expire_date.config(state="disabled")
            self.expire_time.config(state="disabled")
            link = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=id,availableDate,expirationDate,:season,:episodeNumber,:houseID,title,:channel_id,:series-Title&byExpirationDate="+str(self.get_unix_timestamp())+f"&range=1-{self.num_of_item.get()}&token=" + tok + "&account=FoxPlus%20Asia%20DP"
            try:
                self.msg_box.insert(END,"Fetching results from MPX...\n")
                self.msg_box.see("end")
                with urllib.request.urlopen(link) as url:
                    data = json.loads(url.read().decode())
                    for i in data["entries"]:
                        start_date = datetime.fromtimestamp(i.get("availableDate", 1000) / 1000).strftime('%Y-%m-%d %H:%M')
                        end_date = datetime.fromtimestamp(i.get("expirationDate", 1000) / 1000).strftime('%Y-%m-%d %H:%M')
                        if "foxsports" in ",".join(i.get('pl2$channel_id',"")):
                            pass
                        else:
                            if i.get('pl2$series-Title'):
                                title = i.get('pl2$series-Title',"")
                            else:
                                title = i.get("title","")
                            self.tree.insert("",END,text=i.get("pl1$houseID",""),
                                            values=(i.get("id","").split("/")[-1],title,i.get('pl1$season',""),i.get('pl1$episodeNumber',""),
                                                    ",".join(i.get('pl2$channel_id',"")),start_date,end_date))
                self.get_button.config(relief="ridge", state="normal")
                self.msg_box.insert(END,f"Expired title checking has completed. Showing {len(self.tree.get_children())} results.\n")
                self.msg_box.see("end")
                self.num_of_item.config(state="normal")
                self.expire_date.config(state="normal")
                self.expire_time.config(state="normal")
                if vision.nb.index("current") != 3:
                    pop = PopUpMessage(3, "Task completed", "Expired title checking has completed.", "(CLICK TO VIEW)")
                    pop.can.bind("<Button-1>", pop.jump_tab)
                self.top_frame.config(cursor="arrow")
            except:
                self.msg_box.insert(END,f"Cannot connect to MPX.\n")
                self.msg_box.see("end")
                self.num_of_item.config(state="normal")
                self.expire_date.config(state="normal")
                self.expire_time.config(state="normal")
                self.top_frame.config(cursor="arrow")
                self.get_button.config(relief="ridge", state="normal")
        else:
            self.get_button.config(relief="sunken", state="disabled")
            vision.searching = True
            vision.mpx_connect_console()
            self.get_link()

    def create_tree(self):
        self.tree = ttk.Treeview(self.top_frame,height=25,selectmode='extended')
        self.tree.pack(side='left',fill="both")
        vsb = ttk.Scrollbar(self.top_frame, orient="vertical",command=self.tree.yview)
        vsb.pack(side='right', fill='y')
        self.tree.configure(yscrollcommand=vsb.set)
        header = ("Media ID", "Series/Title", "Season", "Ep", "Channel","Available Date","Expiration Date")
        width = (170, 300, 60, 60, 160, 150, 150)
        self.tree["columns"] = header
        self.tree.column("#0", width=150, anchor="w")
        self.tree.heading("#0", text="House ID", anchor="w")
        for i in range(len(header)):
            self.tree.column(header[i], width=width[i], anchor="w")
            self.tree.heading(header[i], text=header[i], anchor='w', command=lambda _col=header[i]: self._tree_sort_column(self.tree, _col, False))
        self.tree.tag_configure("problem", background="red",foreground="white")
        self.tree.bind("<Double-1>",self.double_click)
        self.tree.bind("<Button-3>", self.popup_menu)

    def popup_menu(self, event):
        rmenu = Menu(None, tearoff=0, takefocus=0)
        def expire_export():
            f = filedialog.asksaveasfilename(defaultextension=".xlsx",filetypes=(("Excel file", "*.xlsx"), ("All Files", "*.*")))
            if f:
                try:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Expired title report generated"
                    header = ["House ID","Media ID", "Series/Title", "Season", "Ep", "Channel","Available Date","Expiration Date"]
                    ws.append(header)
                    grey = colors.Color(rgb='444444')
                    my_fill = fills.PatternFill(patternType='solid', fgColor=grey)
                    for cell in ("A1","B1","C1","D1","E1","F1","G1","H1"):
                        ws[cell].fill = my_fill
                        ws[cell].font = Font(color='FFFFFF')
                    ws.auto_filter.ref = 'A:H'
                    for child in self.tree.get_children():
                        try: #FIXME
                            current_value = self.tree.item(child)["values"]
                            house_id = self.tree.item(child)["text"]
                            result = [house_id] + list(current_value)
                            ws.append(result)
                        except:
                            continue
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter  # Get the column name
                        for cell in col:
                            try:  # Necessary to avoid error on empty cells
                                if len(str(cell.value)) > max_length:
                                    max_length = len(cell.value)
                            except:
                                pass
                        adjusted_width = (max_length + 1) * 1.15
                        ws.column_dimensions[column].width = adjusted_width
                    ws.column_dimensions["E"].width = 7
                    wb.save(filename=f)
                    filename = f.split("/")[-1]
                    message = " Report exported as {}".format(filename)
                    self.msg_box.insert(END,message.lstrip()+"\n")
                    self.msg_box.see("end")
                    messagebox.showinfo("Success",message)
                except PermissionError:
                    message = " Please close the file or enable permission."
                    messagebox.showerror("Fail to generate report", message)
                    self.msg_box.insert(END, message.lstrip() + "\n")
                    self.msg_box.see("end")
                except IndexError:
                    message = "Function SearchExpiration.popup_menu.right_click_export: An index error occurred."
                    self.msg_box.insert(END, message + "\n")
                    self.msg_box.see("end")
                    error_log(message)
                except Exception as e:
                    message = "Function SearchExpiration.popup_menu.right_click_export: An unknown error occurred."
                    self.msg_box.insert(END, message + "\n")
                    self.msg_box.see("end")
                    error_log(message)
        rmenu.add_command(label="Query media",command=lambda : self.double_click(event),compound="left")
        rmenu.add_command(label="Delete selected",command=lambda : self.del_selected(event,mode="select"),compound="left")
        rmenu.add_command(label="Delete all", command=lambda: self.del_selected(event, mode="all"),compound="left")
        rmenu.add_command(label="Export result", command=expire_export, compound="left")
        rmenu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")

    def double_click(self,event):
        item = self.tree.identify('item', event.x, event.y)
        try:
            if self.tree.item(item)["text"]:
                try:
                    vision.house_id_event([self.tree.item(item)["text"]])
                except:
                    pass
        except AttributeError:
            pass

    def _tree_sort_column(self,tv, col, reverse):
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)

        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        tv.heading(col, command=lambda: self._tree_sort_column(tv, col, not reverse))

class PopUpMessage:
    count = 0

    def __init__(self,nb_index,title,message,action):
        PopUpMessage.count+=1
        self.can = Toplevel()
        self.nb_index = nb_index
        self.can.attributes("-alpha", 0)
        self.can.attributes('-topmost', 'true')
        self.can.geometry("275x67+{}+{}".format(round(root.winfo_x() + 929), round(root.winfo_y() + 563)))
        self.can.overrideredirect(True)
        self.default_color = vision.default_color
        self.font_color = vision.font_color
        delete = PhotoImage(file="icons/delete.png")
        self.can_frame = Frame(self.can,highlightbackground=settings.default_color,
                               highlightcolor="dark green", highlightthickness=1,bg=self.default_color)
        self.can_frame.pack(fill=X, expand=True)
        del_bot = Button(self.can, image=delete, command=self.can.destroy, relief="flat", bg=self.default_color)
        del_bot.img = delete
        Label(self.can_frame, image=vision.check_icon,bg=self.default_color).grid(row=0,column=0)
        Label(self.can_frame,text=title, font='Helvetica 10 bold',bg=self.default_color,
              fg=self.font_color).grid(row=0,column=1,sticky="w")
        self.body_frame = Frame(self.can,highlightbackground=self.default_color,
                                highlightcolor="dark green", highlightthickness=1,bg="white")
        Label(self.body_frame, text=message, font='Helvetica 8',bg="white").grid(row=1,column=1,sticky="w")
        Label(self.body_frame, text=action, font='Helvetica 7',bg="white",fg="blue2").grid(row=2, column=1, sticky="w")
        self.body_frame.pack(fill=BOTH,expand=True)
        del_bot.place(relx=0.91,rely=0.02)
        root.lift()
        root.attributes('-topmost', True)
        root.after_idle(root.attributes, '-topmost', False)
        self.fade_in()

    def jump_tab(self,event):
        vision.nb.select(self.nb_index)
        self.fade_away()

    def refresh_data(self,event):
        PopUpMessage.count = 0
        vision.start_thread(vision.refresh_inventory)
        self.fade_away()
        vision.import_but.config(relief="flat",state="normal")

    def fade_in(self):
        try:
            alpha = self.can.attributes("-alpha")
            if alpha < 1:
                alpha += .07
                self.can.attributes("-alpha", alpha)
                self.can.update_idletasks()
                root.after(100, self.fade_in)
            #else:
            #    root.after(60000, self.fade_away)
        except TclError:
            pass

    def fade_away(self):
        try:
            alpha = self.can.attributes("-alpha")
            if alpha > 0:
                alpha -= .07
                self.can.attributes("-alpha", alpha)
                root.after(100, self.fade_away)
            else:
                self.can.destroy()
        except TclError:
            pass
        self.count-=1

class ttkTimer(Thread):
    def __init__(self, callback, tick):
        Thread.__init__(self)
        self.callback = callback
        self.stopFlag = Event()
        self.tick = tick
        self.iters = 0

    def run(self):
        try:
            while not self.stopFlag.wait(self.tick):
                self.iters += 1
                self.callback()
        except RuntimeError as e:
            error_log("Function ttkTimer.run: "+str(e))

    def stop(self):
        self.stopFlag.set()

    def get(self):
        return self.iters

class Player:
    def __init__(self, url, title=None):
        self.parent = Toplevel()
        self.parent.grab_set()
        self.parent.protocol("WM_DELETE_WINDOW",self.on_quit)
        self.parent.overrideredirect(True)
        self.top_frame = Frame(self.parent,background="gray28",cursor="size")
        self.top_frame.pack(fill=BOTH)
        self.assv = IntVar()
        self.assv.set(0)
        Label(self.top_frame,text=" Choose subtitle ",background="gray28",foreground="white",cursor="arrow").grid(row=0,column=0)
        self.url = url
        self.ass_list= []
        if not title: #title == None:
            title = "Subtitle Override"
        self.parent.title(title)
        self.xpos, self.ypos = 500, 200
        self.parent.geometry(f"640x480+{self.xpos}+{self.ypos}")
        self.parent.config(highlightbackground="dark green", highlightcolor="dark green", highlightthickness=1)

        delete = PhotoImage(file="icons/exit2.png")
        exit_but = Button(self.parent,image=delete,command=self.on_quit,relief="flat",background="gray28")
        exit_but.place(relx=0.96,rely=0)
        exit_but.img = delete

        # The second panel holds controls
        self.player = None
        self.videopanel = Frame(self.parent,highlightbackground="dark green", highlightcolor="dark green", highlightthickness=1)
        self.canvas = Canvas(self.videopanel).pack(fill=BOTH,expand=1)
        self.videopanel.pack(fill=BOTH,expand=1)
        self.image_holder = [PhotoImage(file="icons/play.png"),PhotoImage(file="icons/pause.png"),
                             PhotoImage(file="icons/stop2.png"),PhotoImage(file="icons/volume.png"),
                             PhotoImage(file="icons/muted.png")]
        ctrlpanel = ttk.Frame(self.parent)
        self.current_time = ttk.Label(ctrlpanel, text="Current Time - 00:00",relief="groove",anchor=CENTER)
        play = ttk.Button(ctrlpanel, text="Play",image=self.image_holder[0],compound="left", command=self.OnPlay,cursor="hand2")
        pause  = ttk.Button(ctrlpanel, text="Pause",image=self.image_holder[1],compound="left", command=self.OnPause,cursor="hand2")
        stop   = ttk.Button(ctrlpanel, text="Stop",image=self.image_holder[2],compound="left", command=self.OnStop,cursor="hand2")
        self.volume = ttk.Button(ctrlpanel, text="Mute",image=self.image_holder[3],compound="left", command=self.OnToggleVolume,cursor="hand2")

        self.current_time.pack(side=LEFT,padx=10,ipadx=10,ipady=6)
        play.pack(side=LEFT)
        pause.pack(side=LEFT)
        stop.pack(side=LEFT)
        self.volume.pack(side=LEFT)
        self.volume_var = IntVar()
        self.volslider = Scale(ctrlpanel, variable=self.volume_var, command=self.volume_sel,
                from_=0, to=100, orient=HORIZONTAL, length=100,cursor="hand2")
        self.volslider.pack(side=LEFT)
        ctrlpanel.pack(side=BOTTOM)

        ctrlpanel2 = ttk.Frame(self.parent)
        self.scale_var = DoubleVar()
        self.timeslider_last_val = ""
        self.timeslider = Scale(ctrlpanel2, variable=self.scale_var, command=self.scale_sel,
                from_=0, to=1000, orient=HORIZONTAL, length=500,showvalue=0,cursor="hand2")
        self.timeslider.pack(side=BOTTOM, fill=X,expand=1)
        self.timeslider_last_update = time.time()
        ctrlpanel2.pack(side=BOTTOM,fill=X)

        # VLC player controls
        self.Instance = vlc.Instance(['--video-on-top'])
        self.player = self.Instance.media_player_new()

        # below is a test, now use the File->Open file menu
        #media = self.Instance.media_new('output.mp4')
        #self.player.set_media(media)
        #self.player.play() # hit the player button
        #self.player.video_set_deinterlace(str_to_bytes('yadif'))

        self.timer = ttkTimer(self.OnTimer, 1.0)
        self.timer.start()
        self.parent.update()
        self.parent.withdraw()
        self.parent.grab_release()
        for i in ("<Left>","<Right>","<Up>","<Down>"):
            self.parent.bind(i,self.move)
        #self.player.set_hwnd(self.GetHandle()) # for windows, OnOpen does does this

    def move(self,event):
        height, rest = self.parent.winfo_geometry().split("x")
        width, xloc, yloc = rest.split("+")
        xloc,yloc = int(xloc), int(yloc)
        try:
            if event.keysym == "Left":
                xloc -= 20 #self.xpos -=20
            elif event.keysym == "Right":
                xloc += 20#self.xpos +=20
            elif event.keysym == "Up":
                yloc -= 20 #self.ypos -=20
            elif event.keysym == "Down":
                yloc += 20#self.ypos +=20
            self.parent.geometry(f"640x480+{xloc}+{yloc}")
        except:
            pass

    def load_url(self,url,ass_list):
        self.url = url
        self.ass_list = ass_list
        self.parent.deiconify()
        self.parent.grab_set()
        self.OnOpen()
        _start_new_thread(self.temp_save_ass,("Thread",1))

    def temp_save_ass(self,a,b):
        self.radio_list = []
        for i, file in enumerate(self.ass_list):
            try:
                fp = urllib.request.urlopen(file)
                sub_read = fp.read().decode("utf8")
                with open(file.split("/")[-1], "w", encoding="utf-8") as f:
                    f.write(sub_read)
            except urllib.error as e:
                error_log("Function Player.temp_save_ass: "+str(e))
            self.radio_but = Radiobutton(self.top_frame, padx=5, pady=1, bd=2, relief="groove", selectcolor="purple3",
                              bg="grey", fg="white",command=self.set_sub,cursor="hand2",
                              text=file.split("/")[-1].split(".")[0][-2:], variable=self.assv, value=i, indicatoron=0)
            self.radio_but.grid(row=0,column=i+1)
            self.radio_list.append(self.radio_but)
        try:
            self.set_sub()
        except:
            pass
        return

    def set_sub(self):
        if self.ass_list:
            self.player.video_set_subtitle_file(self.ass_list[self.assv.get()].split("/")[-1])

    def OnOpen(self):
        self.OnStop()
        self.Media = self.Instance.media_new(self.url)
        self.player.set_media(self.Media)

        if platform.system() == 'Windows':
            self.player.set_hwnd(self.GetHandle())
        else:
            self.player.set_xwindow(self.GetHandle()) # this line messes up windows
        # FIXME: this should be made cross-platform
        self.OnPlay()
        # set the volume slider to the current volume
        #self.volslider.SetValue(self.player.audio_get_volume() / 2)
        self.volslider.set(self.player.audio_get_volume())
        self.volume_var.set(100)

    def OnPlay(self):
        if not self.player.get_media():
            self.OnOpen()
        else:
            if self.player.play() == -1:
                self.errorDialog("Unable to play.")

    def GetHandle(self):
        return self.videopanel.winfo_id()

    def OnPause(self):
        self.player.pause()

    def OnStop(self):
        self.player.stop()
        self.player.video_set_subtitle_file("")
        self.timeslider.set(0)

    def OnTimer(self):
        """Update the time slider according to the current movie time.
        """
        if self.player == None:
            return
        # since the self.player.get_length can change while playing,
        # re-set the timeslider to the correct range.
        length = self.player.get_length()
        dbl = length * 0.001
        self.timeslider.config(to=dbl)

        # update the time on the slider
        tyme = self.player.get_time()
        if tyme == -1:
            tyme = 0
        dbl = tyme * 0.001
        self.timeslider_last_val = ("%.0f" % dbl) + ".0"
        self.current_time.config(text=f"Current Time - {int(dbl/60):02}:{int(dbl%60):02}")
        # don't want to programmatically change slider while user is messing with it.
        # wait 2 seconds after user lets go of slider
        if time.time() > (self.timeslider_last_update + 2.0):
            self.timeslider.set(dbl)

    def scale_sel(self, evt):
        if self.player == None:
            return
        nval = self.scale_var.get()
        sval = str(nval)
        if self.timeslider_last_val != sval:
            self.timeslider_last_update = time.time()
            mval = "%.0f" % (nval * 1000)
            self.player.set_time(int(mval)) # expects milliseconds

    def volume_sel(self, evt):
        if self.player == None:
            return
        volume = self.volume_var.get()
        if volume > 100:
            volume = 100
        if self.player.audio_set_volume(volume) == -1: pass
            #self.errorDialog("Failed to set volume")

    def OnToggleVolume(self):
        """Mute/Unmute according to the audio button."""
        is_mute = self.player.audio_get_mute()
        if is_mute:
            self.volume.config(text="Mute",image=self.image_holder[3])
        else:
            self.volume.config(text="Muted",image=self.image_holder[4])
        self.player.audio_set_mute(not is_mute)
        # update the volume slider;
        # since vlc volume range is in [0, 200],
        # and our volume slider has range [0, 100], just divide by 2.
        self.volume_var.set(self.player.audio_get_volume())

    def OnSetVolume(self):
        """Set the volume according to the volume slider.
        """
        volume = self.volume_var.get()
        # vlc.MediaPlayer.audio_set_volume returns 0 if success, -1 otherwise
        if volume > 100:
            volume = 100
        if self.player.audio_set_volume(volume) == -1:
            self.errorDialog("Failed to set volume")

    def errorDialog(self, errormessage):
        """Display a simple error dialog"""
        messagebox.showerror('Error', errormessage)

    def on_quit(self):
        self.OnStop()
        #self.parent.withdraw()
        #self.parent.grab_release()
        for but in self.radio_list:
            try:
                but.destroy()
            except TclError:
                pass
        for i in self.ass_list:
            try:
                os.remove(i.split("/")[-1])
            except (FileNotFoundError,PermissionError):
                pass
            except TclError:
                pass
        self.timer.stop()
        self.parent.destroy()

    def on_exit(self):
        self.parent.quit()
        self.parent.destroy()
        self.timer.stop()

class Encryption:

    @staticmethod
    def encode(key, clear):
        enc = []
        for i in range(len(clear)):
            key_c = key[i % len(key)]
            enc_c = chr((ord(clear[i]) + ord(key_c)) % 256)
            enc.append(enc_c)
        return base64.urlsafe_b64encode("".join(enc).encode()).decode()

    @staticmethod
    def decode(key, enc):
        dec = []
        enc = base64.urlsafe_b64decode(enc).decode()
        for i in range(len(enc)):
            key_c = key[i % len(key)]
            dec_c = chr((256 + ord(enc[i]) - ord(key_c)) % 256)
            dec.append(dec_c)
        return "".join(dec)

class LoginScreen:
    def __init__(self):
        self.log_win = Toplevel()
        self.log_win.geometry("320x311")
        self.log_win.grab_set()
        self.auto_login = False
        self.top_frame = Frame(self.log_win)
        self.top_frame.pack(fill=X)
        self.top_frame.config(background=vision.default_color)
        self.color = "white"
        self.log_win.config(background=self.color)
        self.bottom_frame = Frame(self.log_win,background=self.color,highlightbackground="green", highlightcolor="green", highlightthickness=1)
        self.bottom_frame.pack(fill=BOTH)
        self.bg_label = Label(self.top_frame,text="Sign In", background=vision.default_color,foreground="white",font="Calibri 22")
        self.bg_label.pack(side=LEFT,padx=10)
        self.log_win.overrideredirect(True)
        self.log_win.geometry(self.center(self.log_win))
        self.exit_img = PhotoImage(file="icons/delete.png")
        self.exit = Label(self.log_win,image=self.exit_img,background=vision.default_color)
        self.exit.img = self.exit_img
        self.exit.place(relx=0.93,rely=0.007)
        self.exit.bind("<Button-1>",lambda e: self.log_win.destroy())
        self.create_bottom_frame(self.bottom_frame)

    def create_bottom_frame(self,parent):
        try:
            with open("login_info.json","r") as f:
                self.login_info = json.load(f)
        except:
            self.login_info = {"autologin": 0,"username":"","password":""}
        self.login_message = Label(parent,text="",foreground="red",background=self.color,font="Calibri 12")
        self.login_message.grid(row=0, column=0,sticky=W,padx=10)
        self.user_label = Label(parent, text="USERNAME",font="Calibri 14",background=self.color)
        self.user_label.grid(row=1, column=0,sticky=W,padx=10,pady=1)
        self.username = Entry(parent, width=29,font="Calibri 14",borderwidth=2)
        self.username.grid(row=2, column=0,padx=11,pady=1,sticky=W,ipady=2)
        self.username.insert(0,Encryption.decode("Infinitywar@2018",self.login_info.get("username")))
        self.username.focus_set()
        self.pass_label = Label(parent,text="PASSWORD",font="Calibri 14",background=self.color)
        self.pass_label.grid(row=3,column=0,sticky=W,padx=10,pady=1)
        self.password = Entry(parent,width=29,show="*",font="Calibri 14",borderwidth=2)
        self.password.grid(row=4,column=0,padx=11,pady=1,sticky=W,ipady=2)
        self.password.insert(0,Encryption.decode("Infinitywar@2018",self.login_info.get("password")))
        self.password.bind('<Lock-KeyPress>', lambda e: self.caps_lock(e))
        self.eye_img = PhotoImage(file="icons/eye.png")
        self.eye_label = Label(parent, image=self.eye_img,background="white")
        self.eye_label.place(relx=0.89, rely=0.49)
        self.eye_label.img = self.eye_img
        self.eye_label.bind("<Button-1>",self.show_password)
        self.var = IntVar()
        self.var.set(self.login_info.get("autologin"))
        self.var_check = Checkbutton(parent, variable=self.var, borderwidth=2,text="Keep me signed in",background=self.color,
                                     font="Calibri 12",command="", anchor=E,justify=RIGHT,selectcolor="white")
        self.var_check.grid(row=5,column=0,sticky=E,padx=3)
        self.forget_pass = Label(parent,text="Forgot your password?",font="Calibri 11",cursor="hand2",foreground="DeepSkyBlue3",background=self.color)
        f = font.Font(self.forget_pass, self.forget_pass.cget("font"))
        f.configure(underline=True)
        self.forget_pass.config(font=f)
        self.forget_pass.grid(row=6,column=0,sticky=E,padx=5)
        self.log_but = Button(parent,text="SIGN IN",relief="groove",background=self.color,foreground="forest green",font="Calibri 14",highlightbackground="forest green",
                           highlightcolor="forest green",width=9)
        self.log_but.grid(row=7,column=0,pady=12,sticky=E,padx=7)

    def show_password(self,event):
        if self.password.cget("show"):
            self.password.config(show="")
        else:
            self.password.config(show="*")

    def callback(event):
        webbrowser.open_new(f"http://mpx.theplatform.com/#/request-reset-password?username={self.username.get()}")

    def caps_lock(self,event):
        if event.keysym == "Caps_Lock":
            print ("Caps lock off")
        else:
            print (event)

    def center(self,win):
        win.update_idletasks()
        width = win.winfo_width()
        height = win.winfo_height()
        x = (root.winfo_screenwidth() // 2) - (width // 2)
        y = (root.winfo_screenheight() // 2) - (height // 2)
        return '{}x{}+{}+{}'.format(width, height, x, y-80)

    def save_login(self):
        self.auto_login = True

    def _quit_login(self):
        self.log_win.destroy()

class SetPref:

    def clear_canvas(self):
        try:
            self.hover_info.destroy()
        except (TclError,AttributeError):
            pass

    def draw_canvas(self,val,iid,x,y):
        self.hover_info = Canvas(root)
        self.value = val
        self.iid = iid
        self.x,self.y = x,y
        hover_frame = Frame(self.hover_info, bg="PaleGreen4", highlightbackground="green", highlightcolor="green",
                            highlightthickness=1)
        hover_frame.pack(fill=X)
        Label(hover_frame, text="Set version preference", bg="PaleGreen4", fg="white").pack(side=LEFT,
                                                                                                          padx=1)
        lower_frame = Frame(self.hover_info, bg=vision.default_bg, highlightbackground="green", highlightcolor="green",
                            highlightthickness=1)
        lower_frame.pack(fill=BOTH)
        Label(lower_frame, text=self.value[0], bg=vision.default_bg,fg=vision.default_font, anchor='w').grid(row=0, column=0, sticky="w",columnspan=4)
        for num, text in enumerate(("Season:","House No:","Box No:"),1):
            Label(lower_frame, text=text, bg=vision.default_bg,fg=vision.default_font, width=9, anchor='w').grid(row=num, column=0, sticky="w")
        for num, text in enumerate(("Episode:", "Version:", "On-Air ID"),1):
            Label(lower_frame, text=text, bg=vision.default_bg,fg=vision.default_font, width=8, anchor='w').grid(row=num, column=2, sticky="w")
        Label(lower_frame, text=self.value[1] if self.value[1] else "N/A", #Season
              bg=vision.default_bg,fg=vision.default_font, width=10, anchor='w').grid(row=1, column=1, sticky="w")
        Label(lower_frame, text=self.value[2] if self.value[2] else "N/A", #ep
              bg=vision.default_bg,fg=vision.default_font, width=10, anchor='w').grid(row=1, column=3, sticky="w")
        self.ver_combo = ttk.Combobox(lower_frame, state="readonly",width=13)
        result = vision.df[vision.df.ID.eq(int(self.value[21]))]
        self.box_combo = ttk.Combobox(lower_frame, state="readonly", width=13)
        self.box_combo.grid(row=3, column=1, sticky="w")
        self.box_combo["values"] = result["Box Number"].tolist()
        self.box_combo.set(self.value[14])
        Label(lower_frame, text=self.value[15] if self.value[15] else "", # house ID
              bg=vision.default_bg,fg=vision.default_font, width=10, anchor='w').grid(row=2, column=1, sticky="w")
        Label(lower_frame, text=self.value[13] if self.value[15] else "", # version
              bg=vision.default_bg, fg=vision.default_font, width=10, anchor='w').grid(row=2, column=3, sticky="w")
        Label(lower_frame, text=self.value[21] if self.value[21] else "", # OnAir ID
              bg=vision.default_bg,fg=vision.default_font, width=10, anchor='w').grid(row=3, column=3, sticky="w")
        self.confirm_but = Button(lower_frame,text="Confirm",compound="left",image=vision.check_icon,command=self.save_pref,relief="flat",bg=vision.default_bg,fg=vision.default_font)
        self.confirm_but.grid(row=4,column=2,columnspan=2)
        cancel_icon = PhotoImage(file="icons/delete.png")
        self.cancel_but = Button(lower_frame, text="Cancel",compound="left",image=cancel_icon,command=self.delete_box, relief="flat", bg=vision.default_bg,fg=vision.default_font)
        self.cancel_but.grid(row=4, column=0, columnspan=2)
        self.cancel_but.img = cancel_icon
        self.hover_info.place(relx=0.6,y=self.y-20) #self.hover_info.place(relx=0.001, rely=0.038) (x=self.x+25,y=self.y+20)

    def save_pref(self):
        result = vision.df.loc[vision.df["Box Number"].eq(self.box_combo.get())]
        try:
            settings.pref.remove(vision.tree.item(self.iid)["values"][14])
        except (KeyError, Exception):
            pass
        settings.pref.add(self.box_combo.get())
        new_value = list(self.value)
        new_value[13] = result.iat[0,13]
        new_value[14] = self.box_combo.get()
        new_value[15] = result.iat[0,15]
        vision.tree.item(self.iid,value=new_value)
        try:
            with open("settings.pkl", "wb") as f:
                pickle.dump(settings, f)
        except PermissionError:
            pass
        self.hover_info.destroy()
        vision.show_pref = False

    def delete_box(self):
        self.hover_info.destroy()
        vision.show_pref = False

def _quit():
    print ("I'm quitting")
    root.quit()
    os._exit(1)

class PurgeSubtitle:
    def __init__(self,master):
        self.master = master
        self.hidden_dict = {}
        self.top_frame = Frame(master,bg=settings.default_color,highlightbackground="green",
                               highlightcolor="green", highlightthickness=1)
        self.top_frame.pack(fill=X)
        Label(self.top_frame,text="Input House/Box number(s) for subtitle purge: ",bg=settings.default_color,
              fg=settings.font_color).grid(row=0,column=0,sticky="w",padx=5)
        self.entry = Entry(self.top_frame,width=112,borderwidth=2,fg="grey",font="TkDefaultFont 9 italic")
        self.entry.grid(row=0,column=1,sticky="w")
        self.message = " Enter Box or House number. Can be separated by ',' or copy & paste from excel/quip."
        self.entry.insert(0, self.message)
        self.entry.bind('<FocusIn>', self.entry_click)
        self.entry.bind('<FocusOut>', lambda e: self.entry_focusout(e, self.message))
        self.submit_icon = PhotoImage(file="icons/submit.png")
        self.submit = Button(self.top_frame,text="Submit",image=self.submit_icon,
                             relief="groove",compound="left",width=65,
                             bg=settings.default_color,fg=settings.font_color,
                             command=lambda: _start_new_thread(self.get_url,("Thread",1)))
        self.submit.grid(row=0,column=2,sticky="w",padx=5)
        self.quip_icon = PhotoImage(file="icons/quip.png").subsample(32)
        self.quip = Button(self.top_frame,text="Get from Quip",image=self.quip_icon,relief="groove",compound="left",
                             bg=settings.default_color,fg=settings.font_color,command="",state="disabled")
        self.quip.grid(row=0,column=3,sticky="w")
        self.purge_icon = PhotoImage(file="icons/recycle.png")
        self.purge = Button(self.top_frame,text="Purge",image=self.purge_icon,
                            relief="groove",compound="left",width=65,
                            bg=settings.default_color,fg=settings.font_color,
                            command=lambda: _start_new_thread(self.purge_subtitle,("Thread",1)))
        self.purge.grid(row=0, column=4,sticky="w",padx=5)
        self.middle_frame = Frame(master,bg="white",highlightbackground="green", highlightcolor="green", highlightthickness=1)
        self.middle_frame.pack(fill=BOTH)
        self.image_holder = [self.submit_icon,self.quip_icon,self.purge_icon]
        self.entry.bind("<Return>",self.return_but)
        self.create_tree()
        self.top_frame.bind("<Visibility>", lambda e: self.tree.focus_set())

    def return_but(self,event):
        _start_new_thread(self.get_url, ("Thread", 1))

    def entry_click(self,event):
        if event.widget["foreground"] == "grey":
            event.widget.delete(0, "end")
            event.widget.insert(0, "")
            event.widget.configure(fg="black",font="TkDefaultFont 9")

    def entry_focusout(self, event, msg):
        if not event.widget.get():
            event.widget.configure(fg="grey",font="TkDefaultFont 9 italic")
            event.widget.insert(0, msg)

    def get_url(self,a,b):
        self.tok = vision.tok
        if not self.entry.get() or not self.tok:
            return
        self.submit.config(state="disabled",relief="sunken")
        self.purge.config(state="disabled",relief="sunken")
        self.middle_frame.config(cursor="wait")
        if "," in self.entry.get():
            result = self.entry.get().replace(" ","").split(",")
        else:
            result = self.entry.get().replace(" ","").split("\n")
        result = set(i.upper() for i in result)
        house_id_combo = []
        for i in result:
            if len(i) == 8 and i.isnumeric():
                house_id_combo.append(i)
            elif i in vision.df["Box Number"].values:
                house_id_combo.append(vision.df.loc[vision.df["Box Number"].eq(i), "House Number"].iat[0])
        self.entry.delete(0,END)
        self.entry.insert(0,",".join(house_id_combo))
        self.entry.config(state="disabled")
        if not house_id_combo:
            self.submit.config(state="normal", relief="groove")
            self.purge.config(state="normal", relief="groove")
            self.entry.config(state="normal")
            self.middle_frame.config(cursor="arrow")
            return
        link = "http://data.media.theplatform.com/media/data/Media?schema=1.10.0&searchSchema=1.0.0&form=cjson&pretty=true&fields=id,added,categories,:season,:episodeNumber,:houseID,title,:channel_id,:series-Title,content&byCustomValue=%7BhouseID%7D%7B" + "%7C".join(
            house_id_combo) + "%7D&token=" + self.tok + "&account=Fox%20Play%20Asia"
        try:
            with urllib.request.urlopen(link) as url:
                data = json.loads(url.read().decode())
                for item in data['entries']:
                    channel_id = ", ".join(
                        item.get("pl3$channel_id", item.get("pl2$channel_id", item.get("pl1$channel_id", ""))))
                    title = item.get("title", "")
                    series_title = item.get("pl3$series-Title",
                                            item.get("pl2$series-Title", item.get("pl1$series-Title", "")))
                    house_id = item.get("pl1$houseID", "")
                    ep = str(item.get("pl1$episodeNumber", ""))
                    season = str(item.get("pl1$season", ""))
                    subtitle = {}
                    sub_date = []
                    for media_item in item["content"]:
                        if media_item["url"].upper().endswith(".DFXP"):
                            date = datetime.fromtimestamp(media_item.get("added", 1000) / 1000).strftime(
                                '%Y-%m-%d %H:%M')
                            subtitle[media_item.get("title")] = {"streamingUrl": media_item.get("streamingUrl", ""),
                                                                 "added": date,"media_id":str(media_item.get("id", "").split("/")[-1]),
                                                                 "asset": ', '.join(media_item.get("assetTypes", "")),}
                            sub_date.append(date)
                    if subtitle:
                        for k, v in subtitle.items():
                            try:
                                self.tree.insert("", END, text=house_id, values=(v.get("media_id"),
                                k, v.get("asset"), series_title if series_title else title, season, ep, channel_id, v.get("streamingUrl"), v.get("added"),""))
                            except:
                                pass
        except KeyError:
            vision.mpx_connect_console()
        finally:
            self.submit.config(state="normal", relief="groove")
            self.purge.config(state="normal", relief="groove")
            self.entry.config(state="normal")
            self.middle_frame.config(cursor="arrow")

    def create_tree(self):
        self.tree = ttk.Treeview(self.middle_frame, height=27, selectmode='extended')
        self.tree.pack(side='left')
        vsb = ttk.Scrollbar(self.middle_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side='right', fill='y')
        self.tree.configure(yscrollcommand=vsb.set)
        header = ("Media ID","Subtitle","Asset Type","Series", "Season", "Ep", "Channel", "URL", "Added Date","Action")
        width = (110,170,150,300,50, 50, 150, 300, 120,130)
        self.tree["columns"] = header
        self.tree.column("#0", width=80, anchor="w")
        self.tree.heading("#0", text="House ID", anchor="w")
        for i in range(len(header)):
            self.tree.column(header[i], width=width[i], anchor="w")
            self.tree.heading(header[i], text=header[i], anchor='w')
            self.tree["displaycolumns"] = (
            "Subtitle","Asset Type","Series", "Season", "Ep", "Channel","Added Date","Action")
        self.tree.tag_configure("problem", background="red", foreground="white")
        self.tree.bind("<Button-3>",self.popup)
        self.tree.bind("<Double-1>", self.double_click)
        for col in self.tree["columns"]:
            self.tree.heading(col, text=col, command=lambda _col=col: self.treeview_sort_column(self.tree, _col, False))

    def purge_subtitle(self,a,b):
        if self.tree.selection():
            self.middle_frame.config(cursor="starting")
            for item in self.tree.selection():
                current_value = self.tree.item(item)["values"]
                current_value[9] = "Requesting..."
                self.tree.item(item, values=current_value)
                try:
                    link = "https://api.asia.fox.com/akamai/purge?url="+str(current_value[7])
                    with urllib.request.urlopen(link) as url:
                        result = json.loads(url.read().decode())
                        current_value[9] = result.get("detail")
                    self.tree.item(item,values=current_value)
                except:
                    current_value[9] = "Request failed"
                    self.tree.item(item, values=current_value,tags=("problem",))
        self.middle_frame.config(cursor="arrow")

    def popup(self,event):
        def right_click_delete():
            if self.tree.selection():
                for item in self.tree.selection():
                    self.tree.delete(item)
        def right_click_delete_all():
            self.tree.delete(*self.tree.get_children())
        def right_click_show_sub(string):
            right_click_show_all()
            try:
                for num, item_id in enumerate(self.tree.get_children()):
                    if string != self.tree.item(item_id)["values"][2]:
                        self.hidden_dict[item_id] = num
                        self.tree.detach(item_id)
            except:
                pass
        def right_click_show_all():
            try:
                for item_id, i_r in self.hidden_dict.items():
                    self.tree.reattach(item_id, '', i_r)
            except:
                pass
        def house_event():
            item = self.tree.identify('item', event.x, event.y)
            if self.tree.item(item)["text"]:
                vision.house_id_event([self.tree.item(item)["text"], ])
        rmenu = Menu(None, tearoff=0, takefocus=0)
        rmenu.add_command(label="Purge", command=lambda: _start_new_thread(self.purge_subtitle,("Thread",1)))
        rmenu.add_command(label="Clear selected",command=right_click_delete)
        rmenu.add_command(label="Clear all", command=right_click_delete_all)
        rmenu.add_command(label="Show all", command=right_click_show_all)
        rmenu.add_command(label="Media info", command=house_event)
        sort_menu = Menu(None, tearoff=0, takefocus=0)
        rmenu.add_cascade(label="Sort by", menu=sort_menu)
        sort_commands = [
            ('Sub Chinese', lambda: right_click_show_sub("Sub Chinese")),
            ('Sub Chinese Simplified', lambda: right_click_show_sub("Sub Chinese Simplified")),
            ('Sub English', lambda: right_click_show_sub("Sub English")),
            ('Sub Thai', lambda: right_click_show_sub("Sub Thai")),
            ('Sub Malay', lambda: right_click_show_sub("Sub Malay")),
            ('Sub Indonesian', lambda: right_click_show_sub("Sub Indonesian")),
            ('Sub Chinese and English', lambda : right_click_show_sub("Sub Chinese and English")),
            ('Sub Arabic', lambda: right_click_show_sub("Sub Arabic")),
            ('Sub Korean', lambda: right_click_show_sub("Sub Korean")),
        ]
        for (txt, cmd) in sort_commands:
            sort_menu.add_command(label=txt, command=cmd, compound="left")
        rmenu.tk_popup(event.x_root + 40, event.y_root + 10, entry="0")

    def double_click(self,event):
        try:
            iid = self.tree.identify_row(event.y)
            url = self.tree.item(iid)["values"][7]
            SubCheck(url)
        except (IndexError,AttributeError):
            pass
        except:
            pass

    def treeview_sort_column(self,tv, col, reverse):
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        l.sort(reverse=reverse)

        # rearrange items in sorted positions
        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        # reverse sort next time
        tv.heading(col, command=lambda: self.treeview_sort_column(tv, col, not reverse))

class VersionValidator:
    def __init__(self):
        self.app_expire = "2019-09-30 23:59:59"

    def check_date(self):
        try:
            res = urllib.request.urlopen('http://just-the-time.appspot.com/')
            result = res.read().strip().decode('utf-8')
            if self.app_expire > result:
                return True, "Valid"
            else:
                return False, "App version has expired.\nPlease acquire an updated version."
        except:
            return False, "Cannot validate app version."

if __name__ == "__main__":
    root = Tk()
    root.withdraw()

    w = 1223  # 1143
    h = 632
    ws = root.winfo_screenwidth()
    hs = root.winfo_screenheight()
    x = (ws / 2) - (w / 2)
    y = (hs / 2) - (h / 2) - 100
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))  # w = 1223 h = 632
    root.iconbitmap('icons/vision3.ico')
    root.resizable(width=False, height=False)
    root.title(app_version)
    series_name, scm_name, movies_name, factual_name, scc_name = set(), set(), set(), set(), set()
    trailer_name = set()

    try:
        with open("settings.pkl", "rb") as f:
            settings = pickle.load(f)
    except FileNotFoundError:
        settings = Settings()

    full_var = [IntVar() for i in range(10)]

    sys_check = SystemCheck()

    vision = MainGui(root)

    root.bind("<Control-f>", vision.filter_focus)

    set_pref = SetPref()

    vision.submenu.add_command(label="Exit", command=_quit)

    root.protocol("WM_DELETE_WINDOW", _quit)

    root.mainloop()